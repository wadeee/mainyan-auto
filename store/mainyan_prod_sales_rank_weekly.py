"""
麦安研产品销售周度排行 - 自动化脚本
=====================================
依赖：pip install playwright && playwright install chromium

自动登录 Pospal 后台，导出麦安研各门店的商品销售和商品报损周度统计数据。

用法：
    python mainyan_prod_sales_rank_weekly.py                    # 导出本周数据（周一到周日）
    python mainyan_prod_sales_rank_weekly.py --weeks -1         # 导出上周数据
    python mainyan_prod_sales_rank_weekly.py --date 2026.06.03  # 根据日期推断所在周
    python mainyan_prod_sales_rank_weekly.py --headless          # 无头模式（不显示浏览器）
"""

import argparse
import copy
import json
import logging
import logging.handlers
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

# ─── 日志配置 ───────────────────────────────────────────────────────────────────
LOG_DIR = Path(__file__).resolve().parent / "log"
LOG_DIR.mkdir(exist_ok=True)

_file_handler = logging.handlers.TimedRotatingFileHandler(
    LOG_DIR / "mainyan_prod_sales_rank_weekly.log",
    when="midnight",
    backupCount=30,
    encoding="utf-8",
)
_file_handler.suffix = "%Y-%m-%d"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[_file_handler, logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# ─── 配置区（按需修改）─────────────────────────────────────────────────────────
# 账号信息同 factory_delivery_tsj_weekly.py
ACCOUNT = "huomimayzb"
WORKER_ID = "M006"
PASSWORD = "tusijia88"

LOGIN_URL = "https://beta69.pospal.cn/"
PRODUCT_SALE_URL = "https://beta69.pospal.cn/ReportV2/ProductSale"
DISCARD_PRODUCT_URL = "https://beta69.pospal.cn/Inventory/DiscardProductCount"

OUTPUT_DIR = Path(__file__).resolve().parent / "麦安研产品销售周度排行"
TEMPLATE_FILE = Path(__file__).resolve().parent / "麦安研产品销售周度排行_格式化模板.xlsx"

STORES = [
    {"full": "3 - 麦安研（东站宝泰店）", "short": "宝泰店"},
    {"full": "5 - 麦安研（顺德龙江店）", "short": "龙江店"},
    {"full": "2 - 麦安研（顺德杏坛店）", "short": "杏坛店"},
]

SALE_CATEGORIES = ["热销酥类", "现烤面包", "包装面包", "裱花自制"]

CATEGORY_ORDER = ["现烤面包", "热销酥类", "裱花自制", "包装面包"]

HEADER_ROW = 2
DATA_START_ROW = 3
SAMPLE_ROWS = {
    "现烤面包": 3,
    "热销酥类": 5,
    "裱花自制": 7,
    "包装面包": 9,
}


# ─── 格式化合并函数 ──────────────────────────────────────────────────────────


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def _to_num(val):
    if val is None:
        return None
    try:
        f = float(val)
        if f == 0:
            return None
        if f == int(f):
            return int(f)
        return f
    except (ValueError, TypeError):
        return val


def read_sale_data(data_file: Path):
    """读取商品销售周度统计 xlsx，返回 list[dict]。

    每行包含：商品名称(A), 商品条码(B), 单位(E), 商品分类(F), 销售数量(H), 商品总售价(I)
    """
    from openpyxl import load_workbook as _lwb
    wb = _lwb(data_file, data_only=True)
    ws = wb.active

    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val:
            headers[str(val).strip()] = c

    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=headers["商品名称"]).value
        if name is None:
            continue
        name = str(name).strip()
        if name in ("-", "合计", "总计", ""):
            continue
        category = ws.cell(row=r, column=headers["商品分类"]).value
        if category:
            category = str(category).strip()
        if category not in CATEGORY_ORDER:
            continue
        barcode = ws.cell(row=r, column=headers["商品条码"]).value
        if barcode is not None:
            barcode = str(barcode).strip()
        rows.append({
            "商品名称": name,
            "商品条码": barcode,
            "单位": ws.cell(row=r, column=headers["单位"]).value,
            "商品分类": category,
            "销售数量": _to_num(ws.cell(row=r, column=headers["销售数量"]).value),
            "商品总售价": _to_num(ws.cell(row=r, column=headers["商品总售价"]).value),
        })
    wb.close()
    return rows


def read_discard_data(data_file: Path):
    """读取商品报损周度统计 xls，返回 {条码: 报损数量}。"""
    import xlrd
    wb = xlrd.open_workbook(str(data_file), encoding_override="gbk")
    ws = wb.sheet_by_index(0)

    headers = {}
    for c in range(ws.ncols):
        val = ws.cell_value(0, c)
        if val:
            headers[str(val).strip()] = c

    barcode_col = headers["条码"]
    qty_col = headers["报损数量"]

    discard_map = {}
    for r in range(1, ws.nrows):
        barcode = ws.cell_value(r, barcode_col)
        if not barcode or str(barcode).strip() in ("-", "合计", "总计", ""):
            continue
        barcode = str(barcode).strip()
        qty = ws.cell_value(r, qty_col)
        discard_map[barcode] = _to_num(qty)
    wb.release_resources()
    return discard_map


def _build_discard_lookup(discard_map, sale_rows):
    """构建报损查找表：通过商品条码一对一匹配销售与报损数据。

    discard_map: {条码: 报损数量}
    sale_rows: list[dict]，每行含 "商品名称" 和 "商品条码"
    返回: {商品名称: 报损数量}
    """
    lookup = {}
    for row in sale_rows:
        barcode = row.get("商品条码")
        if barcode and barcode in discard_map:
            lookup[row["商品名称"]] = discard_map[barcode]
    return lookup


def fill_template(sale_rows, discard_map, template_file: Path, output_file: Path,
                  store_short: str, monday: datetime, sunday: datetime):
    """将销售和报损数据填入模板，按分类分组、按商品总售价降序排列。"""
    from openpyxl import load_workbook as _lwb
    from openpyxl.cell.rich_text import CellRichText, TextBlock

    discard_lookup = _build_discard_lookup(discard_map, sale_rows)

    wb = _lwb(template_file, rich_text=True)
    ws = wb.active
    max_col = 9

    merged_ranges = list(ws.merged_cells.ranges)
    for mr in merged_ranges:
        ws.unmerge_cells(str(mr))

    # 保存每个分类的样式模板（每个分类的第一个样例行）
    category_styles = {}
    for cat, sample_row in SAMPLE_ROWS.items():
        category_styles[cat] = list(ws.iter_rows(min_row=sample_row, max_row=sample_row))[0]
    sample_height = ws.row_dimensions[DATA_START_ROW].height

    # 按分类分组，每个分类内按商品总售价降序排列
    grouped = {cat: [] for cat in CATEGORY_ORDER}
    for row in sale_rows:
        cat = row["商品分类"]
        if cat in grouped:
            grouped[cat].append(row)
    for cat in CATEGORY_ORDER:
        grouped[cat].sort(key=lambda r: float(r.get("销售数量") or 0), reverse=True)

    all_sorted = []
    for cat in CATEGORY_ORDER:
        all_sorted.extend(grouped[cat])

    data_count = len(all_sorted)

    # 删除样例行(3~10)，插入数据行
    sample_count = ws.max_row - DATA_START_ROW + 1
    ws.delete_rows(DATA_START_ROW, sample_count)
    ws.insert_rows(DATA_START_ROW, data_count)

    for i, row_data in enumerate(all_sorted):
        r = DATA_START_ROW + i
        cat = row_data["商品分类"]
        name = row_data["商品名称"]

        ws.cell(row=r, column=1).value = name                          # A: 商品名称
        ws.cell(row=r, column=2).value = row_data.get("单位")          # B: 单位
        ws.cell(row=r, column=3).value = cat                           # C: 商品分类
        ws.cell(row=r, column=4).value = row_data.get("销售数量")      # D: 销售数量
        ws.cell(row=r, column=5).value = row_data.get("商品总售价")    # E: 商品总售价
        ws.cell(row=r, column=6).value = f"=D{r}/7"                   # F: 日均销量（公式）
        ws.cell(row=r, column=7).value = discard_lookup.get(name)       # G: 报废量
        ws.cell(row=r, column=8).value = f"=G{r}/(G{r}+D{r})"        # H: 报废率（公式）
        # I: 备注 留空

        style_cells = category_styles[cat]
        for col_idx in range(1, max_col + 1):
            src_idx = min(col_idx - 1, len(style_cells) - 1)
            copy_cell_style(style_cells[src_idx], ws.cell(row=r, column=col_idx))
        if sample_height:
            ws.row_dimensions[r].height = sample_height

    # 更新标题中的门店名和日期
    cell_a1 = ws.cell(row=1, column=1)
    m_start = monday.month
    d_start = monday.day
    m_end = sunday.month
    d_end = sunday.day

    week_in_month = (monday.day - 1) // 7 + 1
    title_text = f"{store_short}{m_start}月份第{week_in_month}周产品销售排行"

    if isinstance(cell_a1.value, CellRichText):
        old_blocks = list(cell_a1.value)
        if old_blocks:
            first = old_blocks[0]
            if isinstance(first, TextBlock):
                cell_a1.value = CellRichText(TextBlock(first.font, title_text))
            else:
                cell_a1.value = title_text
        else:
            cell_a1.value = title_text
    else:
        cell_a1.value = title_text

    # 更新 sheet 名称
    ws.title = f"{store_short}{m_start}月-{week_in_month}"

    # 恢复合并单元格
    from openpyxl.utils import get_column_letter
    last_col_letter = get_column_letter(max_col)
    ws.merge_cells(f"A1:{last_col_letter}1")

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    logger.info(f"  已生成格式化文件: {output_file}")


# ─── 工具函数 ──────────────────────────────────────────────────────────────────


def get_week_range(weeks: int = 0, date_str: str = None):
    if date_str:
        base = datetime.strptime(date_str, "%Y.%m.%d")
    else:
        base = datetime.now()

    monday = base - timedelta(days=base.weekday())
    monday = monday + timedelta(weeks=weeks)
    sunday = monday + timedelta(days=6)

    return monday, sunday


# ─── 浏览器自动化函数 ────────────────────────────────────────────────────────


def set_date(page, placeholder: str, value: str):
    page.evaluate(f"""
        (function() {{
            var inp = document.querySelector('input.timeInput.hasDatepicker[placeholder="{placeholder}"]');
            if (!inp) return 'not found';
            inp.value = "{value}";
            inp.dispatchEvent(new Event('input',  {{bubbles: true}}));
            inp.dispatchEvent(new Event('change', {{bubbles: true}}));
            inp.dispatchEvent(new Event('blur',   {{bubbles: true}}));
            return inp.value;
        }})()
    """)


def click_by_text(page, text: str, desc: str = ""):
    result = page.evaluate(f"""
        (function() {{
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {{
                if (els[i].textContent.trim() === "{text}" && els[i].children.length === 0) {{
                    els[i].click();
                    return 'clicked tag=' + els[i].tagName + ' class=' + els[i].className;
                }}
            }}
            return 'not found';
        }})()
    """)
    tag = f"[{desc}] " if desc else ""
    logger.info(f"  {tag}→ {result}")
    return "clicked" in result


def login(page):
    logger.info("[1/4] 打开登录页面...")
    page.goto(LOGIN_URL)
    page.wait_for_load_state("networkidle")

    logger.info("[2/4] 切换到工号登录模式...")
    page.evaluate("""
        (function() {
            var els = document.querySelectorAll('div,span,a,button');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '工号登录' && els[i].children.length === 0) {
                    els[i].click();
                    return 'clicked';
                }
            }
            return 'not found';
        })()
    """)
    time.sleep(1.5)

    placeholder = page.evaluate("""
        (function() {
            var inputs = document.querySelectorAll('input');
            var r = [];
            for (var i = 0; i < inputs.length; i++) {
                if (inputs[i].placeholder) r.push(inputs[i].placeholder);
            }
            return r.join(' | ');
        })()
    """)
    if "员工工号" not in placeholder:
        raise RuntimeError(f"工号登录切换失败，当前输入框：{placeholder}")
    logger.info(f"  表单已切换 → {placeholder}")

    logger.info("[3/4] 填入账号/工号/密码...")
    page.evaluate(f"""
        (function() {{
            var inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].placeholder === '请输入账号')     inputs[i].value = "{ACCOUNT}";
                if (inputs[i].placeholder === '请输入员工工号') inputs[i].value = "{WORKER_ID}";
                if (inputs[i].placeholder === '请输入工号密码') inputs[i].value = "{PASSWORD}";
            }}
            inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].value) {{
                    inputs[i].dispatchEvent(new Event('input', {{bubbles: true}}));
                }}
            }}
        }})()
    """)

    time.sleep(1.5)

    logger.info("[4/4] 点击登录按钮...")
    click_by_text(page, "登 录", "登录")
    page.wait_for_load_state("networkidle", timeout=120_000)

    time.sleep(2)


def select_store(page, store_full_name: str):
    """从自定义 div 下拉 (#ddl_subUsers) 中选择指定门店。"""
    logger.info(f"  → 选择门店: {store_full_name}")

    # 1. 点击下拉框展开
    dropdown = page.locator("#ddl_subUsers")
    dropdown.click()
    time.sleep(0.5)

    # 2. 先取消所有已选中的 li（带 on class 的），确保单选
    page.evaluate("""
        (function() {
            var lis = document.querySelectorAll('#ddl_subUsers .selectBox li');
            lis.forEach(function(li) {
                if (li.classList.contains('on')) li.click();
            });
        })()
    """)
    time.sleep(0.3)

    # 3. 点击目标门店的 li
    target_li = page.locator(f"#ddl_subUsers .selectBox li[title='{store_full_name}']")
    if target_li.count() > 0:
        target_li.click()
        logger.info(f"    门店已选中: {store_full_name}")
    else:
        # title 里 &nbsp; 可能导致精确匹配失败，用 JS 模糊匹配
        click_result = page.evaluate(f"""
            (function() {{
                var lis = document.querySelectorAll('#ddl_subUsers .selectBox li');
                for (var i = 0; i < lis.length; i++) {{
                    var t = lis[i].textContent.replace(/\\u00a0/g, ' ').trim();
                    if (t === '{store_full_name}') {{
                        lis[i].click();
                        return 'clicked: ' + t;
                    }}
                }}
                var names = [];
                for (var i = 0; i < lis.length; i++) names.push(lis[i].textContent.replace(/\\u00a0/g, ' ').trim());
                return 'not found in: ' + names.join(', ');
            }})()
        """)
        logger.info(f"    模糊匹配: {click_result}")

    time.sleep(0.3)

    # 4. 点击「关闭」收起下拉
    close_btn = page.locator("#ddl_subUsers .bottomBar .btnGrey14")
    if close_btn.count() > 0:
        close_btn.click()
        logger.info("    下拉框已关闭")
    time.sleep(0.5)


def click_advanced_search(page):
    """点击高级搜索/更多搜索，展开搜索选项。"""
    result = page.evaluate("""
        (function() {
            var labels = ['高级搜索', '更多搜索', '更多搜索▼', '高级搜索▼'];
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {
                var t = els[i].textContent.trim();
                for (var j = 0; j < labels.length; j++) {
                    if (t === labels[j] && els[i].children.length === 0) {
                        els[i].click();
                        return 'clicked: ' + t;
                    }
                }
            }
            var btn = document.getElementById('advancedBtn');
            if (btn) { btn.click(); return 'clicked advancedBtn'; }
            return 'not found';
        })()
    """)
    logger.info(f"  [高级搜索] → {result}")
    time.sleep(0.5)


def select_sale_categories(page, categories: list):
    """打开分类选择弹框，勾选指定分类，关闭弹框。"""
    logger.info(f"  → 选择分类: {categories}")

    select_result = page.evaluate("""
        (function() {
            var btn = document.getElementById('selectCategory');
            if (btn) { btn.click(); return 'clicked selectCategory'; }
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '选择分类' && els[i].children.length === 0) {
                    els[i].click();
                    return 'clicked tag=' + els[i].tagName;
                }
            }
            return 'not found';
        })()
    """)
    logger.info(f"    [选择分类] → {select_result}")
    time.sleep(1.0)

    cats_json = json.dumps(categories, ensure_ascii=False)
    page.evaluate(f"""
        (function() {{
            var target = {cats_json};
            var divs = document.querySelectorAll('.checkBoxDiv');
            divs.forEach(function(d) {{
                var span = d.querySelector('span');
                if (!span) return;
                var name = span.textContent.trim();
                var shouldBeOn = target.indexOf(name) >= 0;
                var isOn = d.classList.contains('on');
                if (shouldBeOn && !isOn) d.click();
                if (!shouldBeOn && isOn) d.click();
            }});
        }})()
    """)

    check_result = page.evaluate(f"""
        (function() {{
            var target = {cats_json};
            var divs = document.querySelectorAll('.checkBoxDiv');
            var r = [];
            divs.forEach(function(d) {{
                var s = d.querySelector('span');
                if (s) {{
                    var name = s.textContent.trim();
                    if (target.indexOf(name) >= 0 || d.classList.contains('on')) {{
                        r.push(name + ':' + d.classList.contains('on'));
                    }}
                }}
            }});
            return r.join(' | ');
        }})()
    """)
    logger.info(f"    勾选验证: {check_result}")

    logger.info("  → 点击「确定」关闭弹框...")
    click_by_text(page, "确定", "确定弹框")
    time.sleep(0.5)


def click_export(page):
    """点击导出按钮。"""
    result = page.evaluate("""
        (function() {
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                    els[i].click();
                    return 'ok';
                }
            }
            return 'not found';
        })()
    """)
    logger.info(f"  点击导出: {result}")
    if result == "not found":
        raise RuntimeError("未找到「导出」按钮")


def click_popup_export(page):
    """点击弹窗中的导出按钮（报损页面导出有二次确认弹窗）。"""
    result = page.evaluate("""
        (function() {
            var layers = document.querySelectorAll('.layui-layer');
            for (var m = layers.length - 1; m >= 0; m--) {
                var btns = layers[m].querySelectorAll('*');
                for (var i = 0; i < btns.length; i++) {
                    if (btns[i].textContent.trim() === '导出' && btns[i].children.length === 0) {
                        btns[i].click();
                        return 'clicked layui popup';
                    }
                }
            }
            var modals = document.querySelectorAll('[class*="modal"], [class*="dialog"], [class*="popup"]');
            for (var m = modals.length - 1; m >= 0; m--) {
                var btns = modals[m].querySelectorAll('*');
                for (var i = 0; i < btns.length; i++) {
                    if (btns[i].textContent.trim() === '导出' && btns[i].children.length === 0) {
                        btns[i].click();
                        return 'clicked modal popup';
                    }
                }
            }
            var matches = [];
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                    matches.push(els[i]);
                }
            }
            if (matches.length > 1) {
                matches[matches.length - 1].click();
                return 'clicked last of ' + matches.length;
            }
            return 'not found';
        })()
    """)
    logger.info(f"  [弹窗导出] → {result}")
    if result == "not found":
        raise RuntimeError("未找到弹窗中的「导出」按钮")


def main():
    parser = argparse.ArgumentParser(description="Pospal 麦安研产品销售周度排行自动化脚本")
    parser.add_argument("--weeks", type=int, default=0, help="周偏移量：0=本周，-1=上周（默认0）")
    parser.add_argument("--date", type=str, help="指定日期推断所在周，格式 YYYY.MM.DD")
    parser.add_argument("--headless", action="store_true", help="无头模式（不显示浏览器窗口）")
    args = parser.parse_args()

    if args.date:
        monday, sunday = get_week_range(date_str=args.date)
    else:
        monday, sunday = get_week_range(weeks=args.weeks)

    monday_str = monday.strftime("%Y.%m.%d")
    sunday_str = sunday.strftime("%Y.%m.%d")
    date_range_str = f"{monday.strftime('%Y-%m-%d')}~{sunday.strftime('%Y-%m-%d')}"

    logger.info(f"{'=' * 55}")
    logger.info(f"  Pospal 麦安研产品销售周度排行")
    logger.info(f"  目标周：{monday_str} ~ {sunday_str}")
    logger.info(f"  输出根目录：{OUTPUT_DIR}")
    logger.info(f"{'=' * 55}")

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        logger.error("未安装 playwright，请先运行: pip install playwright && playwright install chromium")
        sys.exit(1)

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=args.headless,
            args=["--start-maximized"],
        )
        context = browser.new_context(
            viewport={"width": 1600, "height": 900},
            accept_downloads=True,
        )
        page = context.new_page()
        page.set_default_timeout(120000)
        page.set_default_navigation_timeout(120000)

        try:
            login(page)

            output_dir = OUTPUT_DIR / date_range_str / "原始下载"
            output_dir.mkdir(parents=True, exist_ok=True)

            # ── Part 1：商品销售周度统计（3个门店）──────────────────────────
            logger.info(f"{'─' * 55}")
            logger.info(f"  Part 1：下载商品销售周度统计")
            logger.info(f"{'─' * 55}")

            for i, store in enumerate(STORES):
                logger.info(f"{'─' * 40}")
                logger.info(f"  门店 {i + 1}/{len(STORES)}: {store['short']} ({store['full']})")
                logger.info(f"{'─' * 40}")

                logger.info("  [导航] 前往商品销售统计页面...")
                page.goto(PRODUCT_SALE_URL)
                page.wait_for_load_state("networkidle", timeout=120_000)
                logger.info(f"  已到达 → {page.url}")

                select_store(page, store["full"])

                logger.info(f"  → 设置日期: {monday_str} ~ {sunday_str}...")
                set_date(page, "开始日期", f"{monday_str} 00:00")
                set_date(page, "结束日期", f"{sunday_str} 23:59")

                click_advanced_search(page)
                select_sale_categories(page, SALE_CATEGORIES)

                logger.info("  [查询] 执行查询...")
                click_by_text(page, "查询", "查询")
                page.wait_for_load_state("networkidle", timeout=150_000)
                time.sleep(3)

                logger.info("  [导出] 导出文件...")
                with page.expect_download(timeout=180_000) as dl_info:
                    click_export(page)

                download = dl_info.value
                logger.info(f"  下载文件名: {download.suggested_filename}")

                dest = output_dir / f"商品销售周度统计_{store['short']}_{date_range_str}.xlsx"
                download.save_as(dest)
                logger.info(f"  已保存到: {dest}")

            # ── Part 2：商品报损周度统计（3个门店）──────────────────────────
            logger.info(f"{'─' * 55}")
            logger.info(f"  Part 2：下载商品报损周度统计")
            logger.info(f"{'─' * 55}")

            for i, store in enumerate(STORES):
                logger.info(f"{'─' * 40}")
                logger.info(f"  门店 {i + 1}/{len(STORES)}: {store['short']} ({store['full']})")
                logger.info(f"{'─' * 40}")

                logger.info("  [导航] 前往商品报损统计页面...")
                page.goto(DISCARD_PRODUCT_URL)
                page.wait_for_load_state("networkidle", timeout=120_000)
                logger.info(f"  已到达 → {page.url}")

                select_store(page, store["full"])

                logger.info(f"  → 设置日期: {monday_str} ~ {sunday_str}...")
                set_date(page, "开始日期", f"{monday_str} 00:00")
                set_date(page, "结束日期", f"{sunday_str} 23:59")

                logger.info("  [查询] 执行查询...")
                click_by_text(page, "查询", "查询")
                page.wait_for_load_state("networkidle", timeout=150_000)
                time.sleep(3)

                logger.info("  [导出] 点击导出...")
                click_export(page)
                time.sleep(1.5)

                logger.info("  [弹窗导出] 点击弹窗中的导出...")
                with page.expect_download(timeout=180_000) as dl_info:
                    click_popup_export(page)

                download = dl_info.value
                logger.info(f"  下载文件名: {download.suggested_filename}")

                dest = output_dir / f"商品报损周度统计_{store['short']}_{date_range_str}.xls"
                download.save_as(dest)
                logger.info(f"  已保存到: {dest}")

            logger.info(f"{'=' * 55}")
            logger.info(f"  麦安研产品销售周度排行下载全部完成！")
            logger.info(f"  输出目录: {output_dir}")
            logger.info(f"{'=' * 55}\n")

            # ── Part 3：格式化数据并生成报表 ──────────────────────────
            logger.info(f"{'─' * 55}")
            logger.info(f"  Part 3：格式化数据并生成报表")
            logger.info(f"{'─' * 55}")

            for i, store in enumerate(STORES):
                logger.info(f"{'─' * 40}")
                logger.info(f"  门店 {i + 1}/{len(STORES)}: {store['short']}")
                logger.info(f"{'─' * 40}")

                sale_file = output_dir / f"商品销售周度统计_{store['short']}_{date_range_str}.xlsx"
                discard_file = output_dir / f"商品报损周度统计_{store['short']}_{date_range_str}.xls"

                logger.info(f"  读取销售数据: {sale_file.name}")
                sale_rows = read_sale_data(sale_file)
                logger.info(f"  共 {len(sale_rows)} 条销售记录")

                logger.info(f"  读取报损数据: {discard_file.name}")
                discard_map = read_discard_data(discard_file)
                logger.info(f"  共 {len(discard_map)} 条报损记录")

                formatted_output = (
                    OUTPUT_DIR / date_range_str
                    / f"麦安研产品销售周度排行_{store['short']}_{date_range_str}.xlsx"
                )
                fill_template(sale_rows, discard_map, TEMPLATE_FILE,
                              formatted_output, store["short"], monday, sunday)

            logger.info(f"{'=' * 55}")
            logger.info(f"  麦安研产品销售周度排行全部完成！")
            logger.info(f"{'=' * 55}\n")

        except Exception as e:
            logger.error(f"任务失败: {e}")
            try:
                screenshot = OUTPUT_DIR / "error_screenshot.png"
                page.screenshot(path=str(screenshot))
                logger.info(f"  错误截图已保存: {screenshot}")
            except Exception:
                pass
            raise
        finally:
            context.close()
            browser.close()


if __name__ == "__main__":
    main()
