"""
工厂配送兔司家门店客户月度业绩对比表 - 自动化脚本
=====================================
依赖：pip install playwright && playwright install chromium

自动登录 Pospal 后台，依次完成工厂配送兔司家门店客户月度业绩对比表所需的数据导出与处理。

用法：
    python factory_delivery_tsj_store_customer_perform_compare_monthly.py                    # 导出本月数据
    python factory_delivery_tsj_store_customer_perform_compare_monthly.py --months -1        # 导出上月数据
    python factory_delivery_tsj_store_customer_perform_compare_monthly.py --date 2026.06.03  # 根据日期推断所在月
    python factory_delivery_tsj_store_customer_perform_compare_monthly.py --headless          # 无头模式（不显示浏览器）
"""

import argparse
import calendar
import copy
import logging
import logging.handlers
import re
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.utils import get_column_letter

# ─── 日志配置 ───────────────────────────────────────────────────────────────────
LOG_DIR = Path(__file__).resolve().parent / "log"
LOG_DIR.mkdir(exist_ok=True)

_file_handler = logging.handlers.TimedRotatingFileHandler(
    LOG_DIR / "factory_delivery_tsj_store_customer_perform_compare_monthly.log",
    when="midnight",
    backupCount=30,
    encoding="utf-8",
)
_file_handler.suffix = "%Y-%m-%d"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[_file_handler, logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# ─── 配置区（按需修改）─────────────────────────────────────────────────────────
ACCOUNT = "huomimayzb"
WORKER_ID = "M006"
PASSWORD = "tusijia88"

LOGIN_URL = "https://beta69.pospal.cn/"
DELIVERY_COMPARISON_URL = "https://css69.pospal.cn/EnterpriseReport/DeliveryComparison"

OUTPUT_DIR = Path(__file__).resolve().parent / "工厂配送兔司家门店客户月度业绩对比表"
TEMPLATE_FILE = Path(__file__).resolve().parent / "工厂配送兔司家门店客户月度业绩对比表_格式化模板.xlsx"

HEADER_ROW = 2
TOTALS_ROW = 3
DATA_START_ROW = 4
SAMPLE_ROW = 4


# ─── 格式化合并函数 ──────────────────────────────────────────────────────────


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def _read_headers(ws):
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val:
            headers[str(val).strip()] = c
    return headers


def _is_stop_name(name):
    if name is None:
        return True
    s = str(name).strip()
    return s in ("-", "合计")


def _to_num(val):
    if val is None:
        return None
    try:
        f = float(val)
        if f == 0:
            return None
        if f == int(f):
            return int(f)
        return f
    except (ValueError, TypeError):
        return val


def read_delivery_comparison(data_file: Path):
    """读取「仓库配送大客户对比表」，提取全部所需字段。

    返回 list[dict]，每行包含：
        大客户名称, 实际销售金额(C), 销售出库单数(I), 销售退货单数(J),
        销售量(K), 退货量(L), 实际销售量(M), 销售金额(N), 退货金额(O)
    """
    wb = load_workbook(data_file, data_only=True)
    ws = wb.active
    headers = _read_headers(ws)

    sum_cols = ["出库量", "退货量", "实际出库量", "出库金额", "退货金额", "实际出库金额",
                "销售出库单数", "销售退货单数"]

    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=headers["大客户名称"]).value
        if _is_stop_name(name):
            break
        row = {"大客户名称": str(name).strip()}
        for col_name in sum_cols:
            if col_name in headers:
                row[col_name] = _to_num(ws.cell(row=r, column=headers[col_name]).value)
            else:
                row[col_name] = None
        rows.append(row)

    wb.close()
    logger.info(f"  仓库配送大客户对比表: {len(rows)} 行")
    return rows


def read_delivery_fee(data_file: Path):
    wb = load_workbook(data_file, data_only=True)
    ws = wb.active
    headers = _read_headers(ws)

    fee_map = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=headers["大客户名称"]).value
        if _is_stop_name(name):
            break
        fee_map[str(name).strip()] = {
            "配送次数": _to_num(ws.cell(row=r, column=headers["出库量"]).value),
            "运费金额": _to_num(ws.cell(row=r, column=headers["实际出库金额"]).value),
        }
    wb.close()
    return fee_map


def read_self_product(data_file: Path):
    wb = load_workbook(data_file, data_only=True)
    ws = wb.active
    headers = _read_headers(ws)

    sp_map = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=headers["大客户名称"]).value
        if _is_stop_name(name):
            break
        sp_map[str(name).strip()] = _to_num(ws.cell(row=r, column=headers["实际出库金额"]).value)
    wb.close()
    return sp_map


def merge_delivery_into_template(delivery_rows, template_file: Path,
                                  output_file: Path, first_day: datetime, last_day: datetime,
                                  fee_data: dict = None, self_product_data: dict = None):
    wb = load_workbook(template_file, rich_text=True)
    ws = wb.active

    max_col = 15

    merged_ranges = list(ws.merged_cells.ranges)
    for mr in merged_ranges:
        ws.unmerge_cells(str(mr))

    sample_cells = list(ws.iter_rows(min_row=SAMPLE_ROW, max_row=SAMPLE_ROW))[0]
    sample_height = ws.row_dimensions[SAMPLE_ROW].height
    totals_cells = list(ws.iter_rows(min_row=TOTALS_ROW, max_row=TOTALS_ROW))[0]

    data_rows_sorted = sorted(delivery_rows,
                              key=lambda r: float(r.get("实际出库金额", 0) or 0),
                              reverse=True)
    data_count = len(data_rows_sorted)

    ws.delete_rows(SAMPLE_ROW, 2)
    ws.insert_rows(DATA_START_ROW, data_count)

    for i, row_data in enumerate(data_rows_sorted):
        r = DATA_START_ROW + i
        name = str(row_data.get("大客户名称", "")).strip()
        ws.cell(row=r, column=1).value = f"=ROW()-3"
        ws.cell(row=r, column=2).value = name

        # C(3) = 实际出库金额/含运费
        ws.cell(row=r, column=3).value = _to_num(row_data.get("实际出库金额"))

        # E(5) = 配送次数, F(6) = 运费金额
        fee_info = fee_data.get(name) if fee_data else None
        if fee_info:
            ws.cell(row=r, column=5).value = fee_info.get("配送次数")
            ws.cell(row=r, column=6).value = fee_info.get("运费金额")

        # H(8) = 自产品金额
        if self_product_data:
            sp_val = self_product_data.get(name)
            if sp_val is not None:
                ws.cell(row=r, column=8).value = sp_val

        # D(4) = C - F
        ws.cell(row=r, column=4).value = f"=C{r}-F{r}"

        # G(7) = C - F - H
        ws.cell(row=r, column=7).value = f"=C{r}-F{r}-H{r}"

        # I(9) = 销售出库单数, J(10) = 销售退货单数
        ws.cell(row=r, column=9).value = row_data.get("销售出库单数")
        ws.cell(row=r, column=10).value = row_data.get("销售退货单数")

        # K(11)=出库量, L(12)=退货量, M(13)=实际出库量, N(14)=出库金额, O(15)=退货金额
        ws.cell(row=r, column=11).value = _to_num(row_data.get("出库量"))
        ws.cell(row=r, column=12).value = _to_num(row_data.get("退货量"))
        ws.cell(row=r, column=13).value = _to_num(row_data.get("实际出库量"))
        ws.cell(row=r, column=14).value = _to_num(row_data.get("出库金额"))
        ws.cell(row=r, column=15).value = _to_num(row_data.get("退货金额"))

        for col_idx in range(1, max_col + 1):
            src_idx = min(col_idx - 1, len(sample_cells) - 1)
            copy_cell_style(sample_cells[src_idx], ws.cell(row=r, column=col_idx))
        if sample_height:
            ws.row_dimensions[r].height = sample_height

    tr = TOTALS_ROW
    last_data_row = DATA_START_ROW + data_count - 1

    ws.cell(row=tr, column=1).value = "合计"
    for col in [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]:
        letter = get_column_letter(col)
        ws.cell(row=tr, column=col).value = f"=SUM({letter}{DATA_START_ROW}:{letter}{last_data_row})"

    for col_idx in range(1, max_col + 1):
        src_idx = min(col_idx - 1, len(totals_cells) - 1)
        copy_cell_style(totals_cells[src_idx], ws.cell(row=tr, column=col_idx))

    date_sheet = f"{first_day.month}.{first_day.day}-{last_day.month}.{last_day.day}"
    old_sheet_name = ws.title
    new_sheet_name = re.sub(r'\d{1,2}\.\d{1,2}-\d{1,2}\.\d{1,2}', date_sheet, old_sheet_name)
    ws.title = new_sheet_name

    cell_a1 = ws.cell(row=1, column=1)
    if cell_a1.value:
        y = first_day.year
        date_pattern = r'\d{4}年\d{1,2}月\d{1,2}日至\d{1,2}月\d{1,2}日'
        date_replacement = f"{y}年{first_day.month}月{first_day.day}日至{last_day.month}月{last_day.day}日"

        if isinstance(cell_a1.value, CellRichText):
            new_blocks = []
            for block in cell_a1.value:
                if isinstance(block, TextBlock):
                    new_blocks.append(TextBlock(block.font, re.sub(date_pattern, date_replacement, block.text)))
                elif isinstance(block, str):
                    new_blocks.append(re.sub(date_pattern, date_replacement, block))
            cell_a1.value = CellRichText(*new_blocks)
        else:
            cell_a1.value = re.sub(date_pattern, date_replacement, str(cell_a1.value))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.merge_cells(start_row=TOTALS_ROW, start_column=1, end_row=TOTALS_ROW, end_column=2)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    logger.info(f"  已生成格式化文件: {output_file}")


# ─── 工具函数 ──────────────────────────────────────────────────────────────────


def get_month_range(months: int = 0, date_str: str = None):
    if date_str:
        base = datetime.strptime(date_str, "%Y.%m.%d")
    else:
        base = datetime.now()

    year = base.year
    month = base.month + months
    while month < 1:
        year -= 1
        month += 12
    while month > 12:
        year += 1
        month -= 12

    first_day = datetime(year, month, 1)
    _, last = calendar.monthrange(year, month)
    last_day = datetime(year, month, last)

    return first_day, last_day


# ─── 浏览器自动化函数 ────────────────────────────────────────────────────────


def set_date(page, placeholder: str, value: str):
    page.evaluate(f"""
        (function() {{
            var inp = document.querySelector('input.timeInput.hasDatepicker[placeholder="{placeholder}"]');
            if (!inp) return 'not found';
            inp.value = "{value}";
            inp.dispatchEvent(new Event('input',  {{bubbles: true}}));
            inp.dispatchEvent(new Event('change', {{bubbles: true}}));
            inp.dispatchEvent(new Event('blur',   {{bubbles: true}}));
            return inp.value;
        }})()
    """)


def click_by_text(page, text: str, desc: str = ""):
    result = page.evaluate(f"""
        (function() {{
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {{
                if (els[i].textContent.trim() === "{text}" && els[i].children.length === 0) {{
                    els[i].click();
                    return 'clicked tag=' + els[i].tagName + ' class=' + els[i].className;
                }}
            }}
            return 'not found';
        }})()
    """)
    tag = f"[{desc}] " if desc else ""
    logger.info(f"  {tag}→ {result}")
    return "clicked" in result


def login(page):
    logger.info("[1/4] 打开登录页面...")
    page.goto(LOGIN_URL)
    page.wait_for_load_state("networkidle")

    logger.info("[2/4] 切换到工号登录模式...")
    page.evaluate("""
        (function() {
            var els = document.querySelectorAll('div,span,a,button');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '工号登录' && els[i].children.length === 0) {
                    els[i].click();
                    return 'clicked';
                }
            }
            return 'not found';
        })()
    """)
    time.sleep(1.5)

    placeholder = page.evaluate("""
        (function() {
            var inputs = document.querySelectorAll('input');
            var r = [];
            for (var i = 0; i < inputs.length; i++) {
                if (inputs[i].placeholder) r.push(inputs[i].placeholder);
            }
            return r.join(' | ');
        })()
    """)
    if "员工工号" not in placeholder:
        raise RuntimeError(f"工号登录切换失败，当前输入框：{placeholder}")
    logger.info(f"  表单已切换 → {placeholder}")

    logger.info("[3/4] 填入账号/工号/密码...")
    page.evaluate(f"""
        (function() {{
            var inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].placeholder === '请输入账号')     inputs[i].value = "{ACCOUNT}";
                if (inputs[i].placeholder === '请输入员工工号') inputs[i].value = "{WORKER_ID}";
                if (inputs[i].placeholder === '请输入工号密码') inputs[i].value = "{PASSWORD}";
            }}
            inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].value) {{
                    inputs[i].dispatchEvent(new Event('input', {{bubbles: true}}));
                }}
            }}
        }})()
    """)

    time.sleep(1.5)

    logger.info("[4/4] 点击登录按钮...")
    click_by_text(page, "登 录", "登录")
    page.wait_for_load_state("networkidle", timeout=120_000)

    time.sleep(2)


def main():
    parser = argparse.ArgumentParser(description="Pospal 工厂配送兔司家门店客户月度业绩对比表自动化脚本")
    parser.add_argument("--months", type=int, default=0, help="月偏移量：0=本月，-1=上月（默认0）")
    parser.add_argument("--date", type=str, help="指定日期推断所在月，格式 YYYY.MM.DD")
    parser.add_argument("--headless", action="store_true", help="无头模式（不显示浏览器窗口）")
    args = parser.parse_args()

    if args.date:
        first_day, last_day = get_month_range(date_str=args.date)
    else:
        first_day, last_day = get_month_range(months=args.months)

    first_day_str = first_day.strftime("%Y.%m.%d")
    last_day_str = last_day.strftime("%Y.%m.%d")
    date_range_str = f"{first_day.strftime('%Y-%m-%d')}~{last_day.strftime('%Y-%m-%d')}"

    logger.info(f"{'=' * 55}")
    logger.info(f"  Pospal 工厂配送兔司家门店客户月度业绩对比表")
    logger.info(f"  目标月：{first_day_str} ~ {last_day_str}")
    logger.info(f"  输出根目录：{OUTPUT_DIR}")
    logger.info(f"{'=' * 55}")

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        logger.error("未安装 playwright，请先运行: pip install playwright && playwright install chromium")
        sys.exit(1)

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=args.headless,
            args=["--start-maximized"],
        )
        context = browser.new_context(
            viewport={"width": 1600, "height": 900},
            accept_downloads=True,
        )
        page = context.new_page()
        page.set_default_timeout(120000)
        page.set_default_navigation_timeout(120000)

        try:
            login(page)

            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：下载仓库配送大客户对比表")
            logger.info(f"{'─' * 55}")

            logger.info("  [导航] 前往仓库配送大客户对比表...")
            page.goto(DELIVERY_COMPARISON_URL)
            page.wait_for_load_state("networkidle", timeout=120_000)
            logger.info(f"  已到达 → {page.url}")

            logger.info(f"  → 设置完成时间: {first_day_str} ~ {last_day_str}...")
            set_date(page, "开始日期", f"{first_day_str} 00:00")
            set_date(page, "结束日期", f"{last_day_str} 23:59")

            logger.info("  [查询] 执行查询...")
            click_by_text(page, "查询", "查询")
            page.wait_for_load_state("networkidle", timeout=150_000)
            time.sleep(3)

            logger.info("  [导出] 导出文件...")
            output_dir = OUTPUT_DIR / date_range_str / "原始下载"
            output_dir.mkdir(parents=True, exist_ok=True)

            with page.expect_download(timeout=180_000) as dl_info:
                result = page.evaluate("""
                    (function() {
                        var els = document.querySelectorAll('*');
                        for (var i = 0; i < els.length; i++) {
                            if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                                els[i].click();
                                return 'ok';
                            }
                        }
                        return 'not found';
                    })()
                """)
                logger.info(f"  点击导出: {result}")
                if result == "not found":
                    raise RuntimeError("未找到「导出」按钮")

            download = dl_info.value
            logger.info(f"  下载文件名: {download.suggested_filename}")

            dest = output_dir / f"仓库配送大客户对比表_{date_range_str}.xlsx"
            download.save_as(dest)
            logger.info(f"  已保存到: {dest}")

            # ── 步骤 1.5：仓库配送大客户对比表_配送费 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：下载仓库配送大客户对比表_配送费")
            logger.info(f"{'─' * 55}")

            logger.info("  → 点击更多搜索...")
            more_search_result = page.evaluate("""
                (function() {
                    var els = document.querySelectorAll('*');
                    for (var i = 0; i < els.length; i++) {
                        var t = els[i].textContent.trim();
                        if ((t === '更多搜索' || t === '更多搜索▼') && els[i].children.length === 0) {
                            els[i].click();
                            return 'clicked tag=' + els[i].tagName + ' class=' + els[i].className;
                        }
                    }
                    var btn = document.getElementById('advancedBtn');
                    if (btn) { btn.click(); return 'clicked advancedBtn'; }
                    return 'not found';
                })()
            """)
            logger.info(f"  [更多搜索] → {more_search_result}")
            time.sleep(0.5)

            logger.info("  → 打开分类选择弹框...")
            select_cat_result = page.evaluate("""
                (function() {
                    var btn = document.getElementById('selectCategory');
                    if (btn) { btn.click(); return 'clicked selectCategory'; }
                    var els = document.querySelectorAll('*');
                    for (var i = 0; i < els.length; i++) {
                        if (els[i].textContent.trim() === '选择分类' && els[i].children.length === 0) {
                            els[i].click();
                            return 'clicked tag=' + els[i].tagName;
                        }
                    }
                    return 'not found';
                })()
            """)
            logger.info(f"  [选择分类] → {select_cat_result}")
            time.sleep(1.0)

            logger.info("  → 勾选「配送费」...")
            page.evaluate("""
                (function() {
                    var divs = document.querySelectorAll('.checkBoxDiv');
                    divs.forEach(function(d) {
                        var span = d.querySelector('span');
                        if (span && span.textContent.trim() === '配送费') {
                            if (!d.classList.contains('on')) {
                                d.click();
                            }
                        }
                    });
                })()
            """)

            check_result = page.evaluate("""
                (function() {
                    var divs = document.querySelectorAll('.checkBoxDiv');
                    var r = [];
                    divs.forEach(function(d) {
                        var s = d.querySelector('span');
                        if (s && s.textContent.trim() === '配送费') {
                            r.push(s.textContent.trim() + ':' + d.classList.contains('on'));
                        }
                    });
                    return r.join(' | ');
                })()
            """)
            logger.info(f"    勾选验证: {check_result}")

            logger.info("  → 点击「确定」关闭弹框...")
            click_by_text(page, "确定", "确定弹框")
            time.sleep(0.5)

            logger.info("  [查询] 执行查询...")
            click_by_text(page, "查询", "查询")
            page.wait_for_load_state("networkidle", timeout=150_000)
            time.sleep(3)

            logger.info("  [导出] 导出配送费文件...")
            with page.expect_download(timeout=180_000) as dl_info_fee:
                result = page.evaluate("""
                    (function() {
                        var els = document.querySelectorAll('*');
                        for (var i = 0; i < els.length; i++) {
                            if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                                els[i].click();
                                return 'ok';
                            }
                        }
                        return 'not found';
                    })()
                """)
                logger.info(f"  点击导出: {result}")
                if result == "not found":
                    raise RuntimeError("未找到「导出」按钮")

            download_fee = dl_info_fee.value
            logger.info(f"  下载文件名: {download_fee.suggested_filename}")

            dest_fee = output_dir / f"仓库配送大客户对比表_配送费_{date_range_str}.xlsx"
            download_fee.save_as(dest_fee)
            logger.info(f"  已保存到: {dest_fee}")

            # ── 步骤 1.7：仓库配送大客户对比表_自产品 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：下载仓库配送大客户对比表_自产品")
            logger.info(f"{'─' * 55}")

            logger.info("  → 点击重新搜索...")
            click_by_text(page, "重新搜索", "重新搜索")
            time.sleep(1.0)

            logger.info("  → 打开分类选择弹框...")
            select_cat_result = page.evaluate("""
                (function() {
                    var btn = document.getElementById('selectCategory');
                    if (btn) { btn.click(); return 'clicked selectCategory'; }
                    var els = document.querySelectorAll('*');
                    for (var i = 0; i < els.length; i++) {
                        if (els[i].textContent.trim() === '选择分类' && els[i].children.length === 0) {
                            els[i].click();
                            return 'clicked tag=' + els[i].tagName;
                        }
                    }
                    return 'not found';
                })()
            """)
            logger.info(f"  [选择分类] → {select_cat_result}")
            time.sleep(1.0)

            logger.info("  → 取消「配送费」，勾选「冷冻面团」和「蛋糕及面包成品及饼干类」...")
            page.evaluate("""
                (function() {
                    var divs = document.querySelectorAll('.checkBoxDiv');
                    divs.forEach(function(d) {
                        var span = d.querySelector('span');
                        if (!span) return;
                        var name = span.textContent.trim();
                        if (name === '配送费') {
                            if (d.classList.contains('on')) {
                                d.click();
                            }
                        }
                        if (name === '冷冻面团' || name === '蛋糕及面包成品及饼干类') {
                            if (!d.classList.contains('on')) {
                                d.click();
                            }
                        }
                    });
                })()
            """)

            check_result = page.evaluate("""
                (function() {
                    var divs = document.querySelectorAll('.checkBoxDiv');
                    var r = [];
                    divs.forEach(function(d) {
                        var s = d.querySelector('span');
                        if (s) {
                            var name = s.textContent.trim();
                            if (name === '配送费' || name === '冷冻面团' || name === '蛋糕及面包成品及饼干类') {
                                r.push(name + ':' + d.classList.contains('on'));
                            }
                        }
                    });
                    return r.join(' | ');
                })()
            """)
            logger.info(f"    勾选验证: {check_result}")

            logger.info("  → 点击「确定」关闭弹框...")
            click_by_text(page, "确定", "确定弹框")
            time.sleep(0.5)

            logger.info("  [查询] 执行查询...")
            click_by_text(page, "查询", "查询")
            page.wait_for_load_state("networkidle", timeout=150_000)
            time.sleep(3)

            logger.info("  [导出] 导出自产品文件...")
            with page.expect_download(timeout=180_000) as dl_info_self:
                result = page.evaluate("""
                    (function() {
                        var els = document.querySelectorAll('*');
                        for (var i = 0; i < els.length; i++) {
                            if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                                els[i].click();
                                return 'ok';
                            }
                        }
                        return 'not found';
                    })()
                """)
                logger.info(f"  点击导出: {result}")
                if result == "not found":
                    raise RuntimeError("未找到「导出」按钮")

            download_self = dl_info_self.value
            logger.info(f"  下载文件名: {download_self.suggested_filename}")

            dest_self = output_dir / f"仓库配送大客户对比表_自产品_{date_range_str}.xlsx"
            download_self.save_as(dest_self)
            logger.info(f"  已保存到: {dest_self}")

            logger.info(f"{'=' * 55}")
            logger.info(f"  工厂配送兔司家门店客户月度业绩对比表下载完成！")
            logger.info(f"  仓库配送大客户对比表 → {dest}")
            logger.info(f"  仓库配送大客户对比表_配送费 → {dest_fee}")
            logger.info(f"  仓库配送大客户对比表_自产品 → {dest_self}")
            logger.info(f"{'=' * 55}\n")

            # ── 步骤 2：格式化数据并生成报表 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：格式化数据并生成报表")
            logger.info(f"{'─' * 55}")

            logger.info("  读取仓库配送大客户对比表...")
            delivery_rows = read_delivery_comparison(dest)
            logger.info(f"  共 {len(delivery_rows)} 个大客户")

            logger.info("  读取配送费数据...")
            fee_data = read_delivery_fee(dest_fee)
            logger.info(f"  共 {len(fee_data)} 条配送费数据")

            logger.info("  读取自产品数据...")
            self_product_data = read_self_product(dest_self)
            logger.info(f"  共 {len(self_product_data)} 条自产品数据")

            formatted_output = OUTPUT_DIR / date_range_str / f"工厂配送兔司家门店客户月度业绩对比表_{date_range_str}.xlsx"
            merge_delivery_into_template(delivery_rows, TEMPLATE_FILE,
                                         formatted_output, first_day, last_day, fee_data,
                                         self_product_data)
            logger.info(f"  格式化输出 → {formatted_output}")

            logger.info(f"{'=' * 55}")
            logger.info(f"  工厂配送兔司家门店客户月度业绩对比表全部完成！")
            logger.info(f"{'=' * 55}\n")

        except Exception as e:
            logger.error(f"任务失败: {e}")
            try:
                screenshot = OUTPUT_DIR / "error_screenshot.png"
                page.screenshot(path=str(screenshot))
                logger.info(f"  错误截图已保存: {screenshot}")
            except Exception:
                pass
            raise
        finally:
            context.close()
            browser.close()


if __name__ == "__main__":
    main()
