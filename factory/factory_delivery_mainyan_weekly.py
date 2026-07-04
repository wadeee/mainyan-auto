"""
工厂配送麦安研门店周度销售报表 - 自动化脚本
=====================================
依赖：pip install playwright openpyxl && playwright install chromium

自动登录 Pospal 后台，导出仓库配送商品门店对比表并生成格式化报表。

用法：
    python factory_delivery_mainyan_weekly.py                    # 导出本周数据（周一到周日）
    python factory_delivery_mainyan_weekly.py --weeks -1         # 导出上周数据
    python factory_delivery_mainyan_weekly.py --date 2026.06.03  # 根据日期推断所在周
    python factory_delivery_mainyan_weekly.py --headless          # 无头模式（不显示浏览器）
"""

import argparse
import copy
import logging
import logging.handlers
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.utils import get_column_letter

# ─── 日志配置 ───────────────────────────────────────────────────────────────────
LOG_DIR = Path(__file__).resolve().parent / "log"
LOG_DIR.mkdir(exist_ok=True)

_file_handler = logging.handlers.TimedRotatingFileHandler(
    LOG_DIR / "factory_delivery_mainyan_weekly.log",
    when="midnight",
    backupCount=30,
    encoding="utf-8",
)
_file_handler.suffix = "%Y-%m-%d"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[_file_handler, logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# ─── 配置区（按需修改）─────────────────────────────────────────────────────────
ACCOUNT = "huomimayzb"
WORKER_ID = "M006"
PASSWORD = "tusijia88"

LOGIN_URL = "https://beta69.pospal.cn/"
REPORT_URL = "https://css69.pospal.cn/ChainStoreSupplyReport/WarehouseDeliveryProductStoreComparison"
DELIVERY_PRODUCT_COMPARISON_URL = "https://css69.pospal.cn/EnterpriseReport/DeliveryProductComparison"

OUTPUT_DIR = Path(__file__).resolve().parent / "工厂配送麦安研门店周度销售报表"
TEMPLATE_FILE = Path(__file__).resolve().parent / "工厂配送麦安研门店周度销售报表_格式化模板.xlsx"

STORE_NAME_MAP = {
    "麦安研（顺德杏坛店）": "杏坛",
    "麦安研（东站宝泰店）": "宝泰",
    "麦安研（佛山创产店）": "创产",
    "麦安研（顺德龙江店）": "龙江",
    "麦安研（佛山万民金海城店）": "金海城",
    "麦安研门店裱花间": "裱花间",
    "中央工厂冷加工间": "冷加工间",
}

BMX_STORE_ORDER = ["广钢", "滨江", "南州路"]
BMX_CUSTOMER_NAME_MAP = {
    "焙满香广钢店": "广钢",
    "焙满香滨江店": "滨江",
    "焙满香南州路店": "南州路",
}

TARGET_CATEGORIES = [
    "配送费",
    "包材耗材",
    "工衣模具",
    "慕斯+饼干+饮品+其他",
    "原料铺料",
    "冷冻面团",
    "蛋糕及面包成品及饼干类",
]

CATEGORY_ORDER = [
    "冷冻面团",
    "蛋糕及面包成品及饼干类",
    "慕斯+饼干+饮品+其他",
    "原料铺料",
    "包材耗材",
    "工衣模具",
    "配送费",
]

HEADER_ROW = 2
TOTALS_ROW = 3
DATA_START_ROW = 4
SAMPLE_ROW = 4


# ─── 工具函数 ──────────────────────────────────────────────────────────────────


def _to_num(val):
    if val is None:
        return None
    try:
        f = float(val)
        if f == 0:
            return None
        if f == int(f):
            return int(f)
        return f
    except (ValueError, TypeError):
        return val


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def clear_cell(cell):
    from openpyxl.styles import Font, PatternFill, Border, Alignment
    cell.value = None
    cell.font = Font()
    cell.fill = PatternFill()
    cell.border = Border()
    cell.alignment = Alignment()
    cell.number_format = "General"


def _style_ref_idx(col_idx, sample_len):
    """列号 → 样式参照索引。1~8 直接映射；9+ 按 I/J 两列一组循环。"""
    if col_idx <= 8:
        return min(col_idx - 1, sample_len - 1)
    offset = (col_idx - 9) % 2
    ref = 8 + offset
    return min(ref, sample_len - 1)


def _get_short_name(full_name):
    short = STORE_NAME_MAP.get(full_name)
    if short:
        return short
    m = re.search(r'[（(](.+?)[）)]', full_name)
    if m:
        return m.group(1).rstrip('店')
    return full_name


def get_week_range(weeks=0, date_str=None):
    if date_str:
        base = datetime.strptime(date_str, "%Y.%m.%d")
    else:
        base = datetime.now()
    monday = base - timedelta(days=base.weekday())
    monday = monday + timedelta(weeks=weeks)
    sunday = monday + timedelta(days=6)
    return monday, sunday


def _update_rich_text_date(cell, monday, sunday):
    if cell.value is None:
        return
    date_pattern = r'\d{4}年\d{1,2}月\d{1,2}日至\d{1,2}月\d{1,2}日'
    date_replacement = f"{monday.year}年{monday.month}月{monday.day}日至{sunday.month}月{sunday.day}日"
    if isinstance(cell.value, CellRichText):
        new_blocks = []
        for block in cell.value:
            if isinstance(block, TextBlock):
                new_blocks.append(TextBlock(block.font, re.sub(date_pattern, date_replacement, block.text)))
            elif isinstance(block, str):
                new_blocks.append(re.sub(date_pattern, date_replacement, block))
        cell.value = CellRichText(*new_blocks)
    else:
        cell.value = re.sub(date_pattern, date_replacement, str(cell.value))


# ─── 数据读取 ──────────────────────────────────────────────────────────────────


def read_downloaded_data(data_file):
    """读取仓库配送商品门店对比表，返回 (store_order, products)。

    store_order: 活跃门店简称列表
    products: 商品明细列表，每项包含 stores 字典（{简称: {数量, 金额}}）
    """
    wb = load_workbook(data_file, data_only=True)
    ws = wb.active

    # 扫描门店列组：每组 3 列（数量、成本金额、出库金额），跳过合计
    store_groups = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h is None:
            continue
        h_str = str(h).strip()
        if h_str.endswith("_数量"):
            name_part = h_str.rsplit("_", 1)[0]
            if name_part != "合计":
                store_groups.append((name_part, c, c + 2))

    # 过滤：仅保留三列中存在非零数据的门店
    active_stores = []
    for full_name, qty_col, amt_col in store_groups:
        has_data = False
        for r in range(2, ws.max_row + 1):
            for col in [qty_col, qty_col + 1, amt_col]:
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    try:
                        if float(v) != 0:
                            has_data = True
                            break
                    except (ValueError, TypeError):
                        pass
            if has_data:
                break
        if has_data:
            short_name = _get_short_name(full_name)
            active_stores.append((short_name, qty_col, amt_col))

    logger.info(f"  活跃门店 ({len(active_stores)}): {', '.join(s[0] for s in active_stores)}")

    # 读取商品数据（来源列：D=商品大类, E=商品分类, F=商品名称, I=规格, J=单位）
    products = []
    for r in range(2, ws.max_row + 1):
        seq = ws.cell(row=r, column=1).value
        if seq is None:
            break
        product = {
            "商品大类": str(ws.cell(row=r, column=4).value or "").strip(),
            "商品分类": str(ws.cell(row=r, column=5).value or "").strip(),
            "商品名称": str(ws.cell(row=r, column=6).value or "").strip(),
            "规格": ws.cell(row=r, column=9).value,
            "单位": ws.cell(row=r, column=10).value,
            "stores": {},
        }
        for short_name, qty_col, amt_col in active_stores:
            product["stores"][short_name] = {
                "数量": _to_num(ws.cell(row=r, column=qty_col).value),
                "金额": _to_num(ws.cell(row=r, column=amt_col).value),
            }
        products.append(product)

    wb.close()
    store_order = [s[0] for s in active_stores]
    logger.info(f"  共 {len(products)} 条商品数据")
    return store_order, products


def read_bmx_store_data(data_file):
    """读取仓库配送商品大客户对比表，提取焙满香三家门店数据，按商品名称透视。

    返回 {商品名称: {"info": {商品大类, 商品分类, 规格, 单位}, "stores": {store_short: {数量, 金额}}}}
    """
    wb = load_workbook(data_file, data_only=True)
    ws = wb.active

    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val:
            headers[str(val).strip()] = c

    pivot = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=headers["大客户名称"]).value
        if name is None:
            break
        name = str(name).strip()
        if name == "合计" or name == "-":
            break

        store_short = BMX_CUSTOMER_NAME_MAP.get(name)
        if not store_short:
            continue

        product_name = str(ws.cell(row=r, column=headers["商品名称"]).value or "").strip()
        if not product_name:
            continue

        if product_name not in pivot:
            pivot[product_name] = {
                "info": {
                    "商品大类": str(ws.cell(row=r, column=headers["商品大类"]).value or "").strip(),
                    "商品分类": str(ws.cell(row=r, column=headers["商品分类"]).value or "").strip(),
                    "规格": ws.cell(row=r, column=headers["规格"]).value,
                    "单位": ws.cell(row=r, column=headers["单位"]).value,
                },
                "stores": {},
            }
        if store_short not in pivot[product_name]["stores"]:
            pivot[product_name]["stores"][store_short] = {"数量": 0.0, "金额": 0.0}

        qty = ws.cell(row=r, column=headers["实际销售量"]).value
        amt = ws.cell(row=r, column=headers["实际销售金额"]).value
        if qty:
            try:
                pivot[product_name]["stores"][store_short]["数量"] += float(qty)
            except (ValueError, TypeError):
                pass
        if amt:
            try:
                pivot[product_name]["stores"][store_short]["金额"] += float(amt)
            except (ValueError, TypeError):
                pass

    wb.close()

    for pname in pivot:
        for store in pivot[pname]["stores"]:
            pivot[pname]["stores"][store]["数量"] = _to_num(pivot[pname]["stores"][store]["数量"])
            pivot[pname]["stores"][store]["金额"] = _to_num(pivot[pname]["stores"][store]["金额"])

    logger.info(f"  焙满香门店数据: {len(pivot)} 个商品")
    return pivot


def merge_products_with_bmx(products, bmx_data):
    """将焙满香门店数据合并到产品列表中。

    - 已存在的商品：将焙满香门店数据追加到 stores 字典
    - 仅焙满香有的商品：新建产品条目
    返回 (合并后的产品列表, 活跃焙满香门店列表)
    """
    bmx_active_order = []
    for short_name in BMX_STORE_ORDER:
        has_data = any(
            (stores["stores"].get(short_name, {}).get("数量") is not None or
             stores["stores"].get(short_name, {}).get("金额") is not None)
            for stores in bmx_data.values()
        )
        if has_data:
            bmx_active_order.append(short_name)

    if not bmx_active_order:
        return products, []

    logger.info(f"  焙满香活跃门店 ({len(bmx_active_order)}): {', '.join(bmx_active_order)}")

    product_name_set = set()
    for p in products:
        pname = p["商品名称"]
        product_name_set.add(pname)
        if pname in bmx_data:
            for store_short in bmx_active_order:
                sd = bmx_data[pname]["stores"].get(store_short)
                if sd:
                    p["stores"][store_short] = sd

    bmx_only_count = 0
    for pname, data in bmx_data.items():
        if pname in product_name_set:
            continue
        stores = {}
        for store_short in bmx_active_order:
            sd = data["stores"].get(store_short)
            if sd:
                stores[store_short] = sd
        if stores:
            products.append({
                "商品大类": data["info"]["商品大类"],
                "商品分类": data["info"]["商品分类"],
                "商品名称": pname,
                "规格": data["info"]["规格"],
                "单位": data["info"]["单位"],
                "stores": stores,
            })
            bmx_only_count += 1

    if bmx_only_count:
        logger.info(f"  焙满香独有商品: {bmx_only_count} 条（已合并）")

    return products, bmx_active_order


# ─── 排序 ──────────────────────────────────────────────────────────────────────


def _sort_key_s1(product):
    """按商品大类顺序排序，同类内按合计金额降序。"""
    cat = product.get("商品大类", "")
    try:
        cat_idx = CATEGORY_ORDER.index(cat)
    except ValueError:
        cat_idx = len(CATEGORY_ORDER)
    total_amt = sum((s.get("金额") or 0) for s in product.get("stores", {}).values())
    return (cat_idx, -total_amt)


# ─── Sheet 1：货品对比表 ───────────────────────────────────────────────────────


def fill_s1(wb, sorted_products, store_order, monday, sunday):
    ws = wb.worksheets[0]
    original_max_col = ws.max_column

    store_count = len(store_order)
    s1_max_col = 8 + store_count * 2

    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    sample_cells = list(ws.iter_rows(min_row=SAMPLE_ROW, max_row=SAMPLE_ROW))[0]
    sample_height = ws.row_dimensions[SAMPLE_ROW].height

    sample_count = ws.max_row - SAMPLE_ROW + 1
    ws.delete_rows(SAMPLE_ROW, sample_count)
    data_count = len(sorted_products)
    ws.insert_rows(DATA_START_ROW, data_count)

    store_columns = {}
    for idx, short_name in enumerate(store_order):
        qty_col = 9 + idx * 2
        amt_col = 10 + idx * 2
        store_columns[short_name] = (qty_col, amt_col)

    for short_name, (qty_col, amt_col) in store_columns.items():
        ws.cell(row=HEADER_ROW, column=qty_col).value = f"{short_name}数量"
        ws.cell(row=HEADER_ROW, column=amt_col).value = f"{short_name}金额"

    qty_letters = [get_column_letter(9 + i * 2) for i in range(store_count)]
    amt_letters = [get_column_letter(10 + i * 2) for i in range(store_count)]

    for i, product in enumerate(sorted_products):
        r = DATA_START_ROW + i
        ws.cell(row=r, column=1).value = f"=ROW()-3"
        ws.cell(row=r, column=2).value = product["商品大类"]
        ws.cell(row=r, column=3).value = product["商品分类"]
        ws.cell(row=r, column=4).value = product["商品名称"]
        ws.cell(row=r, column=5).value = product["规格"]
        ws.cell(row=r, column=6).value = product["单位"]
        ws.cell(row=r, column=7).value = f"=SUM({','.join(c + str(r) for c in qty_letters)})"
        ws.cell(row=r, column=8).value = f"=SUM({','.join(c + str(r) for c in amt_letters)})"

        for short_name, (qty_col, amt_col) in store_columns.items():
            store_data = product["stores"].get(short_name, {})
            qty_val = store_data.get("数量")
            amt_val = store_data.get("金额")
            if qty_val is not None:
                ws.cell(row=r, column=qty_col).value = qty_val
            if amt_val is not None:
                ws.cell(row=r, column=amt_col).value = amt_val

        for col_idx in range(1, s1_max_col + 1):
            src_idx = _style_ref_idx(col_idx, len(sample_cells))
            copy_cell_style(sample_cells[src_idx], ws.cell(row=r, column=col_idx))
        if sample_height:
            ws.row_dimensions[r].height = sample_height

    last_data_row = DATA_START_ROW + data_count - 1
    ws.cell(row=TOTALS_ROW, column=1).value = "合计"
    for col in range(7, s1_max_col + 1):
        letter = get_column_letter(col)
        ws.cell(row=TOTALS_ROW, column=col).value = (
            f"=SUM({letter}{DATA_START_ROW}:{letter}{last_data_row})"
        )

    for hr in [HEADER_ROW, TOTALS_ROW]:
        ref_qty = ws.cell(row=hr, column=9)
        ref_amt = ws.cell(row=hr, column=10)
        for c in range(9, s1_max_col + 1):
            ref = ref_qty if (c - 9) % 2 == 0 else ref_amt
            copy_cell_style(ref, ws.cell(row=hr, column=c))

    i_width = ws.column_dimensions["I"].width
    j_width = ws.column_dimensions["J"].width
    if i_width and j_width:
        for idx in range(store_count):
            ws.column_dimensions[get_column_letter(9 + idx * 2)].width = i_width
            ws.column_dimensions[get_column_letter(10 + idx * 2)].width = j_width

    if s1_max_col < original_max_col:
        for r in range(1, ws.max_row + 1):
            for c in range(s1_max_col + 1, original_max_col + 1):
                clear_cell(ws.cell(row=r, column=c))

    last_col_letter = get_column_letter(s1_max_col)
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws.merge_cells(f"A{TOTALS_ROW}:F{TOTALS_ROW}")

    date_sheet = f"{monday.month}.{monday.day}-{sunday.month}.{sunday.day}"
    ws.title = re.sub(r'\d{1,2}\.\d{1,2}-\d{1,2}\.\d{1,2}', date_sheet, ws.title)
    _update_rich_text_date(ws.cell(row=1, column=1), monday, sunday)


# ─── Sheet 2：货品销售排行表 ───────────────────────────────────────────────────


def fill_s2(wb, sorted_products, monday, sunday):
    ws = wb.worksheets[1]
    s2_max_col = 8

    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    sample_cells = list(ws.iter_rows(min_row=SAMPLE_ROW, max_row=SAMPLE_ROW))[0]
    sample_height = ws.row_dimensions[SAMPLE_ROW].height

    sample_count = ws.max_row - SAMPLE_ROW + 1
    ws.delete_rows(SAMPLE_ROW, sample_count)
    data_count = len(sorted_products)
    ws.insert_rows(DATA_START_ROW, data_count)

    for i, product in enumerate(sorted_products):
        r = DATA_START_ROW + i
        ws.cell(row=r, column=1).value = f"=ROW()-3"
        ws.cell(row=r, column=2).value = product["商品大类"]
        ws.cell(row=r, column=3).value = product["商品分类"]
        ws.cell(row=r, column=4).value = product["商品名称"]
        ws.cell(row=r, column=5).value = product["规格"]
        ws.cell(row=r, column=6).value = product["单位"]

        total_qty = sum((s.get("数量") or 0) for s in product["stores"].values())
        total_amt = sum((s.get("金额") or 0) for s in product["stores"].values())
        ws.cell(row=r, column=7).value = _to_num(total_qty)
        ws.cell(row=r, column=8).value = _to_num(total_amt)

        for col_idx in range(1, s2_max_col + 1):
            src_idx = min(col_idx - 1, len(sample_cells) - 1)
            copy_cell_style(sample_cells[src_idx], ws.cell(row=r, column=col_idx))
        if sample_height:
            ws.row_dimensions[r].height = sample_height

    last_data_row = DATA_START_ROW + data_count - 1
    ws.cell(row=TOTALS_ROW, column=1).value = "合计"
    for col in [7, 8]:
        letter = get_column_letter(col)
        ws.cell(row=TOTALS_ROW, column=col).value = (
            f"=SUM({letter}{DATA_START_ROW}:{letter}{last_data_row})"
        )

    last_col_letter = get_column_letter(s2_max_col)
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws.merge_cells(f"A{TOTALS_ROW}:F{TOTALS_ROW}")

    date_sheet = f"{monday.month}.{monday.day}-{sunday.month}.{sunday.day}"
    ws.title = re.sub(r'\d{1,2}\.\d{1,2}-\d{1,2}\.\d{1,2}', date_sheet, ws.title)
    _update_rich_text_date(ws.cell(row=1, column=1), monday, sunday)


# ─── Sheet 3+：门店产品销售排行表 ─────────────────────────────────────────────


def fill_s3_sheets(wb, sorted_products, store_order, monday, sunday, brand_map=None):
    """为每个门店创建独立的销售排行 sheet。"""
    s3_template = wb.worksheets[2]

    s3_sheets = [s3_template]
    for _ in range(1, len(store_order)):
        s3_sheets.append(wb.copy_worksheet(s3_template))

    for idx, short_name in enumerate(store_order):
        ws = s3_sheets[idx]

        store_products = []
        for p in sorted_products:
            sd = p["stores"].get(short_name, {})
            qty = sd.get("数量") or 0
            amt = sd.get("金额") or 0
            if qty != 0 or amt != 0:
                store_products.append({
                    "商品大类": p["商品大类"],
                    "商品分类": p["商品分类"],
                    "商品名称": p["商品名称"],
                    "规格": p["规格"],
                    "单位": p["单位"],
                    "数量": _to_num(qty),
                    "金额": _to_num(amt),
                })

        store_products.sort(key=lambda x: -(x["金额"] or 0))
        brand = (brand_map or {}).get(short_name, "麦安研")
        _fill_single_s3(ws, store_products, short_name, monday, sunday, brand=brand)


def _fill_single_s3(ws, store_products, short_name, monday, sunday, brand="麦安研"):
    s3_max_col = 8

    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    sample_cells = list(ws.iter_rows(min_row=SAMPLE_ROW, max_row=SAMPLE_ROW))[0]
    sample_height = ws.row_dimensions[SAMPLE_ROW].height

    sample_count = ws.max_row - SAMPLE_ROW + 1
    ws.delete_rows(SAMPLE_ROW, sample_count)
    data_count = len(store_products)
    if data_count > 0:
        ws.insert_rows(DATA_START_ROW, data_count)

    for i, product in enumerate(store_products):
        r = DATA_START_ROW + i
        ws.cell(row=r, column=1).value = f"=ROW()-3"
        ws.cell(row=r, column=2).value = product["商品大类"]
        ws.cell(row=r, column=3).value = product["商品分类"]
        ws.cell(row=r, column=4).value = product["商品名称"]
        ws.cell(row=r, column=5).value = product["规格"]
        ws.cell(row=r, column=6).value = product["单位"]
        ws.cell(row=r, column=7).value = product["数量"]
        ws.cell(row=r, column=8).value = product["金额"]

        for col_idx in range(1, s3_max_col + 1):
            src_idx = min(col_idx - 1, len(sample_cells) - 1)
            copy_cell_style(sample_cells[src_idx], ws.cell(row=r, column=col_idx))
        if sample_height:
            ws.row_dimensions[r].height = sample_height

    ws.cell(row=TOTALS_ROW, column=1).value = "合计"
    if data_count > 0:
        last_data_row = DATA_START_ROW + data_count - 1
        for col in [7, 8]:
            letter = get_column_letter(col)
            ws.cell(row=TOTALS_ROW, column=col).value = (
                f"=SUM({letter}{DATA_START_ROW}:{letter}{last_data_row})"
            )
    else:
        ws.cell(row=TOTALS_ROW, column=7).value = 0
        ws.cell(row=TOTALS_ROW, column=8).value = 0

    ws.title = f"配送{short_name}店产品销售排行表"
    ws.cell(row=HEADER_ROW, column=7).value = f"{short_name}数量"
    ws.cell(row=HEADER_ROW, column=8).value = f"{short_name}金额"

    cell_a1 = ws.cell(row=1, column=1)
    date_pattern = r'\d{4}年\d{1,2}月\d{1,2}日至\d{1,2}月\d{1,2}日'
    date_replacement = f"{monday.year}年{monday.month}月{monday.day}日至{sunday.month}月{sunday.day}日"
    if cell_a1.value:
        if isinstance(cell_a1.value, CellRichText):
            text = ''.join(b.text if isinstance(b, TextBlock) else b for b in cell_a1.value)
            new_text = re.sub(date_pattern, date_replacement, text)
            new_text = re.sub(r'(麦安研|焙满香)\S+门店', f'{brand}{short_name}门店', new_text)
            cell_a1.value = new_text
        else:
            new_val = re.sub(date_pattern, date_replacement, str(cell_a1.value))
            new_val = re.sub(r'(麦安研|焙满香)\S+门店', f'{brand}{short_name}门店', new_val)
            cell_a1.value = new_val

    last_col_letter = get_column_letter(s3_max_col)
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws.merge_cells(f"A{TOTALS_ROW}:F{TOTALS_ROW}")


# ─── 格式化合并主函数 ──────────────────────────────────────────────────────────


def merge_into_template(data_file, template_file, output_file, monday, sunday, bmx_data=None):
    logger.info("读取下载数据...")
    store_order, products = read_downloaded_data(data_file)

    # 合并焙满香门店数据到统一的产品列表
    bmx_active_order = []
    if bmx_data:
        products, bmx_active_order = merge_products_with_bmx(products, bmx_data)

    if not products:
        logger.warning("没有数据，跳过格式化")
        return

    # 所有门店统一过滤和排序
    all_store_order = store_order + bmx_active_order
    active_products = [
        p for p in products
        if any((s.get("数量") or 0) != 0 or (s.get("金额") or 0) != 0 for s in p["stores"].values())
    ]
    sorted_products = sorted(active_products, key=_sort_key_s1)

    # 构建品牌映射（用于 S3 sheet 标题）
    brand_map = {}
    for name in bmx_active_order:
        brand_map[name] = "焙满香"

    logger.info(f"加载模板: {template_file}")
    wb = load_workbook(template_file, rich_text=True)

    logger.info(f"填充 Sheet 1 (货品对比表), {len(sorted_products)} 条商品, {len(all_store_order)} 个门店...")
    fill_s1(wb, sorted_products, all_store_order, monday, sunday)

    logger.info("填充 Sheet 2 (货品销售排行表)...")
    s2_sorted = sorted(active_products, key=lambda p: -sum((s.get("金额") or 0) for s in p["stores"].values()))
    fill_s2(wb, s2_sorted, monday, sunday)

    logger.info(f"填充 Sheet 3+ ({len(all_store_order)} 个门店销售排行)...")
    fill_s3_sheets(wb, sorted_products, all_store_order, monday, sunday, brand_map)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    logger.info(f"已生成格式化文件: {output_file}")


# ─── 浏览器自动化函数 ────────────────────────────────────────────────────────


def set_date(page, placeholder, value):
    page.evaluate(f"""
        (function() {{
            var inp = document.querySelector('input.timeInput.hasDatepicker[placeholder="{placeholder}"]');
            if (!inp) return 'not found';
            inp.value = "{value}";
            inp.dispatchEvent(new Event('input',  {{bubbles: true}}));
            inp.dispatchEvent(new Event('change', {{bubbles: true}}));
            inp.dispatchEvent(new Event('blur',   {{bubbles: true}}));
            return inp.value;
        }})()
    """)


def click_by_text(page, text, desc=""):
    result = page.evaluate(f"""
        (function() {{
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {{
                if (els[i].textContent.trim() === "{text}" && els[i].children.length === 0) {{
                    els[i].click();
                    return 'clicked tag=' + els[i].tagName + ' class=' + els[i].className;
                }}
            }}
            return 'not found';
        }})()
    """)
    tag = f"[{desc}] " if desc else ""
    logger.info(f"  {tag}→ {result}")
    return "clicked" in result


def login(page):
    logger.info("[1/4] 打开登录页面...")
    page.goto(LOGIN_URL)
    page.wait_for_load_state("networkidle")

    logger.info("[2/4] 切换到工号登录模式...")
    page.evaluate("""
        (function() {
            var els = document.querySelectorAll('div,span,a,button');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '工号登录' && els[i].children.length === 0) {
                    els[i].click();
                    return 'clicked';
                }
            }
            return 'not found';
        })()
    """)
    time.sleep(1.5)

    placeholder = page.evaluate("""
        (function() {
            var inputs = document.querySelectorAll('input');
            var r = [];
            for (var i = 0; i < inputs.length; i++) {
                if (inputs[i].placeholder) r.push(inputs[i].placeholder);
            }
            return r.join(' | ');
        })()
    """)
    if "员工工号" not in placeholder:
        raise RuntimeError(f"工号登录切换失败，当前输入框：{placeholder}")
    logger.info(f"  表单已切换 → {placeholder}")

    logger.info("[3/4] 填入账号/工号/密码...")
    page.evaluate(f"""
        (function() {{
            var inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].placeholder === '请输入账号')     inputs[i].value = "{ACCOUNT}";
                if (inputs[i].placeholder === '请输入员工工号') inputs[i].value = "{WORKER_ID}";
                if (inputs[i].placeholder === '请输入工号密码') inputs[i].value = "{PASSWORD}";
            }}
            inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].value) {{
                    inputs[i].dispatchEvent(new Event('input', {{bubbles: true}}));
                }}
            }}
        }})()
    """)

    time.sleep(1.5)

    logger.info("[4/4] 点击登录按钮...")
    click_by_text(page, "登 录", "登录")
    page.wait_for_load_state("networkidle", timeout=120_000)

    time.sleep(2)


# ─── 主函数 ────────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(description="Pospal 工厂配送麦安研门店周度销售报表自动化脚本")
    parser.add_argument("--weeks", type=int, default=0, help="周偏移量：0=本周，-1=上周（默认0）")
    parser.add_argument("--date", type=str, help="指定日期推断所在周，格式 YYYY.MM.DD")
    parser.add_argument("--headless", action="store_true", help="无头模式（不显示浏览器窗口）")
    args = parser.parse_args()

    if args.date:
        monday, sunday = get_week_range(date_str=args.date)
    else:
        monday, sunday = get_week_range(weeks=args.weeks)

    monday_str = monday.strftime("%Y.%m.%d")
    sunday_str = sunday.strftime("%Y.%m.%d")
    date_range_str = f"{monday.strftime('%Y-%m-%d')}~{sunday.strftime('%Y-%m-%d')}"

    logger.info(f"{'=' * 55}")
    logger.info(f"  Pospal 工厂配送麦安研门店周度销售报表")
    logger.info(f"  目标周：{monday_str} ~ {sunday_str}")
    logger.info(f"  输出根目录：{OUTPUT_DIR}")
    logger.info(f"{'=' * 55}")

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        logger.error("未安装 playwright，请先运行: pip install playwright && playwright install chromium")
        sys.exit(1)

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=args.headless,
            args=["--start-maximized"],
        )
        context = browser.new_context(
            viewport={"width": 1600, "height": 900},
            accept_downloads=True,
        )
        page = context.new_page()
        page.set_default_timeout(120000)
        page.set_default_navigation_timeout(120000)

        try:
            login(page)

            # ── 步骤 1：下载仓库配送商品门店对比表 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：下载仓库配送商品门店对比表")
            logger.info(f"{'─' * 55}")

            logger.info("  [导航] 前往仓库配送商品门店对比表...")
            page.goto(REPORT_URL)
            page.wait_for_load_state("networkidle", timeout=120_000)
            logger.info(f"  已到达 → {page.url}")

            logger.info(f"  → 设置时间: {monday_str} ~ {sunday_str}...")
            set_date(page, "开始日期", f"{monday_str} 00:00")
            set_date(page, "结束日期", f"{sunday_str} 23:59")

            logger.info("  → 点击更多搜索...")
            more_search_result = page.evaluate("""
                (function() {
                    var btn = document.getElementById('advancedBtn');
                    if (btn) { btn.click(); return 'clicked advancedBtn'; }
                    var els = document.querySelectorAll('*');
                    for (var i = 0; i < els.length; i++) {
                        var t = els[i].textContent.trim();
                        if ((t === '更多搜索' || t === '更多搜索▼') && els[i].children.length === 0) {
                            els[i].click();
                            return 'clicked tag=' + els[i].tagName + ' class=' + els[i].className;
                        }
                    }
                    return 'not found';
                })()
            """)
            logger.info(f"  [更多搜索] → {more_search_result}")
            time.sleep(0.5)

            logger.info("  → 打开分类选择弹框...")
            select_cat_result = page.evaluate("""
                (function() {
                    var btn = document.getElementById('selectCategory');
                    if (btn) { btn.click(); return 'clicked selectCategory'; }
                    var els = document.querySelectorAll('*');
                    for (var i = 0; i < els.length; i++) {
                        if (els[i].textContent.trim() === '选择分类' && els[i].children.length === 0) {
                            els[i].click();
                            return 'clicked tag=' + els[i].tagName;
                        }
                    }
                    return 'not found';
                })()
            """)
            logger.info(f"  [选择分类] → {select_cat_result}")
            time.sleep(1.0)

            logger.info(f"  → 勾选 {len(TARGET_CATEGORIES)} 个分类...")
            categories_js = str(TARGET_CATEGORIES).replace("'", '"')
            page.evaluate(f"""
                (function() {{
                    var targets = {categories_js};
                    var divs = document.querySelectorAll('.checkBoxDiv');
                    divs.forEach(function(d) {{
                        var span = d.querySelector('span');
                        if (span && targets.indexOf(span.textContent.trim()) >= 0) {{
                            if (!d.classList.contains('on')) {{
                                d.click();
                            }}
                        }}
                    }});
                }})()
            """)

            check_result = page.evaluate(f"""
                (function() {{
                    var targets = {categories_js};
                    var divs = document.querySelectorAll('.checkBoxDiv');
                    var r = [];
                    divs.forEach(function(d) {{
                        var s = d.querySelector('span');
                        if (s && targets.indexOf(s.textContent.trim()) >= 0) {{
                            r.push(s.textContent.trim() + ':' + d.classList.contains('on'));
                        }}
                    }});
                    return r.join(' | ');
                }})()
            """)
            logger.info(f"    勾选验证: {check_result}")

            if "false" in check_result:
                logger.warning("部分分类未勾选，尝试重试...")
                page.evaluate(f"""
                    (function() {{
                        var targets = {categories_js};
                        var divs = document.querySelectorAll('.checkBoxDiv');
                        divs.forEach(function(d) {{
                            var span = d.querySelector('span');
                            if (span && targets.indexOf(span.textContent.trim()) >= 0 && !d.classList.contains('on')) {{
                                d.click();
                            }}
                        }});
                    }})()
                """)

            logger.info("  → 点击「确定」关闭弹框...")
            click_by_text(page, "确定", "确定弹框")
            time.sleep(0.5)

            logger.info("  [查询] 执行查询...")
            click_by_text(page, "查询", "查询")
            page.wait_for_load_state("networkidle", timeout=150_000)
            time.sleep(3)

            logger.info("  [导出] 导出文件...")
            output_dir = OUTPUT_DIR / date_range_str / "原始下载"
            output_dir.mkdir(parents=True, exist_ok=True)

            with page.expect_download(timeout=180_000) as dl_info:
                result = page.evaluate("""
                    (function() {
                        var els = document.querySelectorAll('*');
                        for (var i = 0; i < els.length; i++) {
                            if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                                els[i].click();
                                return 'ok';
                            }
                        }
                        return 'not found';
                    })()
                """)
                logger.info(f"  点击导出: {result}")
                if result == "not found":
                    raise RuntimeError("未找到「导出」按钮")

            download = dl_info.value
            logger.info(f"  下载文件名: {download.suggested_filename}")

            dest = output_dir / f"仓库配送商品门店对比表_{date_range_str}.xlsx"
            download.save_as(dest)
            logger.info(f"  已保存到: {dest}")

            # ── 步骤 2：下载仓库配送商品大客户对比表 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：下载仓库配送商品大客户对比表")
            logger.info(f"{'─' * 55}")

            logger.info("  [导航] 前往仓库配送商品大客户对比表...")
            page.goto(DELIVERY_PRODUCT_COMPARISON_URL)
            page.wait_for_load_state("networkidle", timeout=120_000)
            logger.info(f"  已到达 → {page.url}")

            logger.info(f"  → 设置时间: {monday_str} ~ {sunday_str}...")
            set_date(page, "开始日期", f"{monday_str} 00:00")
            set_date(page, "结束日期", f"{sunday_str} 23:59")

            logger.info("  [查询] 执行查询...")
            click_by_text(page, "查询", "查询")
            page.wait_for_load_state("networkidle", timeout=150_000)
            time.sleep(3)

            logger.info("  [导出] 导出文件...")
            with page.expect_download(timeout=180_000) as dl_info2:
                result = page.evaluate("""
                    (function() {
                        var els = document.querySelectorAll('*');
                        for (var i = 0; i < els.length; i++) {
                            if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                                els[i].click();
                                return 'ok';
                            }
                        }
                        return 'not found';
                    })()
                """)
                logger.info(f"  点击导出: {result}")
                if result == "not found":
                    raise RuntimeError("未找到「导出」按钮")

            download2 = dl_info2.value
            logger.info(f"  下载文件名: {download2.suggested_filename}")

            dest2 = output_dir / f"仓库配送商品大客户对比表_{date_range_str}.xlsx"
            download2.save_as(dest2)
            logger.info(f"  已保存到: {dest2}")

            # ── 步骤 3：格式化数据并生成报表 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：格式化数据并生成报表")
            logger.info(f"{'─' * 55}")

            logger.info("  读取焙满香门店数据...")
            bmx_data = read_bmx_store_data(dest2)

            formatted_output = OUTPUT_DIR / date_range_str / f"工厂配送麦安研门店周度销售报表_{date_range_str}.xlsx"
            merge_into_template(dest, TEMPLATE_FILE, formatted_output, monday, sunday, bmx_data)

            logger.info(f"{'=' * 55}")
            logger.info(f"  工厂配送麦安研门店周度销售报表全部完成！")
            logger.info(f"  仓库配送商品门店对比表 → {dest}")
            logger.info(f"  仓库配送商品大客户对比表 → {dest2}")
            logger.info(f"  格式化输出 → {formatted_output}")
            logger.info(f"{'=' * 55}\n")

        except Exception as e:
            logger.error(f"任务失败: {e}")
            try:
                screenshot = OUTPUT_DIR / "error_screenshot.png"
                page.screenshot(path=str(screenshot))
                logger.info(f"  错误截图已保存: {screenshot}")
            except Exception:
                pass
            raise
        finally:
            context.close()
            browser.close()


if __name__ == "__main__":
    main()
