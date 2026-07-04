"""
麦安研+焙满香门店配货差异明细表 - 自动化脚本
=====================================
依赖：pip install playwright openpyxl && playwright install chromium

自动登录 Pospal 后台，导出订货配货明细并生成格式化报表。

用法：
    python mainyan_store_distribution_diff_detail.py                    # 导出今天的数据
    python mainyan_store_distribution_diff_detail.py --days -1          # 导出昨天的数据
    python mainyan_store_distribution_diff_detail.py --date 2026.06.20  # 指定日期
    python mainyan_store_distribution_diff_detail.py --headless         # 无头模式（不显示浏览器）
"""

import argparse
import copy
import logging
import logging.handlers
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ─── 日志配置 ───────────────────────────────────────────────────────────────────
LOG_DIR = Path(__file__).resolve().parent / "log"
LOG_DIR.mkdir(exist_ok=True)

_file_handler = logging.handlers.TimedRotatingFileHandler(
    LOG_DIR / "mainyan_store_distribution_diff_detail.log",
    when="midnight",
    backupCount=30,
    encoding="utf-8",
)
_file_handler.suffix = "%Y-%m-%d"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[_file_handler, logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# ─── 配置区（按需修改）─────────────────────────────────────────────────────────
ACCOUNT = "huomimayzb"
WORKER_ID = "M006"
PASSWORD = "tusijia88"

LOGIN_URL = "https://beta69.pospal.cn/"
REPORT_URL = "https://css69.pospal.cn/ChainStoreSupplyReport/ProductRequestItem"

OUTPUT_DIR = Path(__file__).resolve().parent / "麦安研+焙满香门店配货差异明细表"
TEMPLATE_FILE = Path(__file__).resolve().parent / "麦安研+焙满香门店配货差异明细表_格式化模板.xlsx"

STORE_NAME_MAP = {
    "麦安研（顺德杏坛店）": "杏坛",
    "麦安研（东站宝泰店）": "宝泰",
    "麦安研（佛山创产店）": "创产",
    "麦安研（顺德龙江店）": "龙江",
    "麦安研（佛山万民金海城店）": "金海城",
    "麦安研门店裱花间": "裱花间",
    "中央工厂冷加工间": "冷加工间",
}


# ─── 工具函数 ──────────────────────────────────────────────────────────────────

def set_date(page, placeholder: str, value: str):
    """设置日期输入框的值并触发必要事件"""
    page.evaluate(f"""
        (function() {{
            var inp = document.querySelector('input.timeInput.hasDatepicker[placeholder="{placeholder}"]');
            if (!inp) return 'not found';
            inp.value = "{value}";
            inp.dispatchEvent(new Event('input',  {{bubbles: true}}));
            inp.dispatchEvent(new Event('change', {{bubbles: true}}));
            inp.dispatchEvent(new Event('blur',   {{bubbles: true}}));
            return inp.value;
        }})()
    """)


def verify_dates(page, expected_date: str) -> bool:
    """验证两个日期输入框的值是否都包含期望日期"""
    result = page.evaluate("""
        (function() {
            var r = [];
            document.querySelectorAll('input.timeInput.hasDatepicker').forEach(function(inp) {
                r.push(inp.placeholder + '=' + inp.value);
            });
            return r.join(' | ');
        })()
    """)
    logger.info(f"  [日期验证] {result}")
    return expected_date in result and result.count(expected_date) == 2


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def read_downloaded_data(data_file):
    """读取订货配货明细，返回按门店分组的数据字典。

    返回: {门店全称: [产品行列表]}
    """
    wb = load_workbook(data_file, data_only=True)
    ws = wb.active
    stores_data = {}

    for r in range(2, ws.max_row + 1):
        store_name = ws.cell(row=r, column=4).value  # D: 订货组织
        if not store_name or store_name not in STORE_NAME_MAP:
            continue

        product_name = ws.cell(row=r, column=11).value  # K: 商品名称
        if not product_name:
            continue

        category = ws.cell(row=r, column=12).value  # L: 商品分类
        spec = ws.cell(row=r, column=15).value  # O: 规格
        unit = ws.cell(row=r, column=16).value  # P: 单位
        delivery_rate = ws.cell(row=r, column=20).value  # T: 配货率
        order_qty = ws.cell(row=r, column=17).value  # Q: 订货量
        delivered_qty = ws.cell(row=r, column=18).value  # R: 已配货量
        order_price = ws.cell(row=r, column=23).value  # W: 订货价
        order_amount = ws.cell(row=r, column=24).value  # X: 订货金额
        delivered_amount = ws.cell(row=r, column=25).value  # Y: 已配货金额

        # 转换数值类型
        def to_float(val):
            if val is None or val == '':
                return None
            if isinstance(val, (int, float)):
                return float(val)
            s = str(val).strip().replace('%', '').replace(',', '')
            try:
                return float(s) / 100 if '%' in str(val) else float(s)
            except:
                return None

        product = {
            "商品名称": product_name,
            "商品分类": category,
            "规格": spec,
            "单位": unit,
            "配货率": to_float(delivery_rate),
            "订货量": to_float(order_qty),
            "已配货量": to_float(delivered_qty),
            "订货价": to_float(order_price),
            "订货金额": to_float(order_amount),
            "已配货金额": to_float(delivered_amount),
        }

        if store_name not in stores_data:
            stores_data[store_name] = []
        stores_data[store_name].append(product)

    for store_name, products in stores_data.items():
        logger.info(f"  {store_name}: {len(products)} 条记录")

    return stores_data


def sort_products(products):
    """排序：非100%配货率在前，按配货率从高到低排列；100%的在后。"""

    def sort_key(p):
        rate = p["配货率"]
        if rate is None:
            return (1, 0)
        # 兼容小数格式(0.8=80%)和百分比格式(80=80%)
        full = (rate >= 100) if rate > 1 else (rate >= 1.0)
        if full:
            return (1, 0)
        return (0, -rate)  # 非100%，数值从高到低

    return sorted(products, key=sort_key)


def click_by_text(page, text, desc=""):
    result = page.evaluate(f"""
        (function() {{
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {{
                if (els[i].textContent.trim() === "{text}" && els[i].children.length === 0) {{
                    els[i].click();
                    return 'clicked tag=' + els[i].tagName + ' class=' + els[i].className;
                }}
            }}
            return 'not found';
        }})()
    """)
    tag = f"[{desc}] " if desc else ""
    logger.info(f"  {tag}→ {result}")
    return "clicked" in result


def login(page):
    logger.info("[1/4] 打开登录页面...")
    page.goto(LOGIN_URL)
    page.wait_for_load_state("networkidle")

    logger.info("[2/4] 切换到工号登录模式...")
    page.evaluate("""
        (function() {
            var els = document.querySelectorAll('div,span,a,button');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '工号登录' && els[i].children.length === 0) {
                    els[i].click();
                    return 'clicked';
                }
            }
            return 'not found';
        })()
    """)
    time.sleep(1.5)

    placeholder = page.evaluate("""
        (function() {
            var inputs = document.querySelectorAll('input');
            var r = [];
            for (var i = 0; i < inputs.length; i++) {
                if (inputs[i].placeholder) r.push(inputs[i].placeholder);
            }
            return r.join(' | ');
        })()
    """)
    if "员工工号" not in placeholder:
        raise RuntimeError(f"工号登录切换失败，当前输入框：{placeholder}")
    logger.info(f"  表单已切换 → {placeholder}")

    logger.info("[3/4] 填入账号/工号/密码...")
    page.evaluate(f"""
        (function() {{
            var inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].placeholder === '请输入账号')     inputs[i].value = "{ACCOUNT}";
                if (inputs[i].placeholder === '请输入员工工号') inputs[i].value = "{WORKER_ID}";
                if (inputs[i].placeholder === '请输入工号密码') inputs[i].value = "{PASSWORD}";
            }}
            inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].value) {{
                    inputs[i].dispatchEvent(new Event('input', {{bubbles: true}}));
                }}
            }}
        }})()
    """)

    time.sleep(1.5)

    logger.info("[4/4] 点击登录按钮...")
    click_by_text(page, "登 录", "登录")
    page.wait_for_load_state("networkidle", timeout=120_000)

    time.sleep(2)


def merge_into_template(stores_data, template_file, output_file, target_date):
    """将分门店数据填入模板并生成输出文件。"""
    wb_template = load_workbook(template_file)
    wb_output = load_workbook(template_file)

    # 删除输出工作簿中的所有sheet
    for ws_name in wb_output.sheetnames:
        del wb_output[ws_name]

    if not stores_data:
        logger.warning("  无数据，跳过生成报表")
        wb_template.close()
        wb_output.close()
        return

    date_str = target_date.strftime("%Y年%#m月%#d日") if sys.platform == "win32" else target_date.strftime(
        "%Y年%-m月%-d日")

    for store_full_name, products in stores_data.items():
        store_short = STORE_NAME_MAP.get(store_full_name, store_full_name)

        # 复制模板的第一个sheet
        ws_template = wb_template.worksheets[0]
        ws_new = wb_output.create_sheet(title=store_short)

        # 复制列宽
        for col_idx in range(1, 12):
            col_letter = get_column_letter(col_idx)
            ws_new.column_dimensions[col_letter].width = ws_template.column_dimensions[col_letter].width

        # 复制行高和表头
        for row_idx in range(1, 3):
            ws_new.row_dimensions[row_idx].height = ws_template.row_dimensions[row_idx].height
            for col_idx in range(1, 12):
                src_cell = ws_template.cell(row=row_idx, column=col_idx)
                dst_cell = ws_new.cell(row=row_idx, column=col_idx)
                dst_cell.value = src_cell.value
                copy_cell_style(src_cell, dst_cell)

        # 更新标题中的日期和门店
        title_cell = ws_new.cell(row=1, column=1)
        if title_cell.value:
            title_cell.value = f"{date_str}{store_short}配货明细表"

        # 排序产品
        sorted_products = sort_products(products)
        data_row_count = len(sorted_products)

        # 复制合计行（第3行）
        ws_new.row_dimensions[3].height = ws_template.row_dimensions[3].height
        for col_idx in range(1, 12):
            src_cell = ws_template.cell(row=3, column=col_idx)
            dst_cell = ws_new.cell(row=3, column=col_idx)
            dst_cell.value = src_cell.value
            copy_cell_style(src_cell, dst_cell)

        # 更新合计行公式
        if data_row_count > 0:
            ws_new.cell(row=3, column=7).value = f"=SUM(G4:G{3 + data_row_count})"
            ws_new.cell(row=3, column=8).value = f"=SUM(H4:H{3 + data_row_count})"
            ws_new.cell(row=3, column=10).value = f"=SUM(J4:J{3 + data_row_count})"
            ws_new.cell(row=3, column=11).value = f"=SUM(K4:K{3 + data_row_count})"

        # 填充数据行
        for idx, product in enumerate(sorted_products):
            row_idx = 4 + idx

            # 使用模板第4行（偶数样式）或第5行（奇数样式）作为样式参考
            template_row = 4 if idx % 2 == 0 else 5

            # 复制行高
            ws_new.row_dimensions[row_idx].height = ws_template.row_dimensions[template_row].height

            # 填充数据并复制样式
            order_qty = product["订货量"] or 0
            delivered_qty = product["已配货量"] or 0
            unit = product["单位"] or ""

            # 计算备货备注
            if order_qty == delivered_qty:
                remark = "-"
            elif order_qty < delivered_qty:
                remark = f"多送{int(delivered_qty - order_qty)}{unit}"
            else:
                remark = f"少送{int(order_qty - delivered_qty)}{unit}"

            data_map = [
                (1, product["商品名称"]),
                (2, product["商品分类"]),
                (3, product["规格"]),
                (4, product["单位"]),
                (5, product["配货率"]),
                (6, remark),
                (7, product["订货量"]),
                (8, product["已配货量"]),
                (9, product["订货价"]),
                (10, product["订货金额"]),
                (11, product["已配货金额"]),
            ]

            for col_idx, value in data_map:
                src_cell = ws_template.cell(row=template_row, column=col_idx)
                dst_cell = ws_new.cell(row=row_idx, column=col_idx)
                copy_cell_style(src_cell, dst_cell)
                dst_cell.value = value

        logger.info(f"  生成 {store_short} sheet，{data_row_count} 行数据")

    wb_template.close()

    # 保存输出文件
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb_output.save(output_file)
    wb_output.close()

    logger.info(f"  输出文件: {output_file}")


# ─── Playwright 自动化 ─────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="麦安研+焙满香门店配货差异明细表自动导出")
    parser.add_argument("--date", help="指定日期 (如 2026.06.20)")
    parser.add_argument("--days", type=int, default=0, help="相对今天的天数偏移")
    parser.add_argument("--headless", action="store_true", help="无头模式")
    args = parser.parse_args()

    if args.date:
        target_date = datetime.strptime(args.date, "%Y.%m.%d")
    else:
        target_date = datetime.now() + timedelta(days=args.days)

    date_str = target_date.strftime("%Y-%m-%d")
    date_display = target_date.strftime("%Y.%m.%d")

    logger.info(f"{'=' * 55}")
    logger.info(f"  麦安研+焙满香门店配货差异明细表 - {date_str}")
    logger.info(f"{'=' * 55}")

    output_dir = OUTPUT_DIR / date_str
    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        logger.error("未安装 Playwright。请运行：pip install playwright && playwright install chromium")
        sys.exit(1)

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=args.headless,
            args=["--start-maximized"],
        )
        context = browser.new_context(
            viewport={"width": 1600, "height": 900},
            accept_downloads=True,
        )
        page = context.new_page()
        page.set_default_timeout(120000)
        page.set_default_navigation_timeout(120000)

        try:
            login(page)

            # ── 步骤 2：导出订货配货明细 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：导出订货配货明细")
            logger.info(f"{'─' * 55}")
            page.goto(REPORT_URL)
            page.wait_for_load_state("networkidle")
            time.sleep(2)

            # 选择"期望到货时间"
            logger.info("  选择时间类型：期望到货时间")
            page.evaluate("""
                          () => {
                              const els = document.querySelectorAll('[p-model="timeType"] li');
                              for (let i = 0; i < els.length; i++) {
                                  if (els[i].textContent.includes('期望到货时间')) {
                                      els[i].click();
                                      return 'ok';
                                  }
                              }
                              return 'not found';
                          }
                          """)

            # 设置日期范围
            logger.info(f"  设置日期: {date_display}")
            set_date(page, "开始日期", f"{date_display} 00:00")
            set_date(page, "结束日期", f"{date_display} 23:59")

            page.click(".submitBtn")
            page.wait_for_load_state("networkidle")
            time.sleep(2)

            logger.info(f"  设置日期: {date_display}")
            set_date(page, "开始日期", f"{date_display} 00:00")
            set_date(page, "结束日期", f"{date_display} 23:59")

            # 点击查询按钮，验证日期并重试
            logger.info("  点击查询")
            for attempt in range(1, 4):
                logger.info(f"  查询第 {attempt} 次...")
                page.click(".submitBtn")
                page.wait_for_load_state("networkidle")
                time.sleep(2)

                if verify_dates(page, date_display):
                    logger.info(f"  ✅ 查询成功，日期正确")
                    break
                else:
                    logger.warning("  日期被重置！重新设置日期...")
                    set_date(page, "开始日期", f"{date_display} 00:00")
                    set_date(page, "结束日期", f"{date_display} 23:59")
            else:
                raise RuntimeError("查询 3 次后日期仍然不正确")

            # 点击导出按钮
            logger.info("  点击导出")
            with page.expect_download(timeout=60000) as dl_info:
                result = page.evaluate("""
                                       (() => {
                                           const els = document.querySelectorAll('#btn_export');
                                           for (let i = 0; i < els.length; i++) {
                                               els[i].click();
                                               return 'ok';
                                           }
                                           return 'not found';
                                       })()
                                       """)
                logger.info(f"  点击导出: {result}")
                if result == "not found":
                    raise RuntimeError("未找到「导出」按钮")

            download = dl_info.value
            logger.info(f"  下载文件名: {download.suggested_filename}")

            raw_dir = output_dir / "原始下载"
            raw_dir.mkdir(parents=True, exist_ok=True)
            dest = raw_dir / f"订货配货明细_{date_str}.xlsx"
            download.save_as(dest)
            logger.info(f"  已保存到: {dest}")

            # ── 步骤 3：格式化数据并生成报表 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：格式化数据并生成报表")
            logger.info(f"{'─' * 55}")

            stores_data = read_downloaded_data(dest)
            formatted_output = OUTPUT_DIR / date_str / f"麦安研+焙满香门店配货差异明细表_{date_str}.xlsx"
            merge_into_template(stores_data, TEMPLATE_FILE, formatted_output, target_date)

            logger.info(f"{'=' * 55}")
            logger.info(f"  麦安研+焙满香门店配货差异明细表全部完成！")
            logger.info(f"  下载文件 → {dest}")
            logger.info(f"  格式化输出 → {formatted_output}")
            logger.info(f"{'=' * 55}\n")

        except Exception as e:
            logger.error(f"任务失败: {e}")
            try:
                screenshot = OUTPUT_DIR / "error_screenshot.png"
                page.screenshot(path=str(screenshot))
                logger.info(f"  错误截图已保存: {screenshot}")
            except Exception:
                pass
            raise
        finally:
            context.close()
            browser.close()


if __name__ == "__main__":
    main()
