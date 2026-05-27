"""
订货商品汇总看板 - 自动导出 & 格式化脚本
=====================================
依赖：pip install playwright openpyxl && playwright install chromium

自动登录后依次导出：
  1. 订货商品汇总看板 (ProductRequestSummaryBoard)
  2. 订货商品明细看板 (ProductRequestItemBoard)
  3. 将下载数据填入格式模板，生成格式化汇总看板

用法：
    python product_request_export.py                    # 导出后天的数据
    python product_request_export.py --date 2026.05.30  # 指定日期
    python product_request_export.py --days 2           # N天后（默认2=后天）
    python product_request_export.py --headless         # 无头模式（不显示浏览器）
"""

import argparse
import copy
import logging
import logging.handlers
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ─── 日志配置 ───────────────────────────────────────────────────────────────────
LOG_DIR = Path(__file__).resolve().parent / "log"
LOG_DIR.mkdir(exist_ok=True)

_file_handler = logging.handlers.TimedRotatingFileHandler(
    LOG_DIR / "product_request_export.log",
    when="midnight",
    backupCount=30,
    encoding="utf-8",
)
_file_handler.suffix = "%Y-%m-%d"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[_file_handler, logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# ─── 配置区（按需修改）─────────────────────────────────────────────────────────
ACCOUNT = "huomimayzb"
WORKER_ID = "M006"
PASSWORD = "tusijia88"

LOGIN_URL = "https://beta69.pospal.cn/"
SUMMARY_BOARD_URL = "https://css69.pospal.cn/ChainStoreSupplySeller/ProductRequestSummaryBoard"
ITEM_BOARD_URL = "https://css69.pospal.cn/ChainStoreSupplySeller/ProductRequestItemBoard"

OUTPUT_DIR = Path(__file__).resolve().parent / "订货商品汇总看板"
TEMPLATE_FILE = Path(__file__).resolve().parent / "订货商品汇总看板_格式化模板.xlsx"

STATUS_OPTIONS = [
    "待审核",
    "配货中",
    "已配货",
    "已收货",
    "已拒绝",
    "已拒绝出库",
    "已拒绝收货",
    "已作废",
]

TARGET_CATEGORIES = [
    "配送费",
    "包材耗材",
    "工衣模具",
    "慕斯+饼干+饮品+其他",
    "原料铺料",
    "冷冻面团",
    "蛋糕及面包成品及饼干类",
]

# ─── 格式化模板配置 ──────────────────────────────────────────────────────────
ROW_CATEGORY_MAP = {
    2: ["冷冻面团"],
    3: ["成品面包类", "蛋糕类", "饼干类"],
    4: ["热销类", "冷冻肉类", "冷冻馅料类", "冷藏馅料类", "油脂类", "粉类", "糖类", "常温馅料类", "干果类", "饼干类/外",
        "慕斯类/外", "饮品类/外", "其他/外", "专版包材类", "公版包材类", "工衣工帽围裙", "模具", "保洁用品"],
    5: ["配送费"],
}

EXPORT_CATEGORY_MAP = {
    "面团": ["冷冻面团"],
    "成品面包饼干": ["成品面包类", "饼干类"],
    "蛋糕": ["蛋糕类"],
    "物料包材": ["热销类", "冷冻肉类", "冷冻馅料类", "冷藏馅料类", "油脂类", "粉类", "糖类", "常温馅料类", "干果类", "饼干类/外",
         "慕斯类/外", "饮品类/外", "其他/外", "专版包材类", "公版包材类", "工衣工帽围裙", "模具", "保洁用品", "配送费"],
}

SAMPLE_ROW_START = 8
SAMPLE_ROW_COUNT = 2
DATA_START_ROW = 8
HEADER_ROW = 7
TOTAL_ORDER_HEADER = "合计订货量"

FIXED_COLUMN_MAP = [
    (1, 2),  # 商品分类 -> B
    (3, 3),  # 商品名称 -> C
    (4, 4),  # 规格 -> D
    (5, 5),  # 单位 -> E
]


# ─── 格式化合并函数 ──────────────────────────────────────────────────────────


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def compute_category_sums(detail_file: Path, store_names: set[str]) -> dict:
    wb = load_workbook(detail_file, data_only=True)
    ws = wb.active

    cat_to_row = {}
    for row_num, categories in ROW_CATEGORY_MAP.items():
        for cat in categories:
            cat_to_row[cat] = row_num

    sums: dict[tuple[int, str], float] = {}
    for row in ws.iter_rows(min_row=2):
        category = row[1].value  # B列：商品分类
        org = row[5].value  # F列：订货组织
        amount = row[13].value  # N列：订货金额

        if category is None or org is None or amount is None:
            continue

        category_str = str(category).strip()
        org_str = str(org).strip()

        row_num = cat_to_row.get(category_str)
        if row_num is None:
            continue
        if org_str not in store_names:
            continue

        key = (row_num, org_str)
        sums[key] = sums.get(key, 0) + float(amount)

    wb.close()
    return sums


def compute_detail_sums_by_category(detail_file: Path, store_names: set[str]) -> dict:
    wb = load_workbook(detail_file, data_only=True)
    ws = wb.active
    sums: dict[tuple[str, str], float] = {}
    for row in ws.iter_rows(min_row=2):
        category = row[1].value
        org = row[5].value
        amount = row[13].value
        if category is None or org is None or amount is None:
            continue
        cat_str = str(category).strip()
        org_str = str(org).strip()
        if org_str not in store_names:
            continue
        key = (cat_str, org_str)
        sums[key] = sums.get(key, 0) + float(amount)
    wb.close()
    return sums


def aggregate_to_row_sums(detail_sums: dict, allowed_categories: set[str] | None) -> dict:
    cat_to_row = {}
    for row_num, categories in ROW_CATEGORY_MAP.items():
        for cat in categories:
            cat_to_row[cat] = row_num
    sums: dict[tuple[int, str], float] = {}
    for (cat, store), amount in detail_sums.items():
        if allowed_categories is not None and cat not in allowed_categories:
            continue
        row_num = cat_to_row.get(cat)
        if row_num is None:
            continue
        key = (row_num, store)
        sums[key] = sums.get(key, 0) + amount
    return sums


def read_data_file(data_file: Path):
    wb = load_workbook(data_file)
    ws = wb.active

    headers = [cell.value for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]

    total_col_idx = None
    for i, h in enumerate(headers):
        if h and TOTAL_ORDER_HEADER in str(h):
            total_col_idx = i
            break
    if total_col_idx is None:
        raise ValueError(f"数据源中找不到'{TOTAL_ORDER_HEADER}'列")

    store_columns = []
    for i in range(total_col_idx + 1, len(headers)):
        if headers[i] is not None and str(headers[i]).strip():
            store_columns.append((i, str(headers[i]).strip()))

    rows = []
    for row in ws.iter_rows(min_row=2):
        vals = [cell.value for cell in row]
        if any(v is not None and str(v).strip() != "" for v in vals):
            rows.append(vals)

    return total_col_idx, store_columns, rows


def merge_into_template(data_rows, total_col_idx, store_columns, template_file, output_file,
                        category_sums=None, target_date=None):
    wb = load_workbook(template_file)
    ws = wb.active

    store_count = len(store_columns)
    last_col = 6 + store_count
    template_max_col = ws.max_column

    column_map = list(FIXED_COLUMN_MAP)
    column_map.append((total_col_idx, 6))
    for offset, (src_idx, _) in enumerate(store_columns):
        column_map.append((src_idx, 7 + offset))

    for offset, (_, name) in enumerate(store_columns):
        ws.cell(row=HEADER_ROW, column=7 + offset).value = name

    for col in range(last_col + 1, template_max_col + 1):
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.font = copy.copy(cell.font)
            cell.border = copy.copy(cell.border)

    if last_col > template_max_col:
        for row in range(1, ws.max_row + 1):
            src = ws.cell(row=row, column=template_max_col)
            for new_col in range(template_max_col + 1, last_col + 1):
                copy_cell_style(src, ws.cell(row=row, column=new_col))

    first_store = get_column_letter(7)
    last_store = get_column_letter(last_col)
    for r in range(2, 6):
        ws.cell(row=r, column=6).value = f"=SUM({first_store}{r}:{last_store}{r})"
    ws.cell(row=6, column=6).value = f"=SUM({first_store}6:{last_store}6)"

    sample_cells = list(ws.iter_rows(
        min_row=SAMPLE_ROW_START,
        max_row=SAMPLE_ROW_START
    ))[0]
    sample_row_height = ws.row_dimensions[SAMPLE_ROW_START].height

    # 创建分类优先级映射，按照ROW_CATEGORY_MAP中定义的顺序
    category_priority = {}
    priority = 0
    for row_num, categories in ROW_CATEGORY_MAP.items():
        for cat in categories:
            category_priority[cat] = priority
            priority += 1

    # 对数据行按照分类在ROW_CATEGORY_MAP中的顺序进行排序
    def get_category_priority(row_data):
        category = str(row_data[1]).strip() if row_data[1] is not None else ""
        return category_priority.get(category, float('inf'))

    sorted_data_rows = sorted(data_rows, key=get_category_priority)

    ws.delete_rows(SAMPLE_ROW_START, SAMPLE_ROW_COUNT)
    ws.insert_rows(DATA_START_ROW, len(sorted_data_rows))

    for i, row_data in enumerate(sorted_data_rows):
        excel_row = DATA_START_ROW + i
        ws.cell(row=excel_row, column=1).value = f"=ROW()-{HEADER_ROW}"

        for src_idx, dst_col in column_map:
            value = row_data[src_idx] if src_idx < len(row_data) else None
            ws.cell(row=excel_row, column=dst_col).value = value

        for col_idx in range(1, last_col + 1):
            if col_idx <= len(sample_cells):
                src = sample_cells[col_idx - 1]
            else:
                src = sample_cells[len(sample_cells) - 1]
            copy_cell_style(src, ws.cell(row=excel_row, column=col_idx))

        if sample_row_height is not None:
            ws.row_dimensions[excel_row].height = sample_row_height

    last_data_row = DATA_START_ROW + len(data_rows) - 1

    for offset, (_, store_name) in enumerate(store_columns):
        col_num = 7 + offset
        for r in range(2, 6):
            if category_sums:
                ws.cell(row=r, column=col_num).value = category_sums.get((r, store_name), 0)
            else:
                ws.cell(row=r, column=col_num).value = 0

    for col_num in range(7, last_col + 1):
        letter = get_column_letter(col_num)
        ws[f"{letter}6"] = f"=SUM({letter}2:{letter}5)"

    g_width = ws.column_dimensions["G"].width
    if g_width:
        for col_num in range(8, last_col + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = g_width

    # 保存前：将表头行 G7 起的单元格值替换为括号内的内容
    for col in range(7, last_col + 1):
        cell = ws.cell(row=HEADER_ROW, column=col)
        if cell.value:
            m = re.search(r'[（(](.+?)[）)]', str(cell.value))
            if m:
                cell.value = m.group(1)

    # 保存前：更新 A6 和 C6 中的日期
    if target_date:
        target_dt = datetime.strptime(target_date, "%Y.%m.%d")
        prev_day = (target_dt - timedelta(days=2)).strftime("%Y.%m.%d")
        for cell_ref, new_date in [("A6", prev_day), ("C6", target_date)]:
            cell = ws[cell_ref]
            if cell.value:
                cell.value = re.sub(r'\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2}', new_date, str(cell.value))

    wb.save(output_file)
    return last_data_row


def merge_board(data_file: Path, detail_file: Path, template_file: Path, output_file: Path,
                target_date=None):
    logger.info(f"数据源: {data_file}")
    logger.info(f"明细  : {detail_file}")
    logger.info(f"模板  : {template_file}")
    logger.info(f"输出  : {output_file}")

    if not data_file.exists():
        raise FileNotFoundError(f"数据源文件不存在: {data_file}")
    if not template_file.exists():
        raise FileNotFoundError(f"模板文件不存在: {template_file}")

    logger.info("读取数据源...")
    total_col_idx, store_columns, data_rows = read_data_file(data_file)
    logger.info(f"  共 {len(data_rows)} 行数据")
    logger.info(f"  合计订货量列: {get_column_letter(total_col_idx + 1)} (索引 {total_col_idx})")
    logger.info(f"  门店列 ({len(store_columns)}): {', '.join(name for _, name in store_columns)}")

    category_sums = {}
    if detail_file.exists():
        store_names = {name for _, name in store_columns}
        logger.info("读取明细数据，计算分类汇总金额...")
        category_sums = compute_category_sums(detail_file, store_names)
        for (r, store), val in sorted(category_sums.items()):
            logger.info(f"  Row {r} / {store}: {val}")
    else:
        logger.info(f"  明细文件不存在，汇总区将填入 0: {detail_file}")

    logger.info("填入模板...")
    last_row = merge_into_template(data_rows, total_col_idx, store_columns, template_file, output_file,
                                   category_sums, target_date=target_date)
    last_letter = get_column_letter(6 + len(store_columns))
    logger.info(f"  数据区: A{DATA_START_ROW}:{last_letter}{last_row}")
    logger.info(f"已生成: {output_file}")
    return output_file


# ─── 浏览器自动化函数 ────────────────────────────────────────────────────────


def set_date(page, placeholder: str, value: str):
    """设置日期输入框的值并触发必要事件"""
    page.evaluate(f"""
        (function() {{
            var inp = document.querySelector('input.timeInput.hasDatepicker[placeholder="{placeholder}"]');
            if (!inp) return 'not found';
            inp.value = "{value}";
            inp.dispatchEvent(new Event('input',  {{bubbles: true}}));
            inp.dispatchEvent(new Event('change', {{bubbles: true}}));
            inp.dispatchEvent(new Event('blur',   {{bubbles: true}}));
            return inp.value;
        }})()
    """)


def verify_dates(page, expected_date: str) -> bool:
    """验证两个日期输入框的值是否都包含期望日期"""
    result = page.evaluate("""
        (function() {
            var r = [];
            document.querySelectorAll('input.timeInput.hasDatepicker').forEach(function(inp) {
                r.push(inp.placeholder + '=' + inp.value);
            });
            return r.join(' | ');
        })()
    """)
    logger.info(f"  [日期验证] {result}")
    return expected_date in result and result.count(expected_date) == 2


def click_by_text(page, text: str, desc: str = ""):
    """通过完整文本内容点击叶子节点元素"""
    result = page.evaluate(f"""
        (function() {{
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {{
                if (els[i].textContent.trim() === "{text}" && els[i].children.length === 0) {{
                    els[i].click();
                    return 'clicked tag=' + els[i].tagName + ' class=' + els[i].className;
                }}
            }}
            return 'not found';
        }})()
    """)
    tag = f"[{desc}] " if desc else ""
    logger.info(f"  {tag}→ {result}")
    return "clicked" in result


def login(page):
    """登录流程"""
    logger.info("\n[1/4] 打开登录页面...")
    page.goto(LOGIN_URL)
    page.wait_for_load_state("networkidle")

    logger.info("[2/4] 切换到工号登录模式...")
    page.evaluate("""
        (function() {
            var els = document.querySelectorAll('div,span,a,button');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '工号登录' && els[i].children.length === 0) {
                    els[i].click();
                    return 'clicked';
                }
            }
            return 'not found';
        })()
    """)
    time.sleep(1.5)

    placeholder = page.evaluate("""
        (function() {
            var inputs = document.querySelectorAll('input');
            var r = [];
            for (var i = 0; i < inputs.length; i++) {
                if (inputs[i].placeholder) r.push(inputs[i].placeholder);
            }
            return r.join(' | ');
        })()
    """)
    if "员工工号" not in placeholder:
        raise RuntimeError(f"工号登录切换失败，当前输入框：{placeholder}")
    logger.info(f"  表单已切换 → {placeholder}")

    logger.info("[3/4] 填入账号/工号/密码...")
    page.evaluate(f"""
        (function() {{
            var inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].placeholder === '请输入账号')     inputs[i].value = "{ACCOUNT}";
                if (inputs[i].placeholder === '请输入员工工号') inputs[i].value = "{WORKER_ID}";
                if (inputs[i].placeholder === '请输入工号密码') inputs[i].value = "{PASSWORD}";
            }}
            inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].value) {{
                    inputs[i].dispatchEvent(new Event('input', {{bubbles: true}}));
                }}
            }}
        }})()
    """)

    time.sleep(1.5)

    logger.info("[4/4] 点击登录按钮...")
    click_by_text(page, "登 录", "登录")
    page.wait_for_load_state("networkidle", timeout=30_000)

    time.sleep(2)


def navigate_to_board(page, board_url: str, board_name: str):
    """导航到指定看板"""
    logger.info(f"\n  [导航] 前往{board_name}...")
    page.goto(board_url)
    page.wait_for_load_state("networkidle", timeout=30_000)
    logger.info(f"  已到达 → {page.url}")


def setup_filters(page, target_date: str, *, select_status: bool = False):
    """设置筛选条件：期望到货时间、商品分类、日期范围，可选单据状态"""
    logger.info("\n  [筛选] 设置筛选条件...")

    # ── 1. 切换日期类型标签 ──
    logger.info("  → 切换到「期望到货时间」标签...")
    result = page.evaluate("""
        (function() {
            var lis = document.querySelectorAll('li');
            for (var i = 0; i < lis.length; i++) {
                if (lis[i].textContent.trim() === '期望到货时间') {
                    lis[i].click();
                    return 'clicked li[' + i + ']';
                }
            }
            return 'not found';
        })()
    """)
    logger.info(f"    {result}")
    if result == "not found":
        raise RuntimeError("未找到「期望到货时间」标签")

    tab_status = page.evaluate("""
        (function() {
            var r = [];
            document.querySelectorAll('li').forEach(function(li) {
                var t = li.textContent.trim();
                if (t === '订货时间' || t === '期望到货时间') {
                    r.push(t + ':' + li.className);
                }
            });
            return r.join(' | ');
        })()
    """)
    logger.info(f"    标签状态: {tab_status}")

    # ── 2. 展开高级搜索，选择商品分类 ──
    logger.info("  → 展开高级搜索面板...")
    page.evaluate("document.getElementById('advancedBtn').click()")
    time.sleep(0.5)

    logger.info("  → 打开分类选择弹框...")
    page.evaluate("document.getElementById('selectCategory').click()")
    time.sleep(1.0)

    logger.info(f"  → 勾选 {len(TARGET_CATEGORIES)} 个分类...")
    categories_js = str(TARGET_CATEGORIES).replace("'", '"')
    page.evaluate(f"""
        (function() {{
            var targets = {categories_js};
            var divs = document.querySelectorAll('.checkBoxDiv');
            divs.forEach(function(d) {{
                var span = d.querySelector('span');
                if (span && targets.indexOf(span.textContent.trim()) >= 0) {{
                    d.click();
                }}
            }});
        }})()
    """)

    check_result = page.evaluate(f"""
        (function() {{
            var targets = {categories_js};
            var divs = document.querySelectorAll('.checkBoxDiv');
            var r = [];
            divs.forEach(function(d) {{
                var s = d.querySelector('span');
                if (s && targets.indexOf(s.textContent.trim()) >= 0) {{
                    r.push(s.textContent.trim() + ':' + d.classList.contains('on'));
                }}
            }});
            return r.join(' | ');
        }})()
    """)
    logger.info(f"    勾选验证: {check_result}")
    if "false" in check_result or check_result.count(":true") < len(TARGET_CATEGORIES):
        logger.warning("部分分类未勾选，尝试重试...")
        page.evaluate(f"""
            (function() {{
                var targets = {categories_js};
                var divs = document.querySelectorAll('.checkBoxDiv');
                divs.forEach(function(d) {{
                    var span = d.querySelector('span');
                    if (span && targets.indexOf(span.textContent.trim()) >= 0 && !d.classList.contains('on')) {{
                        d.click();
                    }}
                }});
            }})()
        """)

    logger.info("  → 点击「确定」关闭弹框...")
    click_by_text(page, "确定", "确定弹框")
    time.sleep(0.5)

    # ── 3. 选择单据状态（仅明细看板）──
    if select_status:
        logger.info("  → 选择单据状态...")
        page.evaluate("""
            (function() {
                var el = document.getElementById('ddl_productRequestStatus');
                if (el) el.click();
            })()
        """)
        time.sleep(1.0)

        page.evaluate("""
            (function(targets) {
                var items = document.querySelectorAll('#ddl_productRequestStatus .selectBox ul li');
                for (var i = 0; i < items.length; i++) {
                    var title = items[i].getAttribute('title');
                    if (title && targets.indexOf(title.trim()) >= 0 && !items[i].classList.contains('on')) {
                        items[i].click();
                    }
                }
                var closeBtn = document.querySelector('#ddl_productRequestStatus .selectBox .bottomBar .btnGrey14');
                if (closeBtn) closeBtn.click();
            })(%s)
        """ % str(STATUS_OPTIONS).replace("'", '"'))
        time.sleep(1.0)

    # ── 4. 设置日期范围 ──
    logger.info(f"  → 设置日期: {target_date}...")
    set_date(page, "开始日期", f"{target_date} 00:00")
    set_date(page, "结束日期", f"{target_date} 23:59")


def search_and_count_rows(page, target_date: str, btn_id: str, max_retries: int = 3) -> int:
    """点击查询按钮，验证日期未被重置，返回结果行数"""
    logger.info("\n  [查询] 执行查询...")

    for attempt in range(1, max_retries + 1):
        logger.info(f"  查询第 {attempt} 次...")
        page.evaluate(f"document.getElementById('{btn_id}').click()")
        page.wait_for_load_state("networkidle", timeout=90_000)

        result = page.evaluate("""
            (function() {
                var rows = document.querySelectorAll('table tbody tr');
                var r = 'rows:' + rows.length + ' | ';
                document.querySelectorAll('input.timeInput.hasDatepicker').forEach(function(inp) {
                    r += inp.placeholder + '=' + inp.value + ' | ';
                });
                return r;
            })()
        """)
        logger.info(f"    结果: {result}")

        if target_date in result:
            m = re.search(r"rows:(\d+)", result)
            row_count = int(m.group(1)) if m else 0
            logger.info(f"  ✅ 查询成功，共 {row_count} 行数据")
            return row_count
        else:
            logger.warning("日期被重置！重新设置日期...")
            set_date(page, "开始日期", f"{target_date} 00:00")
            set_date(page, "结束日期", f"{target_date} 23:59")
            if not verify_dates(page, target_date):
                logger.warning("日期验证失败")

    raise RuntimeError(f"查询 {max_retries} 次后日期仍然不正确，请手动检查")


def export_and_save(page, target_date: str, file_prefix: str) -> Path:
    """点击导出，等待下载，保存到输出目录，返回文件路径"""
    logger.info("\n  [导出] 导出文件...")
    time.sleep(3)

    date_str = target_date.replace(".", "-")
    daily_output_dir = OUTPUT_DIR / date_str / "原始下载"
    daily_output_dir.mkdir(parents=True, exist_ok=True)
    logger.info(f"  → 输出至: {daily_output_dir}")

    with page.expect_download(timeout=60_000) as dl_info:
        result = page.evaluate("""
            (function() {
                var els = document.querySelectorAll('*');
                for (var i = 0; i < els.length; i++) {
                    if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                        els[i].click();
                        return 'ok';
                    }
                }
                return 'not found';
            })()
        """)
        logger.info(f"  点击导出: {result}")
        if result == "not found":
            raise RuntimeError("未找到「导出」按钮")

    download = dl_info.value
    logger.info(f"  下载文件名: {download.suggested_filename}")

    dest = daily_output_dir / f"{file_prefix}_{date_str}.xlsx"
    download.save_as(dest)
    logger.info(f"  ✅ 已保存到: {dest}")
    return dest


def main():
    parser = argparse.ArgumentParser(description="Pospal 订货商品汇总看板自动导出")
    parser.add_argument("--date", type=str, help="指定日期，格式 YYYY.MM.DD，如 2026.05.30")
    parser.add_argument("--days", type=int, default=2, help="今天之后第N天（默认2=后天）")
    parser.add_argument("--headless", action="store_true", help="无头模式（不显示浏览器窗口）")
    args = parser.parse_args()

    if args.date:
        target_date = args.date
    else:
        target_dt = datetime.now() + timedelta(days=args.days)
        target_date = target_dt.strftime("%Y.%m.%d")

    logger.info(f"{'=' * 55}")
    logger.info(f"  Pospal 订货看板导出")
    logger.info(f"  目标日期：{target_date}")
    logger.info(f"  输出根目录：{OUTPUT_DIR}")
    logger.info(f"{'=' * 55}")

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        logger.error("未安装 playwright，请先运行: pip install playwright && playwright install chromium")
        sys.exit(1)

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=args.headless,
            args=["--start-maximized"],
        )
        context = browser.new_context(
            viewport={"width": 1600, "height": 900},
            accept_downloads=True,
        )
        page = context.new_page()

        try:
            login(page)

            # ── 任务 1：订货商品汇总看板 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  任务 1/2：订货商品汇总看板")
            logger.info(f"{'─' * 55}")

            navigate_to_board(page, SUMMARY_BOARD_URL, "订货商品汇总看板")
            setup_filters(page, target_date)
            summary_row_count = search_and_count_rows(page, target_date, "btnLoadRequestList")
            summary_path = export_and_save(page, target_date, "订货商品汇总看板")

            # ── 任务 2：订货商品明细看板 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  任务 2/2：订货商品明细看板")
            logger.info(f"{'─' * 55}")

            navigate_to_board(page, ITEM_BOARD_URL, "订货商品明细看板")
            setup_filters(page, target_date, select_status=True)
            item_row_count = search_and_count_rows(page, target_date, "btnList")
            item_path = export_and_save(page, target_date, "订货商品明细看板")

            logger.info(f"{'=' * 55}")
            logger.info(f"  ✅ 下载完成！")
            logger.info(f"  订货商品汇总看板：{summary_row_count} 行 → {summary_path}")
            logger.info(f"  订货商品明细看板：{item_row_count} 行 → {item_path}")
            logger.info(f"{'=' * 55}\n")

            # ── 任务 3：格式化输出（按分类分表）──
            logger.info(f"{'─' * 55}")
            logger.info(f"  任务 3/3：生成格式化汇总看板（{len(EXPORT_CATEGORY_MAP)} 个分表）")
            logger.info(f"{'─' * 55}\n")

            date_str = target_date.replace(".", "-")

            logger.info("读取汇总数据源...")
            total_col_idx, store_columns, data_rows = read_data_file(summary_path)
            store_names = {name for _, name in store_columns}
            logger.info(f"  共 {len(data_rows)} 行，{len(store_columns)} 个门店")

            detail_sums = {}
            if item_path.exists():
                logger.info("读取明细数据，按分类汇总金额...")
                detail_sums = compute_detail_sums_by_category(item_path, store_names)

            all_category_sums = aggregate_to_row_sums(detail_sums, None)
            all_output = OUTPUT_DIR / date_str / f"订货商品汇总看板_格式化_全部_{date_str}.xlsx"
            merge_into_template(
                data_rows, total_col_idx, store_columns,
                TEMPLATE_FILE, all_output, all_category_sums,
                target_date=target_date,
            )
            logger.info(f"  [全部] {len(data_rows)} 行 → {all_output}")

            for export_name, export_cats in EXPORT_CATEGORY_MAP.items():
                export_cat_set = set(export_cats)
                filtered_rows = [
                    row for row in data_rows
                    if row[1] is not None and str(row[1]).strip() in export_cat_set
                ]
                logger.info(f"\n  [{export_name}] 匹配 {len(filtered_rows)} 行")
                if not filtered_rows:
                    logger.info(f"  [{export_name}] 无数据，跳过")
                    continue

                output_file = OUTPUT_DIR / date_str / f"订货商品汇总看板_{export_name}_{date_str}.xlsx"
                merge_into_template(
                    filtered_rows, total_col_idx, store_columns,
                    TEMPLATE_FILE, output_file, all_category_sums,
                    target_date=target_date,
                )
                logger.info(f"  [{export_name}] → {output_file}")

            logger.info(f"{'=' * 55}")
            logger.info(f"  ✅ 全部完成！")
            logger.info(f"{'=' * 55}\n")

        except Exception as e:
            logger.error(f"任务失败: {e}")
            try:
                screenshot = OUTPUT_DIR / "error_screenshot.png"
                page.screenshot(path=str(screenshot))
                logger.info(f"  错误截图已保存: {screenshot}")
            except Exception:
                pass
            raise
        finally:
            context.close()
            browser.close()


if __name__ == "__main__":
    main()
