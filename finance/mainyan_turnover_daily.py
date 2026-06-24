"""
麦安研营业每日报表 - 自动化脚本
=====================================
依赖：pip install playwright openpyxl xlrd && playwright install chromium

自动登录银豹后台，导出各类报表数据并生成麦安研营业每日报表。

用法：
    python mainyan_turnover_daily.py                    # 导出今天数据
    python mainyan_turnover_daily.py --days -1          # 导出昨天数据
    python mainyan_turnover_daily.py --date 2026.06.03  # 导出指定日期
    python mainyan_turnover_daily.py --headless          # 无头模式（不显示浏览器）
"""

import argparse
import calendar
import copy
import logging
import logging.handlers
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

import xlrd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

# ─── 日志配置 ───────────────────────────────────────────────────────────────────
LOG_DIR = Path(__file__).resolve().parent / "log"
LOG_DIR.mkdir(exist_ok=True)

_file_handler = logging.handlers.TimedRotatingFileHandler(
    LOG_DIR / "mainyan_turnover_daily.log",
    when="midnight",
    backupCount=30,
    encoding="utf-8",
)
_file_handler.suffix = "%Y-%m-%d"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[_file_handler, logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# ─── 配置区 ──────────────────────────────────────────────────────────────────────
ACCOUNT = "huomimayzb"
WORKER_ID = "M008"
PASSWORD = "Maianyan88"

LOGIN_URL = "https://beta69.pospal.cn/"
DISCARD_HISTORY_URL = "https://beta69.pospal.cn/Inventory/DiscardInventoryHistory"
DISCARD_COUNT_URL = "https://beta69.pospal.cn/Inventory/DiscardProductCount"
PRODUCT_SALE_URL = "https://beta69.pospal.cn/ReportV2/ProductSale"
SALE_ANALYSIS_URL = "https://beta69.pospal.cn/Report/SaleAnalysis"
DELIVERY_REPORT_URL = "https://css69.pospal.cn/ChainStoreSupplyReport/WarehouseDeliveryProductStoreComparison"

OUTPUT_DIR = Path(__file__).resolve().parent / "麦安研营业每日报表"
TEMPLATE_FILE = Path(__file__).resolve().parent / "麦安研营业每日报表_格式化模板.xlsx"
STATS_DIR = Path(__file__).resolve().parent / "麦安研营业统计"

STORES = [
    {
        "store_short": "宝泰店",
        "short_abbr": "宝泰",
        "multi_select_items": ["1 - 麦安研", "3 - 麦安研（东站宝泰店）"],
        "single_select_items": ["1 - 麦安研", "3 - 麦安研（东站宝泰店）"],
        "stats_ws_index": 0,
        "col_offset": 0,
        "hour_row_map": {11: 23, 12: 24, 13: 25, 16: 26, 17: 27, 18: 28},
    },
    {
        "store_short": "龙江店",
        "short_abbr": "龙江",
        "multi_select_items": ["5 - 麦安研（顺德龙江店）"],
        "single_select_items": ["5 - 麦安研（顺德龙江店）"],
        "stats_ws_index": 1,
        "col_offset": 10,
        "hour_row_map": {17: 23, 18: 24, 19: 25, 20: 26, 21: 27, 22: 28},
    },
    {
        "store_short": "杏坛店",
        "short_abbr": "杏坛",
        "multi_select_items": ["2 - 麦安研（顺德杏坛店）"],
        "single_select_items": ["2 - 麦安研（顺德杏坛店）"],
        "stats_ws_index": 2,
        "col_offset": 20,
        "hour_row_map": {17: 23, 18: 24, 19: 25, 20: 26, 21: 27, 22: 28},
    },
]

DISCARD_REASONS = ["报废", "过期"]

DELIVERY_TARGET_CATEGORIES = [
    "包材耗材",
    "工衣模具",
    "慕斯+饼干+饮品+其他",
    "原料铺料",
    "冷冻面团",
    "蛋糕及面包成品及饼干类",
]

DELIVERY_STORE_NAME_MAP = {
    "麦安研（顺德杏坛店）": "杏坛",
    "麦安研（东站宝泰店）": "宝泰",
    "麦安研（佛山创产店）": "创产",
    "麦安研（顺德龙江店）": "龙江",
    "麦安研（佛山万民金海城店）": "金海城",
    "麦安研门店裱花间": "裱花间",
    "中央工厂冷加工间": "冷加工间",
}



# ─── 工具函数 ──────────────────────────────────────────────────────────────────


def get_target_date(days: int = 0, date_str: str = None):
    if date_str:
        return datetime.strptime(date_str, "%Y.%m.%d")
    return datetime.now() + timedelta(days=days)


def retry_until_success(fn, description, max_retries=10, retry_delay=5):
    for attempt in range(1, max_retries + 1):
        try:
            return fn()
        except Exception as e:
            if attempt == max_retries:
                logger.error(f"[重试] {description} 已重试 {max_retries} 次仍失败，放弃: {e}")
                raise
            wait = retry_delay * attempt
            logger.warning(f"[重试] {description} 第 {attempt}/{max_retries} 次失败: {e}")
            logger.info(f"[重试] {wait} 秒后重试...")
            time.sleep(wait)


def _parse_numeric(val):
    if val is None:
        return None
    try:
        return float(str(val).strip().replace(",", ""))
    except (ValueError, TypeError):
        return val


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def _match_store(store_short, name):
    name = str(name).strip()
    if store_short == "宝泰店":
        return name in ("麦安研", "1 - 麦安研") or "宝泰" in name
    if store_short == "龙江店":
        return "龙江" in name
    if store_short == "杏坛店":
        return "杏坛" in name
    return False


def _get_delivery_short(full_name):
    short = DELIVERY_STORE_NAME_MAP.get(full_name)
    if short:
        return short
    m = re.search(r'[（(](.+?)[）)]', full_name)
    if m:
        return m.group(1).rstrip('店')
    return full_name


def _sub_label(item_name):
    if "（" in item_name:
        m = re.search(r'[（(](.+?)[）)]', item_name)
        return m.group(1) if m else item_name
    return item_name.split(" - ")[-1] if " - " in item_name else item_name


def _eval_cell(ws, row, col, cache=None):
    """递归计算单元格值，支持简单加减法公式。"""
    if cache is None:
        cache = {}
    key = (row, col)
    if key in cache:
        return cache[key]

    val = ws.cell(row=row, column=col).value
    if val is None:
        cache[key] = 0
        return 0
    if isinstance(val, (int, float)):
        cache[key] = val
        return val
    if isinstance(val, str) and val.startswith("="):
        result = _eval_formula(ws, val, cache)
        cache[key] = result
        return result
    try:
        result = float(str(val).strip().replace(",", ""))
        cache[key] = result
        return result
    except (ValueError, TypeError):
        cache[key] = 0
        return 0


def _eval_formula(ws, formula, cache):
    """解析并计算仅含 +/- 和单元格引用的公式。"""
    expr = formula.lstrip("=").strip()
    tokens = re.split(r"(?=[+-])", expr)

    total = 0.0
    for token in tokens:
        token = token.strip()
        if not token:
            continue
        sign = 1
        if token.startswith("-"):
            sign = -1
            token = token[1:].strip()
        elif token.startswith("+"):
            token = token[1:].strip()

        m = re.match(r"^([A-Z]+)(\d+)$", token)
        if m:
            col = column_index_from_string(m.group(1))
            row = int(m.group(2))
            total += sign * _eval_cell(ws, row, col, cache)
        else:
            try:
                total += sign * float(token)
            except ValueError:
                pass

    return total


# ─── 浏览器自动化函数 ────────────────────────────────────────────────────────────


def set_date(page, placeholder: str, value: str):
    page.evaluate(f"""
        (function() {{
            var inp = document.querySelector('input.timeInput.hasDatepicker[placeholder="{placeholder}"]');
            if (!inp) return 'not found';
            inp.value = "{value}";
            inp.dispatchEvent(new Event('input',  {{bubbles: true}}));
            inp.dispatchEvent(new Event('change', {{bubbles: true}}));
            inp.dispatchEvent(new Event('blur',   {{bubbles: true}}));
            return inp.value;
        }})()
    """)


def click_by_text(page, text: str, desc: str = ""):
    result = page.evaluate(f"""
        (function() {{
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {{
                if (els[i].textContent.trim() === "{text}" && els[i].children.length === 0) {{
                    els[i].click();
                    return 'clicked tag=' + els[i].tagName + ' class=' + els[i].className;
                }}
            }}
            return 'not found';
        }})()
    """)
    tag = f"[{desc}] " if desc else ""
    logger.info(f"  {tag}→ {result}")
    return "clicked" in result


def login(page):
    logger.info("[1/4] 打开登录页面...")
    page.goto(LOGIN_URL)
    page.wait_for_load_state("networkidle")
    time.sleep(2)

    logger.info("[2/4] 切换到工号登录模式...")
    page.evaluate("""
        (function() {
            var els = document.querySelectorAll('div,span,a,button');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '工号登录' && els[i].children.length === 0) {
                    els[i].click();
                    return 'clicked';
                }
            }
            return 'not found';
        })()
    """)
    page.wait_for_selector('input[placeholder="请输入员工工号"]', timeout=10000)

    placeholder = page.evaluate("""
        (function() {
            var inputs = document.querySelectorAll('input');
            var r = [];
            for (var i = 0; i < inputs.length; i++) {
                if (inputs[i].placeholder) r.push(inputs[i].placeholder);
            }
            return r.join(' | ');
        })()
    """)
    if "员工工号" not in placeholder:
        raise RuntimeError(f"工号登录切换失败，当前输入框：{placeholder}")
    logger.info(f"  表单已切换 → {placeholder}")

    logger.info("[3/4] 填入账号/工号/密码...")
    page.evaluate(f"""
        (function() {{
            var inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].placeholder === '请输入账号')     inputs[i].value = "{ACCOUNT}";
                if (inputs[i].placeholder === '请输入员工工号') inputs[i].value = "{WORKER_ID}";
                if (inputs[i].placeholder === '请输入工号密码') inputs[i].value = "{PASSWORD}";
            }}
            inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].value) {{
                    inputs[i].dispatchEvent(new Event('input', {{bubbles: true}}));
                }}
            }}
        }})()
    """)

    logger.info("[4/4] 点击登录按钮...")
    click_by_text(page, "登 录", "登录")
    page.wait_for_load_state("networkidle", timeout=120_000)
    time.sleep(2)


def select_stores_multi(page, store_names, dropdown_id="ddl_subUsers"):
    logger.info(f"  → 选择门店: {store_names}")
    dropdown = page.locator(f"#{dropdown_id}")
    dropdown.click()
    page.wait_for_selector(f"#{dropdown_id} .selectBox", state="visible", timeout=10000)

    page.evaluate(f"""
        (function() {{
            var lis = document.querySelectorAll('#{dropdown_id} .selectBox li');
            lis.forEach(function(li) {{
                var cb = li.querySelector('.checkBox14');
                if (li.classList.contains('on') || (cb && cb.classList.contains('on'))) {{
                    li.click();
                }}
            }});
        }})()
    """)
    time.sleep(0.3)

    for name in store_names:
        result = page.evaluate(f"""
            (function() {{
                var lis = document.querySelectorAll('#{dropdown_id} .selectBox li');
                for (var i = 0; i < lis.length; i++) {{
                    var t = lis[i].textContent.replace(/\\u00a0/g, ' ').trim();
                    if (t === '{name}') {{
                        lis[i].click();
                        return 'checked: ' + t;
                    }}
                }}
                var names = [];
                for (var i = 0; i < lis.length; i++) {{
                    names.push(lis[i].textContent.replace(/\\u00a0/g, ' ').trim());
                }}
                return 'not found: {name}. Available: ' + names.join(', ');
            }})()
        """)
        logger.info(f"    {result}")

    close_btn = page.locator(f"#{dropdown_id} .bottomBar .btnGrey14")
    if close_btn.count() > 0:
        close_btn.click()
    page.wait_for_selector(f"#{dropdown_id} .selectBox", state="hidden", timeout=5000)


def select_single_option(page, dropdown_id, option_text):
    logger.info(f"  → 选择: {option_text} (#{dropdown_id})")
    dropdown = page.locator(f"#{dropdown_id}")
    dropdown.click()
    page.wait_for_selector(f"#{dropdown_id} .selectBox", state="visible", timeout=10000)

    result = page.evaluate(f"""
        (function() {{
            var lis = document.querySelectorAll('#{dropdown_id} .selectBox li');
            for (var i = 0; i < lis.length; i++) {{
                var t = lis[i].textContent.replace(/\\u00a0/g, ' ').trim();
                if (t === '{option_text}') {{
                    lis[i].click();
                    return 'selected: ' + t;
                }}
            }}
            var names = [];
            for (var i = 0; i < lis.length; i++) {{
                names.push(lis[i].textContent.replace(/\\u00a0/g, ' ').trim());
            }}
            return 'not found: {option_text}. Available: ' + names.join(', ');
        }})()
    """)
    logger.info(f"    {result}")
    time.sleep(0.5)


def export_with_popup(page, export_btn_id):
    page.locator(f"#{export_btn_id}").click()
    logger.info(f"  点击导出: #{export_btn_id}")
    time.sleep(1)

    with page.expect_download(timeout=180_000) as dl_info:
        popup_result = page.evaluate("""
            (function() {
                var btn = document.querySelector('.layui-layer-btn0');
                if (btn) { btn.click(); return 'layui-btn0'; }
                var els = document.querySelectorAll('*');
                var exports = [];
                for (var i = 0; i < els.length; i++) {
                    if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                        exports.push(els[i]);
                    }
                }
                if (exports.length > 1) {
                    exports[exports.length - 1].click();
                    return 'popup-export-' + exports.length;
                }
                if (exports.length === 1) {
                    exports[0].click();
                    return 'single-export';
                }
                for (var i = 0; i < els.length; i++) {
                    if (els[i].textContent.trim() === '确定' && els[i].children.length === 0) {
                        els[i].click();
                        return 'confirm';
                    }
                }
                return 'no popup button found';
            })()
        """)
        logger.info(f"  弹窗确认: {popup_result}")

    return dl_info.value


# ─── 数据下载函数 ──────────────────────────────────────────────────────────────


def download_discard_history(page, target_str, date_label, output_dir):
    """第1个页面：商品报损记录"""
    TAG = "[商品报损记录]"
    logger.info(f"{TAG} 前往页面...")
    page.goto(DISCARD_HISTORY_URL)
    page.wait_for_load_state("networkidle", timeout=120_000)
    time.sleep(2)

    logger.info(f"{TAG} 设置时间: {target_str}...")
    set_date(page, "开始日期", f"{target_str} 00:00")
    set_date(page, "结束日期", f"{target_str} 23:59")
    time.sleep(1)

    logger.info(f"{TAG} 查询...")
    click_by_text(page, "查询", "查询")
    page.wait_for_load_state("networkidle", timeout=150_000)
    time.sleep(3)

    logger.info(f"{TAG} 导出...")
    download = export_with_popup(page, "btnExportDiscardInventoryHistory")
    time.sleep(3)

    dest = output_dir / f"商品报损记录_{date_label}.xls"
    download.save_as(dest)
    logger.info(f"{TAG} 已保存: {dest.name}")
    return dest


def download_discard_count(page, target_str, date_label, store, reason, output_dir):
    """第2个页面：商品报损统计（按门店+报损原因）"""
    abbr = store["short_abbr"]
    TAG = f"[商品报损统计-{abbr}-{reason}]"
    logger.info(f"{TAG} 前往页面...")
    page.goto(DISCARD_COUNT_URL)
    page.wait_for_load_state("networkidle", timeout=120_000)
    time.sleep(2)

    select_stores_multi(page, store["multi_select_items"])
    select_single_option(page, "ddl_reasons", reason)
    time.sleep(2)

    logger.info(f"{TAG} 设置时间: {target_str}...")
    set_date(page, "开始日期", f"{target_str} 00:00")
    set_date(page, "结束日期", f"{target_str} 23:59")
    time.sleep(2)

    logger.info(f"{TAG} 查询...")
    click_by_text(page, "查询", "查询")
    page.wait_for_load_state("networkidle", timeout=150_000)
    time.sleep(3)

    logger.info(f"{TAG} 导出...")
    download = export_with_popup(page, "btnExport")

    dest = output_dir / f"商品报损统计_{abbr}_{reason}_{date_label}.xls"
    download.save_as(dest)
    logger.info(f"{TAG} 已保存: {dest.name}")
    return dest


def download_product_sale(page, target_str, date_label, store, output_dir):
    """第3个页面：商品销售统计（按门店）"""
    abbr = store["short_abbr"]
    TAG = f"[商品销售统计-{abbr}]"
    logger.info(f"{TAG} 前往页面...")
    page.goto(PRODUCT_SALE_URL)
    page.wait_for_load_state("networkidle", timeout=120_000)
    time.sleep(2)

    select_stores_multi(page, store["multi_select_items"])

    logger.info(f"{TAG} 设置时间: {target_str}...")
    set_date(page, "开始日期", f"{target_str} 00:00")
    set_date(page, "结束日期", f"{target_str} 23:59")

    logger.info(f"{TAG} 查询...")
    click_by_text(page, "查询", "查询")
    page.wait_for_load_state("networkidle", timeout=150_000)
    time.sleep(3)

    logger.info(f"{TAG} 导出...")
    download = export_with_popup(page, "btnExport")

    dest = output_dir / f"商品销售统计_{abbr}_{date_label}.xlsx"
    download.save_as(dest)
    logger.info(f"{TAG} 已保存: {dest.name}")
    return dest


def download_sale_analysis(page, target_str, date_label, store, output_dir):
    """第4个页面：销售趋势分析（单选门店，按小时）
    返回 {hour: amount} 合并后的小时数据。
    """
    abbr = store["short_abbr"]
    all_hourly = {}

    for item_name in store["single_select_items"]:
        sub = _sub_label(item_name)
        TAG = f"[销售趋势分析-{abbr}-{sub}]"
        logger.info(f"{TAG} 前往页面...")
        page.goto(SALE_ANALYSIS_URL)
        page.wait_for_load_state("networkidle", timeout=120_000)
        time.sleep(2)

        select_single_option(page, "ddl_subUsers", item_name)
        select_single_option(page, "ddl_countType", "按小时")

        logger.info(f"{TAG} 设置时间: {target_str}...")
        set_date(page, "开始日期", f"{target_str} 00:00")
        set_date(page, "结束日期", f"{target_str} 23:59")

        logger.info(f"{TAG} 统计分析...")
        click_by_text(page, "统计分析", "统计分析")
        page.wait_for_load_state("networkidle", timeout=150_000)
        time.sleep(3)

        # 从页面表格直接抓取小时数据
        hourly_rows = page.evaluate("""
            (function() {
                var table = document.getElementById('saleAnalysisTable');
                if (!table) return [];
                var rows = table.querySelectorAll('tbody tr:not(.totalRow)');
                var result = [];
                for (var i = 0; i < rows.length; i++) {
                    var tds = rows[i].querySelectorAll('td');
                    if (tds.length < 5) continue;
                    result.push({
                        time: tds[1].textContent.trim(),
                        amount: parseFloat(tds[4].textContent.trim().replace(/,/g, '')) || 0
                    });
                }
                return result;
            })()
        """)
        for row in hourly_rows:
            try:
                hour = int(str(row["time"]).strip())
            except ValueError:
                continue
            all_hourly[hour] = all_hourly.get(hour, 0) + row["amount"]
        logger.info(f"{TAG} 抓取到 {len(hourly_rows)} 行小时数据")

        # 下载表格存档（点第一个下载表格按钮）
        logger.info(f"{TAG} 下载表格...")
        try:
            with page.expect_download(timeout=60_000) as dl_info:
                page.locator("#saleAnalysisTableDiv .option.export").click()
            download = dl_info.value
            dest = output_dir / f"销售趋势分析_{abbr}_{sub}_{date_label}.xls"
            download.save_as(dest)
            logger.info(f"{TAG} 已保存: {dest.name}")
        except Exception as e:
            logger.warning(f"{TAG} 下载表格失败（数据已从页面抓取）: {e}")

    return all_hourly


def download_delivery_comparison(page, target_str, date_label, output_dir):
    """第5个页面：仓库配送商品门店对比表（不含配送费）"""
    TAG = "[仓库配送]"
    logger.info(f"{TAG} 前往页面...")
    page.goto(DELIVERY_REPORT_URL)
    page.wait_for_load_state("networkidle", timeout=120_000)
    time.sleep(2)

    logger.info(f"{TAG} 设置时间: {target_str}...")
    set_date(page, "开始日期", f"{target_str} 00:00")
    set_date(page, "结束日期", f"{target_str} 23:59")

    logger.info(f"{TAG} 点击更多搜索...")
    page.evaluate("""
        (function() {
            var btn = document.getElementById('advancedBtn');
            if (btn) { btn.click(); return 'clicked'; }
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {
                var t = els[i].textContent.trim();
                if ((t === '更多搜索' || t === '更多搜索▼') && els[i].children.length === 0) {
                    els[i].click();
                    return 'clicked tag=' + els[i].tagName;
                }
            }
            return 'not found';
        })()
    """)
    time.sleep(0.5)

    logger.info(f"{TAG} 打开分类选择弹框...")
    page.evaluate("""
        (function() {
            var btn = document.getElementById('selectCategory');
            if (btn) { btn.click(); return 'clicked'; }
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '选择分类' && els[i].children.length === 0) {
                    els[i].click();
                    return 'clicked tag=' + els[i].tagName;
                }
            }
            return 'not found';
        })()
    """)
    time.sleep(1)

    categories_js = str(DELIVERY_TARGET_CATEGORIES).replace("'", '"')
    logger.info(f"{TAG} 勾选 {len(DELIVERY_TARGET_CATEGORIES)} 个分类（不含配送费）...")
    page.evaluate(f"""
        (function() {{
            var targets = {categories_js};
            var divs = document.querySelectorAll('.checkBoxDiv');
            divs.forEach(function(d) {{
                var span = d.querySelector('span');
                if (span && targets.indexOf(span.textContent.trim()) >= 0) {{
                    if (!d.classList.contains('on')) d.click();
                }}
            }});
        }})()
    """)
    time.sleep(0.5)

    check_result = page.evaluate(f"""
        (function() {{
            var targets = {categories_js};
            var divs = document.querySelectorAll('.checkBoxDiv');
            var r = [];
            divs.forEach(function(d) {{
                var s = d.querySelector('span');
                if (s && targets.indexOf(s.textContent.trim()) >= 0) {{
                    r.push(s.textContent.trim() + ':' + d.classList.contains('on'));
                }}
            }});
            return r.join(' | ');
        }})()
    """)
    logger.info(f"{TAG} 勾选验证: {check_result}")

    if "false" in check_result:
        logger.warning(f"{TAG} 部分分类未勾选，重试...")
        page.evaluate(f"""
            (function() {{
                var targets = {categories_js};
                var divs = document.querySelectorAll('.checkBoxDiv');
                divs.forEach(function(d) {{
                    var span = d.querySelector('span');
                    if (span && targets.indexOf(span.textContent.trim()) >= 0 && !d.classList.contains('on')) {{
                        d.click();
                    }}
                }});
            }})()
        """)

    click_by_text(page, "确定", "确定弹框")
    time.sleep(0.5)

    logger.info(f"{TAG} 查询...")
    click_by_text(page, "查询", "查询")
    page.wait_for_load_state("networkidle", timeout=150_000)
    time.sleep(3)

    logger.info(f"{TAG} 导出...")
    with page.expect_download(timeout=180_000) as dl_info:
        result = page.evaluate("""
            (function() {
                var els = document.querySelectorAll('*');
                for (var i = 0; i < els.length; i++) {
                    if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                        els[i].click();
                        return 'ok';
                    }
                }
                return 'not found';
            })()
        """)
        logger.info(f"{TAG} 点击导出: {result}")
        if result == "not found":
            raise RuntimeError("未找到「导出」按钮")

    download = dl_info.value
    dest = output_dir / f"仓库配送商品门店对比表_{date_label}.xlsx"
    download.save_as(dest)
    logger.info(f"{TAG} 已保存: {dest.name}")
    return dest


# ─── 数据读取函数 ──────────────────────────────────────────────────────────────


def read_discard_history_data(file_path, store_short, reason):
    """读取商品报损记录（xls），按门店+报损原因筛选，返回报损金额合计。"""
    if not file_path.exists():
        return 0
    wb = xlrd.open_workbook(str(file_path))
    ws = wb.sheet_by_index(0)

    total = 0
    for r in range(1, ws.nrows):
        store_name = str(ws.cell_value(r, 2) or "").strip()  # C列 (0-indexed=2)
        discard_reason = str(ws.cell_value(r, 5) or "").strip()  # F列 (0-indexed=5)
        amount = _parse_numeric(ws.cell_value(r, 3))  # D列 (0-indexed=3)

        if _match_store(store_short, store_name) and discard_reason == reason:
            if amount is not None and isinstance(amount, (int, float)):
                total += amount

    return round(total, 2)


def read_discard_count_data(file_paths):
    """读取商品报损统计文件（xls，报废+过期合并），按报损金额降序返回商品列表。"""
    products = []

    for fp in file_paths:
        if not fp.exists():
            continue
        wb = xlrd.open_workbook(str(fp))
        ws = wb.sheet_by_index(0)

        for r in range(1, ws.nrows):
            name = ws.cell_value(r, 0)  # A列 (0-indexed=0)
            if name is None:
                continue
            name = str(name).strip()
            if not name or name in ("合计", "总计"):
                continue
            qty = _parse_numeric(ws.cell_value(r, 5))  # F列 (0-indexed=5)
            amount = _parse_numeric(ws.cell_value(r, 7))  # H列 (0-indexed=7)
            if amount is not None and isinstance(amount, (int, float)) and amount != 0:
                products.append({
                    "name": name,
                    "qty": qty if isinstance(qty, (int, float)) else 0,
                    "amount": amount,
                })

    products.sort(key=lambda x: abs(x["amount"]), reverse=True)
    return products


def read_product_sale_data(file_path):
    """读取商品销售统计，处理团购核销合并，按实收金额降序返回商品列表。"""
    if not file_path.exists():
        return []

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    rows = []
    for r in range(2, ws.max_row):
        name = ws.cell(row=r, column=1).value
        if name is None:
            continue
        name = str(name).strip()
        if not name or name in ("合计", "总计", "无码商品"):
            continue
        category = str(ws.cell(row=r, column=6).value or "").strip()
        amount = _parse_numeric(ws.cell(row=r, column=10).value)
        if amount is None or not isinstance(amount, (int, float)):
            amount = 0
        rows.append({"name": name, "category": category, "amount": amount})
    wb.close()

    tuangou_rows = [r for r in rows if r["category"] == "团购核销"]
    normal_rows = [r for r in rows if r["category"] != "团购核销"]

    normal_map = {}
    for r in normal_rows:
        if r["name"] in normal_map:
            normal_map[r["name"]]["amount"] += r["amount"]
        else:
            normal_map[r["name"]] = r

    for tg in tuangou_rows:
        base_name = tg["name"].split("-")[0].strip()
        if base_name in normal_map:
            normal_map[base_name]["amount"] += tg["amount"]
        else:
            normal_map[base_name] = {
                "name": base_name,
                "category": "",
                "amount": tg["amount"],
            }

    result = list(normal_map.values())
    result.sort(key=lambda x: x["amount"], reverse=True)
    return result


def read_delivery_store_totals(data_file):
    """读取仓库配送商品门店对比表，返回各门店出库金额合计。"""
    if not data_file.exists():
        return {}

    wb = load_workbook(data_file, data_only=True)
    ws = wb.active

    store_totals = {}
    store_groups = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h and str(h).strip().endswith("_数量"):
            name_part = str(h).strip().rsplit("_", 1)[0]
            if name_part != "合计":
                store_groups.append((name_part, c, c + 2))

    for full_name, qty_col, amt_col in store_groups:
        short = _get_delivery_short(full_name)
        total = 0
        for r in range(2, ws.max_row + 1):
            seq = ws.cell(row=r, column=1).value
            if seq is None:
                break
            val = ws.cell(row=r, column=amt_col).value
            if val is not None:
                try:
                    total += float(val)
                except (ValueError, TypeError):
                    pass
        if total != 0:
            store_totals[short] = round(total, 2)

    wb.close()
    logger.info(f"  仓库配送各门店金额: {store_totals}")
    return store_totals


def read_stats_value(target, store, column):
    """读取麦安研营业统计指定列的数值（支持公式递归计算）。"""
    month_str = f"{target.year}-{target.month:02d}"
    stats_file = STATS_DIR / month_str / f"麦安研营业统计_{month_str}.xlsx"

    if not stats_file.exists():
        logger.warning(f"  营业统计文件不存在: {stats_file}")
        return None

    wb = load_workbook(stats_file, data_only=False)
    if store["stats_ws_index"] >= len(wb.worksheets):
        logger.warning(f"  营业统计文件sheet不足: index={store['stats_ws_index']}")
        wb.close()
        return None

    ws = wb.worksheets[store["stats_ws_index"]]
    row = target.day + 4
    val = _eval_cell(ws, row, column)
    wb.close()
    return round(val, 2) if val else None


def read_stats_bi_month_sum(target, store):
    """读取营业统计 BI 列当月（1号到目标日期）的累计值（支持公式递归计算）。"""
    month_str = f"{target.year}-{target.month:02d}"
    stats_file = STATS_DIR / month_str / f"麦安研营业统计_{month_str}.xlsx"

    if not stats_file.exists():
        return None

    wb = load_workbook(stats_file, data_only=False)
    if store["stats_ws_index"] >= len(wb.worksheets):
        wb.close()
        return None

    ws = wb.worksheets[store["stats_ws_index"]]
    cache = {}

    total = 0
    for day in range(1, target.day + 1):
        row = day + 4
        val = _eval_cell(ws, row, 61, cache)
        total += val

    wb.close()
    return round(total, 2) if total else None


# ─── 模板创建与填充 ────────────────────────────────────────────────────────────


def create_or_open_monthly_file(target):
    year = target.year
    month = target.month
    month_str = f"{year}-{month:02d}"
    month_dir = OUTPUT_DIR / month_str
    month_dir.mkdir(parents=True, exist_ok=True)

    output_file = month_dir / f"麦安研营业每日报表_{month_str}.xlsx"

    if output_file.exists():
        logger.info(f"  月度文件已存在: {output_file.name}")
        return output_file

    logger.info(f"  从模板创建月度文件: {output_file.name}")
    _create_monthly_from_template(output_file, year, month)
    return output_file


def _create_monthly_from_template(output_file, year, month):
    wb = load_workbook(TEMPLATE_FILE)
    days_in_month = calendar.monthrange(year, month)[1]

    if len(wb.worksheets) < 2:
        logger.warning("  模板文件不足2个sheet")
        wb.save(output_file)
        wb.close()
        return

    template_ws = wb.worksheets[1]

    for day in range(3, days_in_month + 1):
        wb.copy_worksheet(template_ws)

    for day in range(2, days_in_month + 1):
        sheet_idx = day - 1
        if sheet_idx < len(wb.worksheets):
            wb.worksheets[sheet_idx].title = f"{month}.{day}"

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    wb.close()
    logger.info(f"  已创建月度文件: {output_file}")


def _set_week_day_cell(ws, row, col, weekday_num):
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont

        existing_font = ws.cell(row=row, column=col).font
        font_name = existing_font.name or "等线"
        font_size = existing_font.size or 11

        ws.cell(row=row, column=col).value = CellRichText(
            TextBlock(InlineFont(rFont=font_name, sz=font_size), "本周第"),
            TextBlock(InlineFont(rFont=font_name, sz=font_size, color="FF0000"), str(weekday_num)),
            TextBlock(InlineFont(rFont=font_name, sz=font_size), "天"),
        )
    except Exception:
        ws.cell(row=row, column=col).value = f"本周第{weekday_num}天"


def fill_daily_sheet(monthly_file, target, download_dir, sale_analysis_data):
    """将所有下载数据填入对应日期的 sheet。"""
    day = target.day
    date_label = target.strftime("%Y-%m-%d")
    sheet_idx = day - 1

    wb = load_workbook(monthly_file)
    if sheet_idx >= len(wb.worksheets):
        logger.error(f"  sheet索引 {sheet_idx} 超出范围 (共 {len(wb.worksheets)} 个sheet)")
        wb.close()
        return

    ws = wb.worksheets[sheet_idx]
    logger.info(f"  填充第 {day} 天数据到 sheet: {ws.title}")

    discard_history_file = download_dir / f"商品报损记录_{date_label}.xls"
    delivery_file = download_dir / f"仓库配送商品门店对比表_{date_label}.xlsx"
    delivery_totals = read_delivery_store_totals(delivery_file)

    for store in STORES:
        offset = store["col_offset"]
        abbr = store["short_abbr"]
        logger.info(f"    → 填充 {store['store_short']} (col_offset={offset})")

        # ── B3: 日期 ──
        ws.cell(row=3, column=offset + 2).value = date_label

        # ── F3: 本周第X天 ──
        weekday_num = target.isoweekday()
        _set_week_day_cell(ws, row=3, col=offset + 6, weekday_num=weekday_num)

        # ── B8: 营业统计 BI 列值 ──
        bi_val = read_stats_value(target, store, column=61)
        if bi_val is not None:
            ws.cell(row=8, column=offset + 2).value = bi_val
            logger.info(f"      B8 (BI): {bi_val}")

        # ── B9: 营业统计 BI 列当月累计（1号到目标日期） ──
        bi_month = read_stats_bi_month_sum(target, store)
        if bi_month is not None:
            ws.cell(row=9, column=offset + 2).value = bi_month
            logger.info(f"      B9 (BI月累计): {bi_month}")

        # ── B14: 仓库配送门店金额合计 ──
        delivery_amount = delivery_totals.get(abbr, 0)
        if delivery_amount:
            ws.cell(row=14, column=offset + 2).value = delivery_amount
            logger.info(f"      B14 (配送): {delivery_amount}")

        # ── B18: 商品报损记录 "试吃" 合计 ──
        if discard_history_file.exists():
            b18_val = read_discard_history_data(discard_history_file, store["store_short"], "试吃")
            if b18_val:
                ws.cell(row=18, column=offset + 2).value = b18_val
                logger.info(f"      B18 (试吃): {b18_val}")

        # ── B19: 商品报损记录 "报废"+"过期" 合计 ──
        if discard_history_file.exists():
            b19_baofei = read_discard_history_data(discard_history_file, store["store_short"], "报废")
            b19_guoqi = read_discard_history_data(discard_history_file, store["store_short"], "过期")
            b19_val = round(b19_baofei + b19_guoqi, 2)
            if b19_val:
                ws.cell(row=19, column=offset + 2).value = b19_val
                logger.info(f"      B19 (报废+过期): {b19_val}")

        # ── B24~D28: 商品报损统计 Top5（报废+过期合并） ──
        discard_files = [
            download_dir / f"商品报损统计_{abbr}_{reason}_{date_label}.xls"
            for reason in DISCARD_REASONS
        ]
        damage_products = read_discard_count_data(discard_files)
        damage_map = {p["name"]: p["amount"] for p in damage_products}

        for i, prod in enumerate(damage_products[:5]):
            row = 24 + i
            ws.cell(row=row, column=offset + 2).value = prod["name"]
            ws.cell(row=row, column=offset + 3).value = prod["qty"]
            ws.cell(row=row, column=offset + 4).value = prod["amount"]
        logger.info(f"      B24-B28: {len(damage_products[:5])} 条报损排行")

        # ── F8~H11: 商品销售统计 Top4 + 对应报损金额 ──
        sale_file = download_dir / f"商品销售统计_{abbr}_{date_label}.xlsx"
        sale_products = read_product_sale_data(sale_file)

        for i, prod in enumerate(sale_products[:4]):
            row = 8 + i
            ws.cell(row=row, column=offset + 6).value = prod["name"]
            ws.cell(row=row, column=offset + 7).value = prod["amount"]
            h_val = damage_map.get(prod["name"])
            if h_val is not None:
                ws.cell(row=row, column=offset + 8).value = h_val
        logger.info(f"      F8-F11: {len(sale_products[:4])} 条销售排行")

        # ── G14: 营业统计 W 列值 ──
        w_val = read_stats_value(target, store, column=23)
        if w_val is not None:
            ws.cell(row=14, column=offset + 7).value = w_val
            logger.info(f"      G14 (W): {w_val}")

        # ── G23~G28: 销售趋势分析小时数据 ──
        hourly_data = sale_analysis_data.get(store["store_short"], {})
        hour_row_map = store["hour_row_map"]
        filled_hours = 0
        max_val = None
        max_row = None
        for hour, row in hour_row_map.items():
            if hour in hourly_data:
                val = hourly_data[hour]
                ws.cell(row=row, column=offset + 7).value = val
                filled_hours += 1
                if max_val is None or val > max_val:
                    max_val = val
                    max_row = row
        if max_row is not None:
            cell = ws.cell(row=max_row, column=offset + 7)
            old_border = copy.copy(cell.border)
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.fill = red_fill
            cell.border = old_border
        logger.info(f"      G23-G28: {filled_hours} 个时段, 最大值行={max_row}")

    wb.save(monthly_file)
    wb.close()
    logger.info(f"  已保存第 {day} 天数据到: {monthly_file.name}")


# ─── 主函数 ────────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(description="麦安研营业每日报表自动化脚本")
    parser.add_argument("--days", type=int, default=0, help="日期偏移：0=今天，-1=昨天（默认0）")
    parser.add_argument("--date", type=str, help="指定日期，格式 YYYY.MM.DD")
    parser.add_argument("--headless", action="store_true", help="无头模式（不显示浏览器窗口）")
    args = parser.parse_args()

    target = get_target_date(days=args.days, date_str=args.date)
    target_str = target.strftime("%Y.%m.%d")
    date_label = target.strftime("%Y-%m-%d")

    logger.info(f"{'=' * 55}")
    logger.info(f"  麦安研营业每日报表")
    logger.info(f"  目标日期：{target_str}")
    logger.info(f"  输出根目录：{OUTPUT_DIR}")
    logger.info(f"{'=' * 55}")

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        logger.error("未安装 playwright，请先运行: pip install playwright && playwright install chromium")
        sys.exit(1)

    download_dir = OUTPUT_DIR / "原始下载" / date_label
    download_dir.mkdir(parents=True, exist_ok=True)

    sale_analysis_data = {}

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=args.headless,
            args=["--start-maximized"],
        )
        context = browser.new_context(
            viewport={"width": 1600, "height": 900},
            accept_downloads=True,
        )
        page = context.new_page()
        page.set_default_timeout(120000)
        page.set_default_navigation_timeout(120000)

        try:
            login(page)

            # ── 步骤1：下载商品报损记录 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤1：下载商品报损记录")
            logger.info(f"{'─' * 55}")
            retry_until_success(
                lambda: download_discard_history(page, target_str, date_label, download_dir),
                "商品报损记录"
            )

            # ── 步骤2：下载商品报损统计 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤2：下载商品报损统计")
            logger.info(f"{'─' * 55}")
            for store in STORES:
                for reason in DISCARD_REASONS:
                    desc = f"商品报损统计-{store['short_abbr']}-{reason}"
                    retry_until_success(
                        lambda s=store, r=reason: download_discard_count(
                            page, target_str, date_label, s, r, download_dir
                        ),
                        desc
                    )

            # ── 步骤3：下载商品销售统计 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤3：下载商品销售统计")
            logger.info(f"{'─' * 55}")
            for store in STORES:
                desc = f"商品销售统计-{store['short_abbr']}"
                retry_until_success(
                    lambda s=store: download_product_sale(
                        page, target_str, date_label, s, download_dir
                    ),
                    desc
                )

            # ── 步骤4：下载销售趋势分析 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤4：下载销售趋势分析")
            logger.info(f"{'─' * 55}")
            for store in STORES:
                desc = f"销售趋势分析-{store['short_abbr']}"

                def _do_sale_analysis(s=store):
                    hourly = download_sale_analysis(
                        page, target_str, date_label, s, download_dir
                    )
                    sale_analysis_data[s["store_short"]] = hourly

                retry_until_success(_do_sale_analysis, desc)

            # ── 步骤5：下载仓库配送商品门店对比表 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤5：下载仓库配送商品门店对比表")
            logger.info(f"{'─' * 55}")
            retry_until_success(
                lambda: download_delivery_comparison(
                    page, target_str, date_label, download_dir
                ),
                "仓库配送"
            )

        except Exception as e:
            logger.error(f"下载任务失败: {e}")
            try:
                screenshot = OUTPUT_DIR / "error_screenshot.png"
                page.screenshot(path=str(screenshot))
                logger.info(f"  错误截图已保存: {screenshot}")
            except Exception:
                pass
            raise
        finally:
            context.close()
            browser.close()

    # ── 步骤6：格式化填充报表 ──
    logger.info(f"{'─' * 55}")
    logger.info(f"  步骤6：格式化填充报表")
    logger.info(f"{'─' * 55}")

    monthly_file = create_or_open_monthly_file(target)
    fill_daily_sheet(monthly_file, target, download_dir, sale_analysis_data)

    logger.info(f"{'=' * 55}")
    logger.info(f"  麦安研营业每日报表全部完成！")
    logger.info(f"  月度文件: {monthly_file}")
    logger.info(f"{'=' * 55}\n")


if __name__ == "__main__":
    main()
