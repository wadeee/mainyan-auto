"""
麦安研营业统计 - 自动化脚本
=====================================
依赖：pip install playwright && playwright install chromium

自动登录银豹后台，导出麦安研各门店的营业概况日度统计数据。

用法：
    python mainyan_turnover_statistics.py                    # 导出今天数据
    python mainyan_turnover_statistics.py --days -1          # 导出昨天数据
    python mainyan_turnover_statistics.py --date 2026.06.03  # 导出指定日期
    python mainyan_turnover_statistics.py --headless          # 无头模式（不显示浏览器）
"""

import argparse
import calendar
import copy
import csv
import json
import logging.handlers
import re
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path

# ─── 日志配置 ───────────────────────────────────────────────────────────────────
LOG_DIR = Path(__file__).resolve().parent / "log"
LOG_DIR.mkdir(exist_ok=True)

_file_handler = logging.handlers.TimedRotatingFileHandler(
    LOG_DIR / "mainyan_turnover_statistics.log",
    when="midnight",
    backupCount=30,
    encoding="utf-8",
)
_file_handler.suffix = "%Y-%m-%d"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[_file_handler, logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# ─── 配置区（按需修改）─────────────────────────────────────────────────────────
ACCOUNT = "huomimayzb"
WORKER_ID = "M008"
PASSWORD = "Maianyan88"

# 银豹系统
LOGIN_URL = "https://beta69.pospal.cn/"
BUSINESS_SUMMARY_URL = "https://beta69.pospal.cn/Report/BusinessSummaryV2"
UNIONPAY_BILL_URL = "https://cloudapp-pay69.pospal.cn/#/additional/fund-summary?oem=0"
CUSTOMER_SUMMARY_URL = "https://beta69.pospal.cn/CustomerReport/CustomerConsumerSummary"

# 美团外卖
# https://e.waimai.meituan.com
# 美团外卖结算账单页
MEITUAN_DOWNLOAD_URL_PRE = "https://e.waimai.meituan.com/#https://waimaieapp.meituan.com/finance/pc/settleBill"
# 美团外卖账单
MEITUAN_DOWNLOAD_URL = "https://waimaieapp.meituan.com/finance/static/gray_html_pc/billReconciliation.html#/daily-bill"
# 美团推广消费
MEITUAN_AD_URL = "https://waimaieapp.meituan.com/ad/v1/pc#/account"

# 美团经营宝
# https://ecom.meituan.com/meishi
# 美团团购每日收益
MEITUAN_JYB_URL = "https://ecom.meituan.com/finance-kdb/profit/home"

# 饿了么/淘宝闪购
# https://melody.shop.ele.me
# 饿了么账单
ELEME_BILL_URL = "https://napos-bill-pc.faas.ele.me/napos-bill-pc/v2/bill-checking?shopType=SINGLE"

# 招行后台
# https://ym.o2o.cmbchina.com/mc/merchant/handms/login.html
# 招行每日汇总
ZHAOHANG_URL = "https://ym.o2o.cmbchina.com/mc/merchant/handms/dailySummary.html"

# 高德口碑
# https://e.koubei.com
# 高德口碑账单汇总
KOUBEI_BILL_URL = "https://e.koubei.com/kb-pc/finance-mono/merchant/bill/realtime/summary"

MEITUAN_STORE_CONFIG = [
    {
        "store_short": "宝泰店",
        "port": 9223,
        "user_data_dir": r"C:\ChromeDebug_BT",
    },
    {
        "store_short": "龙江店",
        "port": 9224,
        "user_data_dir": r"C:\ChromeDebug_LJ",
    },
    {
        "store_short": "杏坛店",
        "port": 9225,
        "user_data_dir": r"C:\ChromeDebug_XT",
    },
]

ELEME_STORE_CONFIG = {"宝泰店"}

ZHAOHANG_STORE_CONFIG = [
    {
        "store_short": "宝泰店",
        "store_value": "002009212000001",
        "store_label": "MAINYAN麦安研(东方宝泰店)",
    }
]

DOUYIN_LOGIN_URL = "https://life.douyin.com/"
DOUYIN_DAILY_BENEFITS_URL = "https://life.douyin.com/p/fulfillsettle/dailyBenifits"
DOUYIN_ACCOUNT = "18688856666"
DOUYIN_PASSWORD = "Maianyan88"

DOUYIN_STORE_CONFIG = [
    {"store_short": "宝泰店", "search_keyword": "宝泰"},
    {"store_short": "龙江店", "search_keyword": "龙江"},
    {"store_short": "杏坛店", "search_keyword": "杏坛"},
]

KOUBEI_STORE_CONFIG = [
    {"store_short": "宝泰店", "store_name": "MAINYAN麦安研(东方宝泰店)"},
    {"store_short": "龙江店", "store_name": "麦安研(顺德龙江店)"},
    {"store_short": "杏坛店", "store_name": "麦安研(顺德杏坛店)"},
]

KOUBEI_CSV_FIELDS = ["订单金额", "商家优惠", "服务费"]

OUTPUT_DIR = Path(__file__).resolve().parent / "麦安研营业统计"
TEMPLATE_FILE = Path(__file__).resolve().parent / "麦安研营业统计_格式化模板.xlsx"

STORES = [
    {"full": "3 - 麦安研（东站宝泰店）", "short": "宝泰店", "template_name": "东站宝泰店"},
    {"full": "5 - 麦安研（顺德龙江店）", "short": "龙江店", "template_name": "顺德龙江店"},
    {"full": "2 - 麦安研（顺德杏坛店）", "short": "杏坛店", "template_name": "顺德杏坛店"},
]

UNIONPAY_STORE_CONFIG = [
    {
        "store_short": "宝泰店",
        "parent_node": "总部",
        "select_items": ["麦安研", "麦安研（东站宝泰店）"],
    },
    {
        "store_short": "龙江店",
        "parent_node": None,
        "select_items": ["麦安研（顺德龙江店）"],
    },
    {
        "store_short": "杏坛店",
        "parent_node": "总部",
        "select_items": ["麦安研（顺德杏坛店）"],
    },
]

CUSTOMER_SUMMARY_STORE_CONFIG = [
    {
        "store_short": "宝泰店",
        "select_items": ["1 - 麦安研", "3 - 麦安研（东站宝泰店）"],
    },
    {
        "store_short": "龙江店",
        "select_items": ["5 - 麦安研（顺德龙江店）"],
    },
    {
        "store_short": "杏坛店",
        "select_items": ["2 - 麦安研（顺德杏坛店）"],
    },
]

TEMPLATE_STORE_NAME = "东方宝泰店"

UNIONPAY_CSV_FIELDS = ["交易金额", "交易退款金额", "有效交易金额", "交易手续费", "优惠金额", "优惠退款金额"]

DOUYIN_CSV_FIELDS = ["订单实收", "佣金/服务费支出"]

WEEKDAY_NAMES = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]

CDP_PORTS = {cfg["port"] for cfg in MEITUAN_STORE_CONFIG} | {9226}


# ─── 工具函数 ──────────────────────────────────────────────────────────────────


def get_target_date(days: int = 0, date_str: str = None):
    if date_str:
        return datetime.strptime(date_str, "%Y.%m.%d")
    return datetime.now() + timedelta(days=days)


def retry_until_success(fn, description, max_retries=10, retry_delay=5):
    for attempt in range(1, max_retries + 1):
        try:
            return fn()
        except Exception as e:
            if attempt == max_retries:
                logger.error(f"[重试] {description} 已重试 {max_retries} 次仍失败，放弃: {e}")
                raise
            wait = retry_delay * attempt
            logger.warning(f"[重试] {description} 第 {attempt}/{max_retries} 次失败: {e}")
            logger.info(f"[重试] {wait} 秒后重试...")
            time.sleep(wait)


def _connect_cdp_with_retry(pw, port, max_retries=12, interval=5):
    for attempt in range(1, max_retries + 1):
        try:
            browser = pw.chromium.connect_over_cdp(f"http://localhost:{port}")
            context = browser.contexts[0]
            existing_pages = context.pages[:]
            page = context.new_page()
            for p in existing_pages:
                try:
                    p.close()
                except Exception:
                    pass
            page.set_default_timeout(120000)
            page.set_default_navigation_timeout(120000)
            return browser, page
        except Exception as e:
            if attempt == max_retries:
                logger.error(f"[CDP] 连接 port={port} 失败 {max_retries} 次，放弃: {e}")
                raise
            logger.warning(f"[CDP] 连接 port={port} 第 {attempt} 次失败: {e}，{interval}s 后重试...")
            time.sleep(interval)


def cleanup_stale_browsers():
    """在任务开始前清理残留的 Chrome/Chromium 进程，避免端口冲突和 user-data-dir 锁。"""
    killed_pids = set()

    # 1. 通过 netstat 查找占用 CDP 端口的进程
    try:
        output = subprocess.check_output(
            ["netstat", "-ano"], text=True, errors="replace"
        )
        for line in output.splitlines():
            for port in CDP_PORTS:
                if f":{port} " in line and "LISTENING" in line:
                    parts = line.split()
                    try:
                        pid = int(parts[-1])
                    except (ValueError, IndexError):
                        continue
                    if pid > 0 and pid not in killed_pids:
                        logger.info(f"  端口 {port} 被 PID {pid} 占用，终止进程树...")
                        subprocess.run(
                            ["taskkill", "/F", "/T", "/PID", str(pid)],
                            capture_output=True, timeout=10,
                        )
                        killed_pids.add(pid)
    except Exception as e:
        logger.warning(f"  检查端口占用失败: {e}")

    # 2. 查找使用 ChromeDebug 目录的 chrome.exe（可能端口未绑定但锁住了 profile）
    try:
        output = subprocess.check_output(
            ["wmic", "process", "where", "name='chrome.exe'", "get", "processid,commandline"],
            text=True, errors="replace", timeout=10,
        )
        for line in output.splitlines():
            if "ChromeDebug" in line:
                parts = line.strip().split()
                try:
                    pid = int(parts[-1])
                except (ValueError, IndexError):
                    continue
                if pid > 0 and pid not in killed_pids:
                    logger.info(f"  发现 ChromeDebug chrome.exe PID {pid}，终止进程树...")
                    subprocess.run(
                        ["taskkill", "/F", "/T", "/PID", str(pid)],
                        capture_output=True, timeout=10,
                    )
                    killed_pids.add(pid)
    except Exception as e:
        logger.warning(f"  检查 ChromeDebug 进程失败: {e}")

    # 3. 清理 Playwright 残留的 chromium.exe
    try:
        output = subprocess.check_output(
            ["tasklist", "/FI", "IMAGENAME eq chromium.exe"],
            text=True, errors="replace",
        )
        for line in output.splitlines():
            if "chromium.exe" in line.lower():
                parts = line.split()
                try:
                    pid = int(parts[1])
                except (ValueError, IndexError):
                    continue
                if pid not in killed_pids:
                    logger.info(f"  发现残留 chromium.exe PID {pid}，终止进程树...")
                    subprocess.run(
                        ["taskkill", "/F", "/T", "/PID", str(pid)],
                        capture_output=True, timeout=10,
                    )
                    killed_pids.add(pid)
    except Exception as e:
        logger.warning(f"  检查 chromium.exe 进程失败: {e}")

    if killed_pids:
        logger.info(f"  已清理 {len(killed_pids)} 个残留进程，等待 3 秒释放资源...")
        time.sleep(3)
    else:
        logger.info(f"  未发现残留浏览器进程")


# ─── 格式化合并函数 ──────────────────────────────────────────────────────────


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def _parse_numeric(val):
    if val is None:
        return None
    try:
        return float(str(val).strip().replace(",", ""))
    except (ValueError, TypeError):
        return val


def _parse_recharge_amount(text):
    if not text or not isinstance(text, str):
        return None
    m = re.search(r"充值\s*([\d.]+)", text)
    return float(m.group(1)) if m else None


def read_business_summary(data_file: Path):
    from openpyxl import load_workbook as _lwb
    wb = _lwb(data_file, data_only=True)
    ws1 = wb.worksheets[0]
    ws2 = wb.worksheets[1] if len(wb.worksheets) > 1 else None
    result = {
        "c16": _parse_numeric(ws1.cell(row=16, column=3).value),
        "ws2_c4": _parse_numeric(ws2.cell(row=4, column=3).value) if ws2 else None,
        "e4": _parse_numeric(ws1.cell(row=4, column=5).value),
        "b4_recharge": _parse_recharge_amount(ws1.cell(row=4, column=2).value),
    }
    wb.close()
    return result


def create_or_open_monthly_file(stores, target: datetime, template_file: Path, output_dir: Path):
    year = target.year
    month = target.month
    month_str = f"{year}-{month:02d}"
    month_dir = output_dir / month_str
    month_dir.mkdir(parents=True, exist_ok=True)

    output_file = month_dir / f"麦安研营业统计_{month_str}.xlsx"

    if output_file.exists():
        logger.info(f"  月度文件已存在，直接打开: {output_file.name}")
        return output_file, False

    logger.info(f"  月度文件不存在，从模板创建: {output_file.name}")
    _create_monthly_from_template(template_file, output_file, stores, year, month)
    return output_file, True


def _create_monthly_from_template(template_file: Path, output_file: Path, stores, year: int, month: int):
    from openpyxl import load_workbook as _lwb

    wb = _lwb(template_file)
    days_in_month = calendar.monthrange(year, month)[1]

    template_ws = wb.worksheets[0]
    original_title = template_ws.title
    for _ in range(len(stores) - 1):
        wb.copy_worksheet(template_ws)

    for idx, store in enumerate(stores):
        store_name = store["template_name"]
        ws = wb.worksheets[idx]

        ws.title = original_title.replace(TEMPLATE_STORE_NAME, store_name).replace("2026年6月", f"{year}年{month}月")

        a1_val = ws.cell(row=1, column=1).value
        if a1_val and isinstance(a1_val, str):
            ws.cell(row=1, column=1).value = a1_val.replace(TEMPLATE_STORE_NAME, store_name).replace("2026年6月",
                                                                                                     f"{year}年{month}月")

        for day in range(1, 32):
            row_s1 = day + 4
            if day <= days_in_month:
                dt = datetime(year, month, day)
                ws.cell(row=row_s1, column=2).value = f"{year}.{month}.{day}"
                ws.cell(row=row_s1, column=3).value = WEEKDAY_NAMES[dt.weekday()]
            else:
                ws.cell(row=row_s1, column=2).value = None
                ws.cell(row=row_s1, column=3).value = None

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    wb.close()
    logger.info(f"  已创建月度文件: {output_file}")


def fill_daily_data(monthly_file: Path, target: datetime, store, daily_download_dir: Path, store_index: int = 0):
    from openpyxl import load_workbook as _lwb

    day = target.day
    date_label = target.strftime("%Y-%m-%d")

    summary_file = daily_download_dir / f"营业概况_{store['short']}_{date_label}.xlsx"
    if not summary_file.exists():
        logger.warning(f"  营业概况文件不存在，跳过填写: {summary_file.name}")
        summary = None
    else:
        summary = read_business_summary(summary_file)
        logger.info(f"  读取 sheet2 C4 (现金充值) = {summary['ws2_c4']}")
        logger.info(f"  读取 C16 (现金合计) = {summary['c16']}")
        logger.info(f"  读取 E4 (充值金额) = {summary['e4']}")
        logger.info(f"  读取 B4 解析充值金额 = {summary['b4_recharge']}")

    unionpay_file = daily_download_dir / f"银豹付交易账单_{store['short']}_{date_label}.csv"
    if not unionpay_file.exists():
        logger.warning(f"  银豹付交易账单不存在，跳过填写: {unionpay_file.name}")
        unionpay = None
    else:
        unionpay = read_unionpay_bill_csv(unionpay_file)
        logger.info(f"  读取银豹付 交易金额 = {unionpay.get('交易金额')}")
        logger.info(f"  读取银豹付 交易退款金额 = {unionpay.get('交易退款金额')}")
        logger.info(f"  读取银豹付 交易手续费 = {unionpay.get('交易手续费')}")

    customer_file = daily_download_dir / f"会员消费汇总表_{store['short']}_{date_label}.xlsx"
    if not customer_file.exists():
        logger.warning(f"  会员消费汇总表不存在，跳过填写: {customer_file.name}")
        customer = None
    else:
        customer = read_customer_summary(customer_file)

    meituan_file = daily_download_dir / f"账单明细_{store['short']}_{date_label}.csv"
    if not meituan_file.exists():
        logger.warning(f"  美团账单明细不存在，跳过填写: {meituan_file.name}")
        meituan = None
    else:
        meituan = read_unionpay_bill_csv(meituan_file)
        logger.info(f"  读取美团 商品总价 = {meituan.get('商品总价')}")
        logger.info(f"  读取美团 打包费 = {meituan.get('打包费')}")
        logger.info(f"  读取美团 其他类 = {meituan.get('其他类')}")

    ad_promo_file = daily_download_dir / f"美团外卖推广消费_{store['short']}_{date_label}.csv"
    if not ad_promo_file.exists():
        logger.warning(f"  美团推广消费不存在，跳过填写: {ad_promo_file.name}")
        ad_promo_total = None
    else:
        ad_promo_total = read_ad_promo_csv(ad_promo_file)
        logger.info(f"  读取美团推广 变化金额合计 = {ad_promo_total}")

    jyb_file = daily_download_dir / f"美团经营宝_每日收益_{store['short']}_{date_label}.csv"
    if not jyb_file.exists():
        logger.warning(f"  美团经营宝收益不存在，跳过填写: {jyb_file.name}")
        jyb = None
    else:
        jyb = read_unionpay_bill_csv(jyb_file)
        logger.info(f"  读取经营宝 售价={jyb.get('售价')}, 促销费={jyb.get('促销费')}, "
                    f"服务费={jyb.get('服务费')}, 其他费用={jyb.get('其他费用')}")

    eleme_file = daily_download_dir / f"饿了么账单_{store['short']}_{date_label}.csv"
    if not eleme_file.exists():
        logger.warning(f"  饿了么账单不存在，跳过填写: {eleme_file.name}")
        eleme = None
    else:
        eleme = read_unionpay_bill_csv(eleme_file)
        logger.info(f"  读取饿了么 订单类={eleme.get('订单类')}, 其他类={eleme.get('其他类')}")

    zhaohang_file = daily_download_dir / f"招行每日汇总_{store['short']}_{date_label}.csv"
    if not zhaohang_file.exists():
        logger.warning(f"  招行每日汇总不存在，跳过填写: {zhaohang_file.name}")
        zhaohang = None
    else:
        zhaohang = read_unionpay_bill_csv(zhaohang_file)
        logger.info(f"  读取招行 商户实收总计={zhaohang.get('商户实收总计')}")

    douyin_file = daily_download_dir / f"抖音每日收益_{store['short']}_{date_label}.csv"
    if not douyin_file.exists():
        logger.warning(f"  抖音每日收益不存在，跳过填写: {douyin_file.name}")
        douyin = None
    else:
        douyin = read_unionpay_bill_csv(douyin_file)
        logger.info(f"  读取抖音 订单实收={douyin.get('订单实收')}, 佣金/服务费支出={douyin.get('佣金/服务费支出')}")

    koubei_file = daily_download_dir / f"高德口碑账单汇总_{store['short']}_{date_label}.csv"
    if not koubei_file.exists():
        logger.warning(f"  高德口碑账单汇总不存在，跳过填写: {koubei_file.name}")
        koubei = None
    else:
        koubei = read_unionpay_bill_csv(koubei_file)
        logger.info(
            f"  读取口碑 订单金额={koubei.get('订单金额')}, 商家优惠={koubei.get('商家优惠')}, 服务费={koubei.get('服务费')}")

    wb = _lwb(monthly_file)
    ws1 = wb.worksheets[store_index]

    row_s1 = day + 4
    if summary:
        if summary["ws2_c4"] is not None:
            ws1.cell(row=row_s1, column=5).value = summary["ws2_c4"]
        if summary["c16"] is not None:
            ws1.cell(row=row_s1, column=6).value = summary["c16"]
            ws1.cell(row=row_s1, column=7).value = summary["c16"]
        if summary["e4"] is not None:
            ws1.cell(row=row_s1, column=22).value = summary["e4"]
        if summary["b4_recharge"] is not None:
            ws1.cell(row=row_s1, column=23).value = summary["b4_recharge"]

    if unionpay:
        if unionpay.get("交易金额") is not None:
            ws1.cell(row=row_s1, column=10).value = unionpay["交易金额"]
        if unionpay.get("交易退款金额") is not None:
            ws1.cell(row=row_s1, column=11).value = unionpay["交易退款金额"]
        if unionpay.get("交易手续费") is not None:
            ws1.cell(row=row_s1, column=13).value = unionpay["交易手续费"]

    if customer:
        if customer["principal"] is not None:
            ws1.cell(row=row_s1, column=26).value = customer["principal"]
        if customer["gift"] is not None:
            ws1.cell(row=row_s1, column=27).value = customer["gift"]

    if meituan:
        if meituan.get("商品总价") is not None:
            ws1.cell(row=row_s1, column=32).value = abs(meituan["商品总价"])
        if meituan.get("打包费") is not None:
            ws1.cell(row=row_s1, column=33).value = abs(meituan["打包费"])
        if meituan.get("商家对顾客的活动补贴") is not None:
            ws1.cell(row=row_s1, column=34).value = abs(meituan["商家对顾客的活动补贴"])
        if meituan.get("佣金") is not None:
            ws1.cell(row=row_s1, column=35).value = abs(meituan["佣金"])
        if meituan.get("配送服务费") is not None:
            ws1.cell(row=row_s1, column=36).value = abs(meituan["配送服务费"])
        if meituan.get("其他类") is not None:
            ws1.cell(row=row_s1, column=37).value = abs(meituan["其他类"])

    if ad_promo_total is not None:
        ws1.cell(row=row_s1, column=31).value = abs(ad_promo_total)

    if jyb:
        if jyb.get("售价") is not None:
            ws1.cell(row=row_s1, column=40).value = jyb["售价"]
        if jyb.get("促销费") is not None:
            ws1.cell(row=row_s1, column=41).value = jyb["促销费"]
        if jyb.get("服务费") is not None:
            ws1.cell(row=row_s1, column=42).value = jyb["服务费"]
        if jyb.get("其他费用") is not None:
            ws1.cell(row=row_s1, column=43).value = jyb["其他费用"]

    if eleme:
        if eleme.get("订单类") is not None:
            ws1.cell(row=row_s1, column=46).value = abs(eleme["订单类"])
        if eleme.get("其他类") is not None:
            ws1.cell(row=row_s1, column=47).value = abs(eleme["其他类"])

    if zhaohang:
        if zhaohang.get("商户实收总计") is not None:
            ws1.cell(row=row_s1, column=50).value = zhaohang["商户实收总计"]

    if douyin:
        if douyin.get("订单实收") is not None:
            ws1.cell(row=row_s1, column=52).value = douyin["订单实收"]
        if douyin.get("佣金/服务费支出") is not None:
            ws1.cell(row=row_s1, column=53).value = douyin["佣金/服务费支出"]

    if koubei:
        if koubei.get("订单金额") is not None:
            ws1.cell(row=row_s1, column=56).value = koubei["订单金额"]
        if koubei.get("商家优惠") is not None:
            ws1.cell(row=row_s1, column=57).value = koubei["商家优惠"]
        if koubei.get("服务费") is not None:
            ws1.cell(row=row_s1, column=58).value = koubei["服务费"]

    wb.save(monthly_file)
    wb.close()
    logger.info(f"  已将第 {day} 天的数据写入: {monthly_file.name}")


# ─── 浏览器自动化函数 ────────────────────────────────────────────────────────


def set_date(page, placeholder: str, value: str):
    page.evaluate(f"""
        (function() {{
            var inp = document.querySelector('input.timeInput.hasDatepicker[placeholder="{placeholder}"]');
            if (!inp) return 'not found';
            inp.value = "{value}";
            inp.dispatchEvent(new Event('input',  {{bubbles: true}}));
            inp.dispatchEvent(new Event('change', {{bubbles: true}}));
            inp.dispatchEvent(new Event('blur',   {{bubbles: true}}));
            return inp.value;
        }})()
    """)


def click_by_text(page, text: str, desc: str = ""):
    result = page.evaluate(f"""
        (function() {{
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {{
                if (els[i].textContent.trim() === "{text}" && els[i].children.length === 0) {{
                    els[i].click();
                    return 'clicked tag=' + els[i].tagName + ' class=' + els[i].className;
                }}
            }}
            return 'not found';
        }})()
    """)
    tag = f"[{desc}] " if desc else ""
    logger.info(f"  {tag}→ {result}")
    return "clicked" in result


def login(page):
    logger.info("[1/4] 打开登录页面...")
    page.goto(LOGIN_URL)
    page.wait_for_load_state("networkidle")

    logger.info("[2/4] 切换到工号登录模式...")
    page.evaluate("""
        (function() {
            var els = document.querySelectorAll('div,span,a,button');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '工号登录' && els[i].children.length === 0) {
                    els[i].click();
                    return 'clicked';
                }
            }
            return 'not found';
        })()
    """)
    page.wait_for_selector('input[placeholder="请输入员工工号"]', timeout=10000)

    placeholder = page.evaluate("""
        (function() {
            var inputs = document.querySelectorAll('input');
            var r = [];
            for (var i = 0; i < inputs.length; i++) {
                if (inputs[i].placeholder) r.push(inputs[i].placeholder);
            }
            return r.join(' | ');
        })()
    """)
    if "员工工号" not in placeholder:
        raise RuntimeError(f"工号登录切换失败，当前输入框：{placeholder}")
    logger.info(f"  表单已切换 → {placeholder}")

    logger.info("[3/4] 填入账号/工号/密码...")
    page.evaluate(f"""
        (function() {{
            var inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].placeholder === '请输入账号')     inputs[i].value = "{ACCOUNT}";
                if (inputs[i].placeholder === '请输入员工工号') inputs[i].value = "{WORKER_ID}";
                if (inputs[i].placeholder === '请输入工号密码') inputs[i].value = "{PASSWORD}";
            }}
            inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].value) {{
                    inputs[i].dispatchEvent(new Event('input', {{bubbles: true}}));
                }}
            }}
        }})()
    """)

    logger.info("[4/4] 点击登录按钮...")
    click_by_text(page, "登 录", "登录")
    page.wait_for_load_state("networkidle", timeout=120_000)


def select_store(page, store_full_name: str):
    """从自定义 div 下拉 (#ddl_subUsers) 中选择指定门店。"""
    logger.info(f"  → 选择门店: {store_full_name}")

    dropdown = page.locator("#ddl_subUsers")
    dropdown.click()
    page.wait_for_selector("#ddl_subUsers .selectBox", state="visible", timeout=10000)

    page.evaluate("""
        (function() {
            var lis = document.querySelectorAll('#ddl_subUsers .selectBox li');
            lis.forEach(function(li) {
                if (li.classList.contains('on')) li.click();
            });
        })()
    """)

    target_li = page.locator(f"#ddl_subUsers .selectBox li[title='{store_full_name}']")
    if target_li.count() > 0:
        target_li.click()
        logger.info(f"    门店已选中: {store_full_name}")
    else:
        click_result = page.evaluate(f"""
            (function() {{
                var lis = document.querySelectorAll('#ddl_subUsers .selectBox li');
                for (var i = 0; i < lis.length; i++) {{
                    var t = lis[i].textContent.replace(/\\u00a0/g, ' ').trim();
                    if (t === '{store_full_name}') {{
                        lis[i].click();
                        return 'clicked: ' + t;
                    }}
                }}
                var names = [];
                for (var i = 0; i < lis.length; i++) names.push(lis[i].textContent.replace(/\\u00a0/g, ' ').trim());
                return 'not found in: ' + names.join(', ');
            }})()
        """)
        logger.info(f"    模糊匹配: {click_result}")

    close_btn = page.locator("#ddl_subUsers .bottomBar .btnGrey14")
    if close_btn.count() > 0:
        close_btn.click()
        logger.info("    下拉框已关闭")
    page.wait_for_selector("#ddl_subUsers .selectBox", state="hidden", timeout=5000)


def select_stores_ddl(page, store_names):
    """从自定义 div 下拉 (#ddl_subUsers) 中选择多个门店。"""
    logger.info(f"  → 选择门店: {store_names}")

    dropdown = page.locator("#ddl_subUsers")
    dropdown.click()
    page.wait_for_selector("#ddl_subUsers .selectBox", state="visible", timeout=10000)

    page.evaluate("""
        (function() {
            var lis = document.querySelectorAll('#ddl_subUsers .selectBox li');
            lis.forEach(function(li) {
                if (li.classList.contains('on')) li.click();
            });
        })()
    """)

    for name in store_names:
        target_li = page.locator(f"#ddl_subUsers .selectBox li[title='{name}']")
        if target_li.count() > 0:
            target_li.click()
            logger.info(f"    门店已选中: {name}")
        else:
            click_result = page.evaluate(f"""
                (function() {{
                    var lis = document.querySelectorAll('#ddl_subUsers .selectBox li');
                    for (var i = 0; i < lis.length; i++) {{
                        var t = lis[i].textContent.replace(/\\u00a0/g, ' ').trim();
                        if (t === '{name}') {{
                            lis[i].click();
                            return 'clicked: ' + t;
                        }}
                    }}
                    var names = [];
                    for (var i = 0; i < lis.length; i++) names.push(lis[i].textContent.replace(/\\u00a0/g, ' ').trim());
                    return 'not found in: ' + names.join(', ');
                }})()
            """)
            logger.info(f"    模糊匹配: {click_result}")

    close_btn = page.locator("#ddl_subUsers .bottomBar .btnGrey14")
    if close_btn.count() > 0:
        close_btn.click()
        logger.info("    下拉框已关闭")
    page.wait_for_selector("#ddl_subUsers .selectBox", state="hidden", timeout=5000)


def click_export(page):
    """点击导出按钮。"""
    result = page.evaluate("""
        (function() {
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                    els[i].click();
                    return 'ok';
                }
            }
            return 'not found';
        })()
    """)
    logger.info(f"  点击导出: {result}")
    if result == "not found":
        raise RuntimeError("未找到「导出」按钮")


def scrape_unionpay_bill(page):
    """从银豹付交易账单页面抓取汇总数据。"""
    data = page.evaluate("""
        (function() {
            var items = document.querySelectorAll('.data-dz-item');
            var result = {};
            for (var i = 0; i < items.length; i++) {
                var titleEl = items[i].querySelector('.title');
                if (!titleEl) continue;
                var title = '';
                var childNodes = titleEl.childNodes;
                for (var j = 0; j < childNodes.length; j++) {
                    if (childNodes[j].nodeType === 3) title += childNodes[j].textContent;
                }
                title = title.replace(/\\s+/g, ' ').trim();
                var numEl = items[i].querySelector('.num');
                if (!numEl) continue;
                var numText = numEl.textContent.replace(/[￥,\\s]/g, '').trim();
                result[title] = parseFloat(numText) || 0;
            }
            return result;
        })()
    """)
    if not data:
        logger.warning("  未抓取到银豹付数据，可能页面未加载完成")
    else:
        logger.info(f"  抓取到银豹付数据: {json.dumps(data, ensure_ascii=False)}")
    return data


def save_unionpay_bill_csv(data, store_short, date_label, output_dir):
    """将银豹付交易账单数据保存为 CSV。"""
    csv_file = output_dir / f"银豹付交易账单_{store_short}_{date_label}.csv"
    with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["项目", "金额"])
        for key in UNIONPAY_CSV_FIELDS:
            writer.writerow([key, data.get(key, 0)])
    logger.info(f"  已保存CSV: {csv_file.name}")
    return csv_file


def read_unionpay_bill_csv(csv_file):
    """读取银豹付交易账单 CSV。"""
    result = {}
    with open(csv_file, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            result[row["项目"]] = _parse_numeric(row["金额"])
    return result


def read_ad_promo_csv(csv_file):
    total = 0
    with open(csv_file, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            amount = _parse_numeric(row["变化金额(元)"])
            if amount is not None:
                total += amount
    return total


def read_customer_summary(data_file: Path):
    """读取会员消费汇总表，返回合计行的 F列（本金消费）和 G列（赠送消费）。"""
    from openpyxl import load_workbook as _lwb
    wb = _lwb(data_file, data_only=True)
    ws = wb.worksheets[0]

    total_row = None
    for row in range(ws.max_row, 0, -1):
        for col in range(1, min(ws.max_column + 1, 10)):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val and "合计" in str(cell_val):
                total_row = row
                break
        if total_row is not None:
            break

    if total_row is None:
        total_row = ws.max_row

    result = {
        "principal": _parse_numeric(ws.cell(row=total_row, column=6).value),
        "gift": _parse_numeric(ws.cell(row=total_row, column=7).value),
    }
    wb.close()
    logger.info(f"  合计行={total_row}, 本金消费(F)={result['principal']}, 赠送消费(G)={result['gift']}")
    return result


def scrape_zhaohang_daily_summary(page, target_str, date_label, zh_config, output_dir):
    """在招行每日汇总页面选择门店、设置日期、查询并抓取商户实收总计。"""
    store_short = zh_config["store_short"]
    store_value = zh_config["store_value"]
    store_label = zh_config["store_label"]

    logger.info(f"  → 选择门店: {store_label}")
    page.evaluate(f"""
        (function() {{
            var select = document.getElementById('storelist');
            if (!select) return 'select not found';
            select.value = '{store_value}';
            var event = new Event('change', {{bubbles: true}});
            select.dispatchEvent(event);

            var items = document.querySelectorAll('.searchable-select-item');
            for (var i = 0; i < items.length; i++) {{
                if (items[i].getAttribute('data-value') === '{store_value}') {{
                    items[i].click();
                    break;
                }}
            }}
        }})()
    """)
    time.sleep(0.5)

    date_dash = date_label
    logger.info(f"  → 设置日期: {date_dash}...")
    page.evaluate(f"""
        (function() {{
            var startInput = document.getElementById('start');
            var endInput = document.getElementById('end');
            if (startInput) {{
                startInput.readOnly = false;
                startInput.value = '{date_dash}';
                startInput.readOnly = true;
            }}
            if (endInput) {{
                endInput.readOnly = false;
                endInput.value = '{date_dash}';
                endInput.readOnly = true;
            }}
        }})()
    """)
    time.sleep(0.5)

    logger.info("  [查询] 点击查询...")
    page.locator("button.J-query").click()
    page.wait_for_selector('.record-tbody tr.code-item', timeout=30000)
    time.sleep(2)

    logger.info("  → 抓取商户实收总计...")
    zh_data = page.evaluate("""
        (function() {
            var result = {};
            var rows = document.querySelectorAll('.record-tbody tr.code-item');
            var totalReceived = 0;
            for (var i = 0; i < rows.length; i++) {
                var tds = rows[i].querySelectorAll('td');
                if (tds.length >= 16) {
                    var val = parseFloat(tds[15].textContent.trim().replace(/,/g, '')) || 0;
                    totalReceived += val;
                }
            }
            result['商户实收总计'] = Math.round(totalReceived * 100) / 100;
            result['row_count'] = rows.length;
            return result;
        })()
    """)
    logger.info(f"  抓取数据: 共{zh_data.get('row_count', 0)}行, 商户实收总计={zh_data.get('商户实收总计', 0)}")

    csv_file = output_dir / f"招行每日汇总_{store_short}_{date_label}.csv"
    with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["项目", "金额"])
        writer.writerow(["商户实收总计", zh_data.get("商户实收总计", 0)])
    logger.info(f"  已保存CSV: {csv_file}")

    return zh_data


def select_unionpay_store(page, config):
    """在银豹付交易账单页面选择门店（Element UI Cascader 多选组件）。"""
    logger.info(f"  → 选择银豹付门店: {config['select_items']}")

    cascader = page.locator(".el-cascader")
    existing_tags = cascader.locator(".el-tag__close")
    tag_count = existing_tags.count()
    if tag_count > 0:
        logger.info(f"    清除已有选项 ({tag_count} 个)...")
        for _ in range(tag_count):
            close_btn = cascader.locator(".el-tag__close").first
            if close_btn.count() > 0:
                close_btn.click()
                page.wait_for_timeout(150)

    search_input = cascader.locator(".el-cascader__search-input")
    if search_input.count() > 0:
        search_input.click()
    else:
        cascader.locator(".el-input__inner").click(force=True)

    try:
        page.wait_for_selector(".el-cascader__dropdown, .el-popper .el-cascader-panel", state="visible", timeout=5000)
    except Exception:
        if search_input.count() > 0:
            search_input.click()
        else:
            cascader.locator(".el-input__inner").click(force=True)
        page.wait_for_selector(".el-cascader__dropdown, .el-popper .el-cascader-panel", state="visible", timeout=5000)

    cleared = page.evaluate("""
        (function() {
            var count = 0;
            var checked = document.querySelectorAll('.el-cascader-node .el-checkbox__input.is-checked');
            for (var i = 0; i < checked.length; i++) {
                checked[i].click();
                count++;
            }
            return count;
        })()
    """)
    if cleared > 0:
        logger.info(f"    清除面板中已选复选框: {cleared} 个")

    if config.get("parent_node"):
        parent = config["parent_node"]
        result = page.evaluate(f"""
            (function() {{
                var menus = document.querySelectorAll('.el-cascader-menu');
                if (menus.length === 0) return 'no cascader menu found';
                var firstMenu = menus[menus.length > 1 ? menus.length - 1 : 0];
                for (var m = 0; m < menus.length; m++) {{
                    var nodes = menus[m].querySelectorAll('.el-cascader-node');
                    for (var i = 0; i < nodes.length; i++) {{
                        var label = nodes[i].querySelector('.el-cascader-node__label');
                        var text = label ? label.textContent.trim() : nodes[i].textContent.trim();
                        if (text === '{parent}') {{
                            nodes[i].click();
                            return 'clicked parent: ' + text;
                        }}
                    }}
                }}
                var allLabels = [];
                var nodes = document.querySelectorAll('.el-cascader-node');
                for (var i = 0; i < nodes.length; i++) {{
                    var label = nodes[i].querySelector('.el-cascader-node__label');
                    allLabels.push(label ? label.textContent.trim() : nodes[i].textContent.trim());
                }}
                return 'parent not found: {parent}. Available: ' + allLabels.join(', ');
            }})()
        """)
        logger.info(f"    展开父节点: {result}")
        page.wait_for_timeout(200)

    for item_name in config["select_items"]:
        result = page.evaluate(f"""
            (function() {{
                var nodes = document.querySelectorAll('.el-cascader-node');
                for (var i = 0; i < nodes.length; i++) {{
                    var label = nodes[i].querySelector('.el-cascader-node__label');
                    var text = label ? label.textContent.trim() : nodes[i].textContent.trim();
                    if (text === '{item_name}') {{
                        var cb = nodes[i].querySelector('.el-checkbox__input');
                        if (cb) {{
                            cb.click();
                            return 'checked: ' + text;
                        }}
                        nodes[i].click();
                        return 'clicked: ' + text;
                    }}
                }}
                var allLabels = [];
                for (var i = 0; i < nodes.length; i++) {{
                    var label = nodes[i].querySelector('.el-cascader-node__label');
                    allLabels.push(label ? label.textContent.trim() : nodes[i].textContent.trim());
                }}
                return 'not found: {item_name}. Available: ' + allLabels.join(', ');
            }})()
        """)
        logger.info(f"    选择: {result}")
        page.wait_for_timeout(150)

    page.locator("body").click(position={"x": 0, "y": 0})
    page.wait_for_selector(".el-cascader__dropdown", state="hidden", timeout=5000)


def select_store_type_consumption(page):
    """在会员消费汇总表页面选择门店类型为'消费门店'。"""
    logger.info("  → 选择门店类型: 消费门店")

    selector = page.locator("[p-single-selector='userTypeOpts']")
    if selector.count() > 0:
        selector.click()
        page.wait_for_selector("[p-single-selector='userTypeOpts'] .selectBox", state="visible", timeout=5000)

    result = page.evaluate("""
        (function() {
            var box = document.querySelector('[p-single-selector="userTypeOpts"] .selectBox');
            if (box) box.style.display = 'block';
            var lis = document.querySelectorAll('[p-single-selector="userTypeOpts"] .selectBox li');
            for (var i = 0; i < lis.length; i++) {
                if (lis[i].textContent.trim() === '消费门店') {
                    lis[i].click();
                    return 'clicked: 消费门店 (optionvalue=' + lis[i].getAttribute('optionvalue') + ')';
                }
            }
            var names = [];
            for (var i = 0; i < lis.length; i++) names.push(lis[i].textContent.trim());
            return 'not found in: ' + names.join(', ');
        })()
    """)
    logger.info(f"  门店类型: {result}")


def select_stores_multi(page, store_names):
    """在 #queryStoreDiv 中选择多个门店。"""
    logger.info(f"  → 选择门店范围: {store_names}")

    page.evaluate("""
        (function() {
            var lis = document.querySelectorAll('#queryStoreDiv li');
            for (var i = 0; i < lis.length; i++) {
                if (lis[i].classList.contains('on') || lis[i].classList.contains('selected')) {
                    lis[i].click();
                }
            }
        })()
    """)

    for name in store_names:
        result = page.evaluate(f"""
            (function() {{
                var lis = document.querySelectorAll('#queryStoreDiv li');
                for (var i = 0; i < lis.length; i++) {{
                    var text = lis[i].textContent.trim();
                    if (text === '{name}') {{
                        lis[i].click();
                        return 'clicked: ' + text + ' (data=' + lis[i].getAttribute('data') + ')';
                    }}
                }}
                var names = [];
                for (var i = 0; i < lis.length; i++) names.push(lis[i].textContent.trim());
                return 'not found: ' + name + '. Available: ' + names.join(', ');
            }})()
        """)
        logger.info(f"    {result}")


def set_vue_date(page, target_str):
    """在 Vue/Element UI 页面设置日期。"""
    date_dash = target_str.replace(".", "-")

    range_inputs = page.locator("input.el-range-input")
    if range_inputs.count() >= 2:
        for idx in range(2):
            range_inputs.nth(idx).click()
            range_inputs.nth(idx).press("Control+a")
            range_inputs.nth(idx).type(date_dash, delay=50)
        range_inputs.nth(1).press("Enter")
        logger.info(f"  日期已设置(range): {date_dash}")
        return

    date_inputs = page.locator(".el-date-editor input.el-input__inner")
    if date_inputs.count() >= 2:
        for idx in range(min(date_inputs.count(), 2)):
            date_inputs.nth(idx).click()
            date_inputs.nth(idx).press("Control+a")
            date_inputs.nth(idx).type(date_dash, delay=50)
            date_inputs.nth(idx).press("Enter")
        logger.info(f"  日期已设置(editor): {date_dash}")
        return

    logger.warning("  未找到日期输入框，尝试 fallback...")
    page.evaluate(f"""
        (function() {{
            var inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                var ph = (inputs[i].placeholder || '');
                if (ph.indexOf('日期') >= 0 || ph.indexOf('date') >= 0) {{
                    var nativeSet = Object.getOwnPropertyDescriptor(
                        window.HTMLInputElement.prototype, 'value').set;
                    nativeSet.call(inputs[i], '{date_dash}');
                    inputs[i].dispatchEvent(new Event('input', {{bubbles: true}}));
                    inputs[i].dispatchEvent(new Event('change', {{bubbles: true}}));
                }}
            }}
        }})()
    """)
    logger.info(f"  日期已设置(fallback): {date_dash}")


def douyin_set_date(page, target):
    """设置抖音每日收益页面的核销日期（范围选择器，起止日期相同）。"""
    target_year = target.year
    target_month = target.month
    target_day = str(target.day)
    date_str = target.strftime("%Y-%m-%d")

    logger.info(f"  → 设置核销日期: {date_str} ~ {date_str}...")

    # 等待页面稳定后再操作日期选择器
    for attempt in range(3):
        try:
            page.wait_for_selector('.byted-date-picker', timeout=10000)
            result = page.evaluate("""
                (function() {
                    var picker = document.querySelector('.byted-date-picker');
                    if (!picker) return 'picker not found';
                    var trigger = picker.querySelector('.byted-popper-trigger');
                    if (trigger) { trigger.click(); return 'clicked trigger'; }
                    var input = picker.querySelector('input.byted-input');
                    if (input) { input.click(); return 'clicked input'; }
                    return 'no clickable element found';
                })()
            """)
            logger.info(f"    打开日期选择器: {result}")
            break
        except Exception as e:
            if attempt < 2:
                logger.warning(f"    日期选择器操作失败(重试 {attempt + 1}/3): {e}")
                page.wait_for_timeout(1000)
            else:
                raise

    try:
        page.wait_for_selector('.byted-date-container', timeout=5000)
    except Exception:
        logger.warning("  日期弹窗未出现，尝试再次点击...")
        page.evaluate("""
            (function() {
                var picker = document.querySelector('.byted-date-picker');
                if (picker) picker.click();
            })()
        """)
        try:
            page.wait_for_selector('.byted-date-container', timeout=5000)
        except Exception:
            logger.error("  日期弹窗仍未出现，跳过日期设置")
            return

    days_diff = (datetime.now().date() - target.date()).days
    if days_diff == 0:
        page.locator('.byted-date-panel-preset-item a', has_text='今天').click()
        logger.info(f"  使用预设「今天」")
        page.wait_for_selector('.byted-date-container', state='hidden', timeout=5000)
        return
    if days_diff == 1:
        page.locator('.byted-date-panel-preset-item a', has_text='昨天').click()
        logger.info(f"  使用预设「昨天」")
        page.wait_for_selector('.byted-date-container', state='hidden', timeout=5000)
        return

    current = page.evaluate("""
        (function() {
            var views = document.querySelectorAll('.byted-date-view');
            var r = {};
            for (var i = 0; i < views.length; i++) {
                var titles = views[i].querySelectorAll('.byted-date-title-item');
                if (titles.length < 2) continue;
                var y = parseInt(titles[0].textContent), m = parseInt(titles[1].textContent);
                if (views[i].classList.contains('byted-date-position-start')) {
                    r.sy = y; r.sm = m;
                } else if (views[i].classList.contains('byted-date-position-end')) {
                    r.ey = y; r.em = m;
                }
            }
            return r;
        })()
    """)
    logger.info(f"    当前日历面板: {current}")

    if not current or 'sy' not in current:
        logger.warning("  无法读取日历面板年月")
        return

    if current['sy'] == target_year and current['sm'] == target_month:
        panel = '.byted-date-position-start'
    elif current.get('ey') == target_year and current.get('em') == target_month:
        panel = '.byted-date-position-end'
    else:
        months_diff = (current['sy'] - target_year) * 12 + (current['sm'] - target_month)
        if months_diff > 0:
            logger.info(f"  往前导航 {months_diff} 个月...")
            for _ in range(months_diff):
                page.locator('.byted-date-position-start .byted-icon-left-o').click()
                page.wait_for_timeout(200)
        elif months_diff < 0:
            logger.info(f"  往后导航 {-months_diff} 个月...")
            for _ in range(-months_diff):
                page.locator('.byted-date-position-end .byted-icon-right-o').click()
                page.wait_for_timeout(200)
        page.wait_for_timeout(200)
        panel = '.byted-date-position-start'

    selector = f"{panel} .byted-date-item:not(.byted-date-grid-prev):not(.byted-date-grid-next):not(.byted-date-disabled)"
    for click_label in ("start", "end"):
        items = page.locator(selector)
        count = items.count()
        clicked = False
        for i in range(count):
            item = items.nth(i)
            if item.text_content().strip() == target_day:
                item.click()
                clicked = True
                logger.info(f"    clicked {click_label}: day {target_day}")
                break
        if not clicked:
            logger.warning(f"    {click_label}: day {target_day} not found in panel")
        page.wait_for_timeout(300)


def douyin_select_store(page, store_config):
    """在抖音每日收益页面选择门店（按省市筛选弹窗）。"""
    store_short = store_config["store_short"]
    search_keyword = store_config["search_keyword"]

    logger.info(f"  → 选择门店: {store_short} (搜索: {search_keyword})...")

    # 点击门店选择器（poi_dropdown 触发器）
    result = page.evaluate("""
        (function() {
            var trigger = document.querySelector('.poi_dropdown_4a1a7');
            if (trigger) { trigger.click(); return 'clicked poi_dropdown'; }
            var container = document.querySelector('[class*="accountSelect"]');
            if (!container) container = document.querySelector('[class*="indicators"]');
            if (!container) return 'container not found';
            var detail = container.querySelector('[class*="detail-indicator"]');
            if (detail) { detail.click(); return 'clicked detail-indicator'; }
            var triggers = container.querySelectorAll('.byted-dropdown-trigger');
            if (triggers.length >= 2) { triggers[1].click(); return 'clicked trigger[1]'; }
            return 'trigger not found';
        })()
    """)
    logger.info(f"    打开门店选择器: {result}")
    time.sleep(1)

    try:
        page.wait_for_selector('.menu_4a1a7', state="visible", timeout=5000)
    except Exception:
        logger.warning("    门店弹窗未出现")

    # 先清除已有选项
    result = page.evaluate("""
        (function() {
            var els = document.querySelectorAll('[class*="clean-btn"], [class*="clean"]');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '清除') {
                    els[i].click();
                    return 'clicked';
                }
            }
            var spans = document.querySelectorAll('span');
            for (var i = 0; i < spans.length; i++) {
                if (spans[i].textContent.trim() === '清除' && spans[i].children.length === 0) {
                    spans[i].click();
                    return 'clicked span';
                }
            }
            return 'not found';
        })()
    """)
    logger.info(f"    清除已有选项: {result}")
    time.sleep(1)

    # 搜索门店
    result = page.evaluate("""
        (function() {
            var input = document.querySelector('input[placeholder*="门店"]');
            if (!input) return 'search input not found';
            var nativeInputValueSetter = Object.getOwnPropertyDescriptor(
                window.HTMLInputElement.prototype, 'value').set;
            nativeInputValueSetter.call(input, '');
            input.dispatchEvent(new Event('input', { bubbles: true }));
            return 'cleared';
        })()
    """)
    logger.info(f"    清空搜索框: {result}")
    time.sleep(0.5)

    result = page.evaluate(f"""
        (function() {{
            var input = document.querySelector('input[placeholder*="门店"]');
            if (!input) return 'search input not found';
            input.focus();
            var nativeInputValueSetter = Object.getOwnPropertyDescriptor(
                window.HTMLInputElement.prototype, 'value').set;
            nativeInputValueSetter.call(input, '{search_keyword}');
            input.dispatchEvent(new Event('input', {{ bubbles: true }}));
            return 'filled: {search_keyword}';
        }})()
    """)
    logger.info(f"    搜索门店: {result}")
    time.sleep(3)
    page.wait_for_selector('[class*="account-select-item"]', timeout=10000)

    # 选择第一个可用的门店 checkbox
    result = page.evaluate("""
        (function() {
            var items = document.querySelectorAll('[class*="account-select-item-pc__warpper"], [class*="account-select-item"]');
            for (var i = 0; i < items.length; i++) {
                var cb = items[i].querySelector('.byted-checkbox');
                if (!cb) continue;
                if (cb.classList.contains('byted-checkbox-disabled')) continue;
                var icon = cb.querySelector('.byted-checkbox-icon');
                if (icon) { icon.click(); }
                else { cb.click(); }
                var nameEl = items[i].querySelector('span[style*="overflow"]');
                return 'checked: ' + (nameEl ? nameEl.textContent.trim() : 'unknown');
            }
            // 如果都是disabled，尝试强制点击第一个
            for (var i = 0; i < items.length; i++) {
                var cb = items[i].querySelector('.byted-checkbox');
                if (!cb) continue;
                var input = cb.querySelector('input[type="checkbox"]');
                if (input) {
                    input.disabled = false;
                    input.click();
                    var nameEl = items[i].querySelector('span[style*="overflow"]');
                    return 'force-checked: ' + (nameEl ? nameEl.textContent.trim() : 'unknown');
                }
            }
            return 'no checkbox found';
        })()
    """)
    logger.info(f"    选择门店: {result}")
    time.sleep(1)

    # 点击确认
    result = page.evaluate("""
        (function() {
            var footer = document.querySelector('.account-select-footer-pc');
            if (footer) {
                var btns = footer.querySelectorAll('button');
                for (var i = 0; i < btns.length; i++) {
                    if (btns[i].textContent.trim() === '确认') {
                        btns[i].click();
                        return 'clicked';
                    }
                }
            }
            var btns = document.querySelectorAll('button.byted-btn-type-primary');
            for (var i = 0; i < btns.length; i++) {
                if (btns[i].textContent.trim() === '确认') {
                    btns[i].click();
                    return 'clicked fallback';
                }
            }
            return 'not found';
        })()
    """)
    logger.info(f"    确认: {result}")
    page.wait_for_selector('.menu_4a1a7', state="hidden", timeout=10000)
    time.sleep(3)


def scrape_douyin_daily_benefits(page):
    """从抖音每日收益页面抓取订单实收和佣金/服务费支出。"""
    data = page.evaluate("""
        (function() {
            var result = {};
            var containers = document.querySelectorAll('[class*="container--"]');
            for (var i = 0; i < containers.length; i++) {
                var textEl = containers[i].querySelector('[class*="text--"]');
                var numEl = containers[i].querySelector('[class*="number--"]');
                if (!textEl || !numEl) continue;
                var label = textEl.textContent.trim();
                var value = parseFloat(numEl.textContent.trim().replace(/,/g, '')) || 0;
                if (label === '订单实收' || label === '佣金/服务费支出') {
                    result[label] = value;
                }
            }
            return result;
        })()
    """)
    if not data:
        logger.warning("  未抓取到抖音数据，可能页面未加载完成")
    else:
        logger.info(f"  抓取到抖音数据: {json.dumps(data, ensure_ascii=False)}")
    return data


def save_douyin_daily_benefits_csv(data, store_short, date_label, output_dir):
    """将抖音每日收益数据保存为CSV。"""
    csv_file = output_dir / f"抖音每日收益_{store_short}_{date_label}.csv"
    with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["项目", "金额"])
        for key in DOUYIN_CSV_FIELDS:
            writer.writerow([key, data.get(key, 0)])
    logger.info(f"  已保存CSV: {csv_file.name}")
    return csv_file


def dismiss_koubei_popup(page):
    """检查并关闭高德口碑页面可能弹出的「设置账号信息」弹窗。"""
    try:
        modal = page.locator('.crmhome-main-modal-wrap')
        if modal.count() > 0 and modal.is_visible():
            cancel_btn = modal.locator('button.crmhome-main-btn-default')
            if cancel_btn.count() > 0:
                cancel_btn.click()
                logger.info("  已关闭「设置账号信息」弹窗")
                page.wait_for_selector('.crmhome-main-modal-wrap', state='hidden', timeout=5000)
    except Exception:
        pass


def koubei_set_date(page, target):
    """在高德口碑账单汇总页面设置日期范围（起止日期相同）。"""
    date_start = target.strftime("%Y-%m-%d") + " 00:00:00"
    date_end = target.strftime("%Y-%m-%d") + " 23:59:59"
    logger.info(f"  → 设置日期: {date_start} ~ {date_end}...")
    time.sleep(5)

    start_input = page.locator('input[date-range="start"]')
    start_input.click()
    page.wait_for_selector('.aamf-picker-dropdown', state='visible', timeout=20000)

    start_input.press("Control+a")
    start_input.type(date_start, delay=30)

    logger.info("    点击确定（开始时间）...")
    page.locator('.aamf-picker-dropdown .aamf-picker-ok button').click()
    time.sleep(0.5)

    end_input = page.locator('input[date-range="end"]')
    end_input.click()
    page.wait_for_selector('.aamf-picker-dropdown', state='visible', timeout=10000)
    end_input.press("Control+a")
    end_input.type(date_end, delay=30)

    logger.info("    点击确定（结束时间）...")
    page.locator('.aamf-picker-dropdown .aamf-picker-ok button').click()
    time.sleep(0.5)

    logger.info(f"  日期已设置: {date_start} ~ {date_end}")


def koubei_select_query_type(page):
    """在高德口碑页面将查询方式切换为"按门店查询"。"""
    logger.info("  → 选择查询方式: 按门店查询...")

    select_el = page.locator('.aamf-select-single').first
    select_el.click()
    page.wait_for_selector('.aamf-select-item', state='visible', timeout=5000)
    time.sleep(0.5)

    result = page.evaluate("""
        (function() {
            var items = document.querySelectorAll('.aamf-select-item');
            for (var i = 0; i < items.length; i++) {
                if (items[i].textContent.trim() === '按门店查询') {
                    items[i].click();
                    return 'selected: 按门店查询';
                }
            }
            var names = [];
            for (var i = 0; i < items.length; i++) names.push(items[i].textContent.trim());
            return 'not found. Available: ' + names.join(', ');
        })()
    """)
    logger.info(f"    查询方式: {result}")
    time.sleep(1)


def koubei_select_store(page, store_config):
    """在高德口碑账单汇总页面选择单个门店（通过弹窗选择）。"""
    store_name = store_config["store_name"]
    logger.info(f"  → 选择门店: {store_name}...")

    store_input = page.locator('input[placeholder="请选择"][readonly]')
    store_input.click(force=True)
    time.sleep(0.5)

    try:
        page.wait_for_selector('.aamf-modal', timeout=20000)
    except Exception:
        logger.warning("  门店选择弹窗未出现，尝试再次点击...")
        store_input.click(force=True)
        page.wait_for_selector('.aamf-modal', timeout=20000)

    page.wait_for_selector('.aamf-modal .aamf-table-tbody tr.aamf-table-row', timeout=10000)

    page.evaluate("""
        (function() {
            var rows = document.querySelectorAll('.aamf-modal .aamf-table-tbody tr.aamf-table-row');
            for (var i = 0; i < rows.length; i++) {
                var cb = rows[i].querySelector('.aamf-checkbox-input');
                if (cb && cb.checked) {
                    cb.click();
                }
            }
        })()
    """)

    result = page.evaluate(f"""
        (function() {{
            var rows = document.querySelectorAll('.aamf-modal .aamf-table-tbody tr.aamf-table-row');
            for (var i = 0; i < rows.length; i++) {{
                var tds = rows[i].querySelectorAll('td.aamf-table-cell');
                for (var j = 0; j < tds.length; j++) {{
                    if (tds[j].textContent.trim() === '{store_name}') {{
                        var cb = rows[i].querySelector('.aamf-checkbox-input');
                        if (cb && !cb.checked) cb.click();
                        return 'checked: ' + tds[j].textContent.trim();
                    }}
                }}
            }}
            var names = [];
            for (var i = 0; i < rows.length; i++) {{
                var tds = rows[i].querySelectorAll('td.aamf-table-cell');
                if (tds.length >= 3) names.push(tds[2].textContent.trim());
            }}
            return 'not found: {store_name}. Available: ' + names.join(', ');
        }})()
    """)
    logger.info(f"    选择门店: {result}")
    time.sleep(0.5)

    confirm_btn = page.locator('.aamf-modal-footer button.aamf-btn-primary')
    confirm_btn.click()
    page.wait_for_selector('.aamf-modal', state='hidden', timeout=10000)
    logger.info("    弹窗已确认")


def scrape_koubei_bill(page):
    """从高德口碑账单汇总页面抓取汇总区域的数据。"""
    data = page.evaluate("""
        (function() {
            var result = {};
            var nameEls = document.querySelectorAll('.name--pnzqmgDZ');
            var valueEls = document.querySelectorAll('.value--NiKhfQ57');
            for (var i = 0; i < nameEls.length; i++) {
                var rawName = nameEls[i].textContent.trim();
                var name = rawName.replace(/（元）/g, '').replace(/\\(元\\)/g, '');
                if (i < valueEls.length) {
                    var val = parseFloat(valueEls[i].textContent.trim().replace(/,/g, '')) || 0;
                    result[name] = val;
                }
            }
            return result;
        })()
    """)
    if not data:
        logger.warning("  未抓取到口碑数据，可能页面未加载完成")
    else:
        logger.info(f"  抓取到口碑数据: {json.dumps(data, ensure_ascii=False)}")
    return data


def save_koubei_bill_csv(data, store_short, date_label, output_dir):
    """将高德口碑账单汇总数据保存为 CSV。"""
    csv_file = output_dir / f"高德口碑账单汇总_{store_short}_{date_label}.csv"
    with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["项目", "金额"])
        for key in KOUBEI_CSV_FIELDS:
            writer.writerow([key, data.get(key, 0)])
    logger.info(f"  已保存CSV: {csv_file.name}")
    return csv_file


def _launch_chrome(port, user_data_dir, headless=False):
    process = subprocess.Popen([
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        *(["--headless=new"] if headless else []),
        f"--remote-debugging-port={port}",
        f"--user-data-dir={user_data_dir}",
    ])
    time.sleep(3)
    return process


def _kill_process_tree(process):
    try:
        subprocess.run(
            ["taskkill", "/F", "/T", "/PID", str(process.pid)],
            capture_output=True,
            timeout=10,
        )
    except Exception:
        try:
            process.kill()
        except Exception:
            pass


def _cleanup_chrome(page, browser, process):
    if page:
        try:
            page.close()
        except Exception:
            pass
    if browser:
        try:
            browser.close()
        except Exception:
            pass
    if process:
        _kill_process_tree(process)


# ─── Group A: 银豹系统（Playwright 浏览器）──────────────────────────────────────


def run_pospal_tasks(args, target, target_str, date_label, output_dir):
    TAG = "[银豹]"
    from playwright.sync_api import sync_playwright

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=args.headless,
            args=["--start-maximized"],
        )
        context = browser.new_context(
            viewport={"width": 1600, "height": 900},
            accept_downloads=True,
        )
        page = context.new_page()
        page.set_default_timeout(120000)
        page.set_default_navigation_timeout(120000)

        try:
            login(page)

            # ── 营业概况 ─────────────────────────────
            logger.info(f"{TAG} {'─' * 45}")
            logger.info(f"{TAG}  下载营业概况日度统计")
            logger.info(f"{TAG} {'─' * 45}")

            for i, cs_config in enumerate(CUSTOMER_SUMMARY_STORE_CONFIG):
                store_short = cs_config["store_short"]
                desc = f"营业概况-{store_short}"

                def _do_business_summary(cfg=cs_config, ss=store_short):
                    logger.info(f"{TAG}  门店 {i + 1}/{len(CUSTOMER_SUMMARY_STORE_CONFIG)}: {ss}")
                    page.goto(BUSINESS_SUMMARY_URL)
                    page.wait_for_load_state("networkidle", timeout=120_000)
                    select_stores_ddl(page, cfg["select_items"])
                    set_date(page, "开始日期", f"{target_str} 00:00")
                    set_date(page, "结束日期", f"{target_str} 23:59")
                    click_by_text(page, "查询", "查询")
                    page.wait_for_load_state("networkidle", timeout=150_000)
                    time.sleep(3)
                    with page.expect_download(timeout=180_000) as dl_info:
                        click_export(page)
                    download = dl_info.value
                    dest = output_dir / f"营业概况_{ss}_{date_label}.xlsx"
                    download.save_as(dest)
                    logger.info(f"{TAG}  已保存: {dest.name}")

                retry_until_success(_do_business_summary, desc)

            logger.info(f"{TAG}  营业概况下载全部完成！")

            # ── 银豹付交易账单 ─────────────────────────────
            logger.info(f"{TAG} {'─' * 45}")
            logger.info(f"{TAG}  下载银豹付交易账单")
            logger.info(f"{TAG} {'─' * 45}")

            for i, up_config in enumerate(UNIONPAY_STORE_CONFIG):
                store_short = up_config["store_short"]
                desc = f"银豹付-{store_short}"

                def _do_unionpay(cfg=up_config, ss=store_short):
                    logger.info(f"{TAG}  门店 {i + 1}/{len(UNIONPAY_STORE_CONFIG)}: {ss}")
                    page.goto(UNIONPAY_BILL_URL)
                    page.wait_for_load_state("networkidle", timeout=120_000)
                    page.wait_for_selector(".el-cascader", timeout=30000)
                    set_vue_date(page, target_str)
                    time.sleep(1)
                    select_unionpay_store(page, cfg)
                    time.sleep(0.5)
                    click_by_text(page, "搜索", "搜索")
                    page.wait_for_load_state("networkidle", timeout=150_000)
                    page.wait_for_selector(".data-dz-item", timeout=30000)
                    time.sleep(5)
                    bill_data = scrape_unionpay_bill(page)
                    save_unionpay_bill_csv(bill_data, ss, date_label, output_dir)

                retry_until_success(_do_unionpay, desc)

            logger.info(f"{TAG}  银豹付交易账单下载全部完成！")

            # ── 会员消费汇总表 ─────────────────────────────
            logger.info(f"{TAG} {'─' * 45}")
            logger.info(f"{TAG}  下载会员消费汇总表")
            logger.info(f"{TAG} {'─' * 45}")

            for i, cs_config in enumerate(CUSTOMER_SUMMARY_STORE_CONFIG):
                store_short = cs_config["store_short"]
                desc = f"会员消费-{store_short}"

                def _do_customer(cfg=cs_config, ss=store_short):
                    logger.info(f"{TAG}  门店 {i + 1}/{len(CUSTOMER_SUMMARY_STORE_CONFIG)}: {ss}")
                    page.goto(CUSTOMER_SUMMARY_URL)
                    page.wait_for_load_state("networkidle", timeout=120_000)
                    select_store_type_consumption(page)
                    select_stores_multi(page, cfg["select_items"])
                    date_dash = target_str.replace(".", "-")
                    result = page.evaluate(f"""
                        (function() {{
                            function setVal(id, val) {{
                                var inp = document.getElementById(id);
                                if (!inp) return 'not found: ' + id;
                                if (window.jQuery && jQuery.fn.datepicker) {{
                                    jQuery(inp).datepicker('setDate', val);
                                }}
                                var nativeSet = Object.getOwnPropertyDescriptor(
                                    HTMLInputElement.prototype, 'value').set;
                                nativeSet.call(inp, val);
                                inp.dispatchEvent(new Event('input', {{bubbles: true}}));
                                inp.dispatchEvent(new Event('change', {{bubbles: true}}));
                                inp.dispatchEvent(new Event('blur', {{bubbles: true}}));
                                return inp.value;
                            }}
                            var r1 = setVal('txt_startDatetime', '{date_dash}');
                            var r2 = setVal('txt_endDatetime', '{date_dash}');
                            return 'start=' + r1 + ', end=' + r2;
                        }})()
                    """)
                    logger.info(f"{TAG}  统计时间: {result}")
                    page.locator("#btnSearch").click()
                    page.wait_for_load_state("networkidle", timeout=150_000)
                    time.sleep(2)
                    click_by_text(page, "导出销售单据", "导出销售单据")
                    time.sleep(1)
                    with page.expect_download(timeout=180_000) as dl_info:
                        click_by_text(page, "导出", "弹窗导出")
                    download = dl_info.value
                    dest = output_dir / f"会员消费汇总表_{ss}_{date_label}.xlsx"
                    download.save_as(dest)
                    logger.info(f"{TAG}  已保存: {dest.name}")

                retry_until_success(_do_customer, desc)

            logger.info(f"{TAG}  会员消费汇总表下载全部完成！")

        finally:
            context.close()
            browser.close()


# ─── Group B: 美团外卖（每店独立 Chrome）──────────────────────────────────────


def _run_meituan_waimai_store(mt_config, args, target_str, date_label, output_dir):
    mt_store_short = mt_config["store_short"]
    mt_port = mt_config["port"]
    mt_user_data_dir = mt_config["user_data_dir"]
    TAG = f"[美团外卖-{mt_store_short}]"

    from playwright.sync_api import sync_playwright

    chrome_process = _launch_chrome(mt_port, mt_user_data_dir, args.headless)
    chrome_browser = None
    chrome_page = None

    with sync_playwright() as pw:
        try:
            logger.info(f"{TAG}  连接到 Chrome (port={mt_port})...")
            chrome_browser, chrome_page = _connect_cdp_with_retry(pw, mt_port)

            # ── 美团外卖账单 ──────────────────────────────
            def _do_meituan_bill():
                logger.info(f"{TAG}  [导航] 先进入美团外卖结算账单页...")
                chrome_page.goto(MEITUAN_DOWNLOAD_URL_PRE)
                chrome_page.wait_for_selector('#hashframe', timeout=15000)
                time.sleep(2)
                # chrome_page.wait_for_load_state("networkidle", timeout=60000)

                logger.info(f"{TAG}  [导航] 前往美团外卖账单明细页...")
                chrome_page.goto(MEITUAN_DOWNLOAD_URL)
                chrome_page.wait_for_selector('.select-input-wrapper', timeout=15000)
                time.sleep(2)
                # chrome_page.wait_for_load_state("networkidle", timeout=60000)

                date_input_value = f"{date_label} 日账单"
                logger.info(f"{TAG}  → 设置日期: {date_input_value}...")
                date_input = chrome_page.locator(".select-input-wrapper .roo-input")
                date_input.click()
                time.sleep(0.3)
                date_input.press("Control+a")
                date_input.type(date_input_value, delay=50)
                time.sleep(0.5)
                date_input.press("Enter")
                time.sleep(0.3)
                chrome_page.locator("body").click(position={"x": 0, "y": 0})

                logger.info(f"{TAG}  等待数据刷新...")
                chrome_page.wait_for_selector('.bill-charge-table tfoot', timeout=30000)
                time.sleep(5)

                logger.info(f"{TAG}  → 抓取账单数据...")
                bill_data = chrome_page.evaluate("""
                    (function() {
                        var result = {};
                        var tfoot = document.querySelector('.bill-charge-table tfoot tr');
                        if (tfoot) {
                            var tds = tfoot.querySelectorAll('td');
                            result['商品总价'] = parseFloat(tds[1].textContent.trim().replace(/,/g, '')) || 0;
                            result['打包费'] = parseFloat(tds[2].textContent.trim().replace(/,/g, '')) || 0;
                            result['商家对顾客的活动补贴'] = parseFloat(tds[3].textContent.trim().replace(/,/g, '')) || 0;
                            result['佣金'] = parseFloat(tds[6].textContent.trim().replace(/,/g, '')) || 0;
                            result['配送服务费'] = parseFloat(tds[7].textContent.trim().replace(/,/g, '')) || 0;
                        }
                        var tabs = document.querySelectorAll('.roo-tabs-nav .tab-item a');
                        for (var i = 0; i < tabs.length; i++) {
                            var text = tabs[i].textContent.trim();
                            if (text.indexOf('其它类') >= 0 || text.indexOf('其他类') >= 0) {
                                var match = text.match(/([-\\d.]+)\\s*$/);
                                result['其他类'] = match ? parseFloat(match[1]) : 0;
                                break;
                            }
                        }
                        return result;
                    })()
                """)
                logger.info(f"{TAG}  抓取数据: {json.dumps(bill_data, ensure_ascii=False)}")

                meituan_csv_fields = ["商品总价", "打包费", "商家对顾客的活动补贴", "佣金", "配送服务费", "其他类"]
                csv_file = output_dir / f"账单明细_{mt_store_short}_{date_label}.csv"
                with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)
                    writer.writerow(["项目", "金额"])
                    for key in meituan_csv_fields:
                        writer.writerow([key, bill_data.get(key, 0)])
                logger.info(f"{TAG}  已保存CSV: {csv_file.name}")

            retry_until_success(_do_meituan_bill, f"美团外卖账单-{mt_store_short}")

            # ── 美团推广消费 ──────────────────────────────
            def _do_meituan_ad():
                logger.info(f"{TAG}  [导航] 前往美团推广账户详情页...")
                chrome_page.goto(MEITUAN_AD_URL)
                chrome_page.wait_for_load_state("networkidle", timeout=120_000)
                chrome_page.wait_for_selector('.panel-body table tbody tr', timeout=30000)
                time.sleep(5)

                logger.info(f"{TAG}  → 逐页查找 {date_label} 的推广消费数据...")
                ad_found_rows = []
                target_date = datetime.strptime(date_label, "%Y-%m-%d").date()
                prev_page_key = None

                for ad_page_num in range(1, 100):
                    if ad_page_num > 1:
                        logger.info(f"{TAG}  翻到第 {ad_page_num} 页...")
                        chrome_page.evaluate(f"""
                            (function() {{
                                var input = document.querySelector('.jump input.form-control');
                                var nativeSet = Object.getOwnPropertyDescriptor(
                                    HTMLInputElement.prototype, 'value').set;
                                nativeSet.call(input, '{ad_page_num}');
                                input.dispatchEvent(new Event('input', {{bubbles: true}}));
                                input.dispatchEvent(new Event('change', {{bubbles: true}}));
                                document.querySelector('.jump button.btn').click();
                            }})()
                        """)
                        chrome_page.wait_for_load_state("networkidle", timeout=30000)
                        time.sleep(5)

                    page_rows = chrome_page.evaluate("""
                        (function() {
                            var rows = document.querySelectorAll('.panel-body table tbody tr');
                            var items = [];
                            for (var i = 0; i < rows.length; i++) {
                                var tds = rows[i].querySelectorAll('td');
                                if (tds.length < 4) continue;
                                var rawDate = tds[0].textContent.trim();
                                var amountSpan = tds[2].querySelector('span');
                                items.push({
                                    rawDate: rawDate,
                                    type: tds[1].textContent.trim(),
                                    amount: (amountSpan ? amountSpan.textContent.trim() : tds[2].textContent.trim())
                                });
                            }
                            return items;
                        })()
                    """)

                    if not page_rows:
                        logger.info(f"{TAG}  第 {ad_page_num} 页无数据，停止搜索")
                        break

                    cur_page_key = [(r["rawDate"], r["type"], r["amount"]) for r in page_rows]
                    if cur_page_key == prev_page_key:
                        logger.info(f"{TAG}  第 {ad_page_num} 页数据与上页相同（已到最后一页），停止搜索")
                        break
                    prev_page_key = cur_page_key

                    logger.info(f"{TAG}  第 {ad_page_num} 页: {len(page_rows)} 行, "
                                f"首行日期={page_rows[0]['rawDate']}, 末行日期={page_rows[-1]['rawDate']}")

                    date_passed = False
                    for row in page_rows:
                        row_date_match = re.search(r"(\d{4})\D(\d{1,2})\D(\d{1,2})", row["rawDate"])
                        if not row_date_match:
                            continue
                        row_date = datetime(
                            int(row_date_match.group(1)),
                            int(row_date_match.group(2)),
                            int(row_date_match.group(3)),
                        ).date()
                        if row_date == target_date:
                            amount = float(row["amount"].replace(",", "")) if row["amount"] else 0
                            if amount < 0:
                                ad_found_rows.append({
                                    "时间": row["rawDate"],
                                    "类型": row["type"],
                                    "变化金额(元)": amount,
                                })
                                logger.info(f"{TAG}  第 {ad_page_num} 页命中: {row['type']}, 金额={amount}")
                        if row_date < target_date:
                            date_passed = True

                    if date_passed:
                        logger.info(f"{TAG}  已翻过目标日期 {date_label}，停止搜索")
                        break

                if ad_found_rows:
                    ad_total = sum(r["变化金额(元)"] for r in ad_found_rows)
                    logger.info(f"{TAG}  推广消费数据: {len(ad_found_rows)} 条, 合计变化金额={ad_total}")
                    ad_csv_file = output_dir / f"美团外卖推广消费_{mt_store_short}_{date_label}.csv"
                    with open(ad_csv_file, "w", newline="", encoding="utf-8-sig") as f:
                        writer = csv.writer(f)
                        writer.writerow(["时间", "类型", "变化金额(元)"])
                        for r in ad_found_rows:
                            writer.writerow([r["时间"], r["类型"], r["变化金额(元)"]])
                    logger.info(f"{TAG}  已保存推广消费CSV: {ad_csv_file.name}")
                else:
                    logger.warning(f"{TAG}  未找到 {date_label} 的推广消费数据（该日可能无推广消费）")

            retry_until_success(_do_meituan_ad, f"美团推广消费-{mt_store_short}")

            # ── 饿了么账单 ──────────────────────────────
            if mt_store_short in ELEME_STORE_CONFIG:
                def _do_eleme():
                    logger.info(f"{TAG}  [导航] 前往饿了么账单页面...")
                    chrome_page.goto(ELEME_BILL_URL)
                    chrome_page.wait_for_selector('input[placeholder="开始日期"]', timeout=30000)
                    time.sleep(3)

                    def _eleme_set_date_and_query():
                        logger.info(f"{TAG}  → 设置账单日期: {date_label}...")
                        start_input = chrome_page.locator('input[placeholder="开始日期"]')
                        start_input.click()
                        start_input.press("Control+a")
                        start_input.type(date_label, delay=50)
                        time.sleep(0.5)

                        end_input = chrome_page.locator('input[placeholder="结束日期"]')
                        end_input.click()
                        end_input.press("Control+a")
                        end_input.type(date_label, delay=50)
                        time.sleep(0.5)
                        end_input.press("Enter")

                        logger.info(f"{TAG}  [查询] 点击查询...")
                        chrome_page.locator("button.cook-btn-primary").click()
                        chrome_page.wait_for_selector('.ant-table-thead', timeout=30000)
                        time.sleep(5)

                    def _eleme_scrape():
                        return chrome_page.evaluate("""
                            (function() {
                                var result = {};
                                var thead = document.querySelector('.ant-table-thead');
                                if (!thead) return result;
                                var rows = thead.querySelectorAll('tr');
                                if (rows.length < 2) return result;
                                var ths = rows[1].querySelectorAll('th');
                                if (ths.length >= 4) {
                                    result['结算金额'] = parseFloat(ths[1].textContent.trim().replace(/,/g, '')) || 0;
                                    result['订单类'] = parseFloat(ths[2].textContent.trim().replace(/,/g, '')) || 0;
                                    result['其他类'] = parseFloat(ths[3].textContent.trim().replace(/,/g, '')) || 0;
                                }
                                return result;
                            })()
                        """)

                    _eleme_set_date_and_query()

                    has_data = chrome_page.locator('.ant-table-tbody tr.ant-table-row').count() > 0
                    if not has_data:
                        logger.info(f"{TAG}  已结算账单中未找到 {date_label} 的数据，切换到待结算账单...")
                        chrome_page.locator('[data-node-key="TO_BE_SETTLED"] .cook-tabs-tab-btn').click()
                        chrome_page.wait_for_selector('input[placeholder="开始日期"]', timeout=30000)
                        time.sleep(3)
                        _eleme_set_date_and_query()

                    logger.info(f"{TAG}  → 抓取饿了么账单数据...")
                    eleme_data = _eleme_scrape()
                    logger.info(f"{TAG}  抓取数据: {json.dumps(eleme_data, ensure_ascii=False)}")

                    eleme_csv_fields = ["结算金额", "订单类", "其他类"]
                    eleme_csv_file = output_dir / f"饿了么账单_{mt_store_short}_{date_label}.csv"
                    with open(eleme_csv_file, "w", newline="", encoding="utf-8-sig") as f:
                        writer = csv.writer(f)
                        writer.writerow(["项目", "金额"])
                        for key in eleme_csv_fields:
                            writer.writerow([key, eleme_data.get(key, 0)])
                    logger.info(f"{TAG}  已保存CSV: {eleme_csv_file.name}")

                retry_until_success(_do_eleme, f"饿了么账单-{mt_store_short}")

        except Exception as e:
            logger.error(f"{TAG} 下载失败: {e}")
            if chrome_page:
                try:
                    screenshot = OUTPUT_DIR / f"meituan_error_{mt_store_short}.png"
                    chrome_page.screenshot(path=str(screenshot))
                except Exception:
                    pass
            raise
        finally:
            _cleanup_chrome(chrome_page, chrome_browser, chrome_process)


def run_meituan_waimai_all(args, target, target_str, date_label, output_dir):
    logger.info(f"[美团外卖] 启动 {len(MEITUAN_STORE_CONFIG)} 个门店并行下载...")

    def _run_store_with_retry(mt_config):
        retry_until_success(
            lambda cfg=mt_config: _run_meituan_waimai_store(cfg, args, target_str, date_label, output_dir),
            f"美团外卖-{mt_config['store_short']}"
        )

    with ThreadPoolExecutor(max_workers=len(MEITUAN_STORE_CONFIG)) as executor:
        futures = {
            executor.submit(_run_store_with_retry, mt_config): mt_config["store_short"]
            for mt_config in MEITUAN_STORE_CONFIG
        }
        for future in as_completed(futures):
            store = futures[future]
            future.result()
            logger.info(f"[美团外卖-{store}] 完成")

    logger.info("[美团外卖] 全部门店下载完成！")


# ─── Group C: 共享 Chrome（经营宝/招行/抖音/口碑）────────────────────────────


def _run_jyb_task(args, target, target_str, date_label, output_dir):
    TAG = "[经营宝]"
    from playwright.sync_api import sync_playwright

    jyb_browser = None
    jyb_page = None

    with sync_playwright() as pw:
        try:
            jyb_browser, jyb_page = _connect_cdp_with_retry(pw, 9226)

            logger.info(f"{TAG}  [导航] 前往美团经营宝每日收益页...")
            jyb_page.goto(MEITUAN_JYB_URL)
            # jyb_page.wait_for_load_state("networkidle", timeout=120_000)
            jyb_page.wait_for_selector('.mtd-table-body', timeout=30000)
            time.sleep(2)

            days_diff = (datetime.now().date() - target.date()).days
            logger.info(f"{TAG}  → 设置收益时间: {target.strftime('%Y/%m/%d')} (距今 {days_diff} 天)...")

            if days_diff in (0, 1, 2):
                btn_map = {0: "今日", 1: "昨日", 2: "前日"}
                jyb_page.locator(f'button[value="{btn_map[days_diff]}"]').click()
                time.sleep(0.5)
                logger.info(f"{TAG}  已点击快捷按钮: {btn_map[days_diff]}")
            else:
                jyb_page.locator('.mtd-date-picker input').click()
                jyb_page.wait_for_selector('.mtd-singleRangePicker-pop', state='visible', timeout=5000)

                current_info = jyb_page.evaluate("""
                    (function() {
                        var popup = document.querySelector('.mtd-singleRangePicker-pop');
                        if (!popup) return null;
                        var leftCal = popup.querySelector('.mtd-date-calendar');
                        if (!leftCal) return null;
                        var yearBtn = leftCal.querySelector('.mtd-date-calendar-year-btn');
                        var monthBtn = leftCal.querySelector('.mtd-date-calendar-month-btn');
                        if (!yearBtn || !monthBtn) return null;
                        return { year: parseInt(yearBtn.textContent), month: parseInt(monthBtn.textContent) };
                    })()
                """)

                if not current_info:
                    raise RuntimeError(f"{TAG} 日历面板未打开或无法读取年月")

                logger.info(f"{TAG}  当前日历: {current_info['year']}年{current_info['month']}月")
                months_back = (current_info['year'] - target.year) * 12 + (current_info['month'] - target.month)

                if months_back > 0:
                    logger.info(f"{TAG}  往前导航 {months_back} 个月...")
                    for _ in range(months_back):
                        jyb_page.locator(
                            '.mtd-singleRangePicker-pop .mtd-date-calendar:first-child '
                            '.mtd-date-calendar-month-switcher.left-switcher'
                        ).click()
                        jyb_page.wait_for_timeout(200)
                        time.sleep(0.5)

                target_day_str = str(target.day)

                start_result = jyb_page.evaluate(f"""
                    (function() {{
                        var popup = document.querySelector('.mtd-singleRangePicker-pop');
                        if (!popup) return 'popup not found';
                        var leftCal = popup.querySelector('.mtd-date-calendar');
                        var activePanel = leftCal.querySelector('.mtd-date-calendar-content.active');
                        if (!activePanel) return 'no active panel';
                        var wrappers = activePanel.querySelectorAll('.mtd-date-panel-data-wrapper');
                        for (var j = 0; j < wrappers.length; j++) {{
                            if (wrappers[j].classList.contains('not-current-month')) continue;
                            if (wrappers[j].classList.contains('disabled-date')) continue;
                            var btn = wrappers[j].querySelector('.mtd-date-panel-data');
                            if (btn && btn.textContent.trim() === '{target_day_str}') {{
                                btn.click();
                                return 'clicked start: day ' + btn.textContent.trim();
                            }}
                        }}
                        return 'start day {target_day_str} not found';
                    }})()
                """)
                logger.info(f"{TAG}  {start_result}")
                time.sleep(0.5)

                end_result = jyb_page.evaluate(f"""
                    (function() {{
                        var popup = document.querySelector('.mtd-singleRangePicker-pop');
                        if (!popup) return 'popup closed';
                        var leftCal = popup.querySelector('.mtd-date-calendar');
                        var activePanel = leftCal.querySelector('.mtd-date-calendar-content.active');
                        if (!activePanel) return 'no active panel';
                        var wrappers = activePanel.querySelectorAll('.mtd-date-panel-data-wrapper');
                        for (var j = 0; j < wrappers.length; j++) {{
                            if (wrappers[j].classList.contains('not-current-month')) continue;
                            var btn = wrappers[j].querySelector('.mtd-date-panel-data');
                            if (btn && btn.textContent.trim() === '{target_day_str}') {{
                                btn.click();
                                return 'clicked end: day ' + btn.textContent.trim();
                            }}
                        }}
                        return 'end day {target_day_str} not found';
                    }})()
                """)
                logger.info(f"{TAG}  {end_result}")
                time.sleep(0.5)

            logger.info(f"{TAG}  等待数据刷新...")
            jyb_page.wait_for_load_state("networkidle", timeout=30000)
            time.sleep(2)

            logger.info(f"{TAG}  → 抓取经营宝收益数据...")
            jyb_rows = jyb_page.evaluate("""
                (function() {
                    var rows = document.querySelectorAll('.mtd-table-body tbody tr');
                    var result = [];
                    for (var i = 0; i < rows.length; i++) {
                        var tds = rows[i].querySelectorAll('td');
                        if (tds.length < 9) continue;
                        var storeName = tds[0].textContent.trim();
                        var priceDiv = tds[2].querySelector('div > div');
                        var price = priceDiv
                            ? parseFloat(priceDiv.textContent.trim().replace(/,/g, '')) || 0
                            : parseFloat(tds[2].textContent.trim().replace(/,/g, '')) || 0;
                        var promotion = parseFloat(tds[3].textContent.trim().replace(/,/g, '')) || 0;
                        var service = parseFloat(tds[4].textContent.trim().replace(/,/g, '')) || 0;
                        var other = parseFloat(tds[7].textContent.trim().replace(/,/g, '')) || 0;
                        result.push({
                            storeName: storeName,
                            price: price,
                            promotion: promotion,
                            service: service,
                            other: other
                        });
                    }
                    return result;
                })()
            """)
            logger.info(f"{TAG}  抓取到 {len(jyb_rows)} 行数据")

            store_keywords = {"宝泰": "宝泰店", "龙江": "龙江店", "杏坛": "杏坛店"}
            for row in jyb_rows:
                store_short = None
                for keyword, short in store_keywords.items():
                    if keyword in row["storeName"]:
                        store_short = short
                        break
                if not store_short:
                    logger.warning(f"{TAG}  未匹配门店: {row['storeName']}")
                    continue

                logger.info(f"{TAG}  {store_short}: 售价={row['price']}, 促销费={row['promotion']}, "
                            f"服务费={row['service']}, 其他费用={row['other']}")

                jyb_csv = output_dir / f"美团经营宝_每日收益_{store_short}_{date_label}.csv"
                with open(jyb_csv, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)
                    writer.writerow(["项目", "金额"])
                    writer.writerow(["售价", row["price"]])
                    writer.writerow(["促销费", row["promotion"]])
                    writer.writerow(["服务费", row["service"]])
                    writer.writerow(["其他费用", row["other"]])
                logger.info(f"{TAG}  已保存CSV: {jyb_csv.name}")

        finally:
            if jyb_page:
                try:
                    jyb_page.close()
                except Exception:
                    pass
            if jyb_browser:
                try:
                    jyb_browser.close()
                except Exception:
                    pass

    logger.info(f"{TAG}  经营宝下载完成！")


def _run_zhaohang_task(args, target, target_str, date_label, output_dir):
    TAG = "[招行]"
    from playwright.sync_api import sync_playwright

    zh_browser = None
    zh_page = None

    with sync_playwright() as pw:
        try:
            zh_browser, zh_page = _connect_cdp_with_retry(pw, 9226)

            for zh_idx, zh_config in enumerate(ZHAOHANG_STORE_CONFIG):
                zh_store_short = zh_config["store_short"]
                desc = f"招行-{zh_store_short}"

                def _do_zhaohang(cfg=zh_config, idx=zh_idx, ss=zh_store_short):
                    logger.info(f"{TAG}  门店 {idx + 1}/{len(ZHAOHANG_STORE_CONFIG)}: {ss}")
                    zh_page.goto(ZHAOHANG_URL)
                    zh_page.wait_for_load_state("networkidle", timeout=120_000)
                    # zh_page.wait_for_selector('#storelist', timeout=30000)
                    zh_page.wait_for_selector('#J-search-form', timeout=30000)
                    scrape_zhaohang_daily_summary(zh_page, target_str, date_label, cfg, output_dir)

                retry_until_success(_do_zhaohang, desc)

        finally:
            if zh_page:
                try:
                    zh_page.close()
                except Exception:
                    pass
            if zh_browser:
                try:
                    zh_browser.close()
                except Exception:
                    pass

    logger.info(f"{TAG}  招行下载完成！")


def _run_douyin_task(args, target, target_str, date_label, output_dir):
    TAG = "[抖音]"
    from playwright.sync_api import sync_playwright

    dy_browser = None
    dy_page = None

    with sync_playwright() as pw:
        try:
            dy_browser, dy_page = _connect_cdp_with_retry(pw, 9226)

            logger.info(f"{TAG}  [导航] 前往抖音每日收益页面...")
            dy_page.goto(DOUYIN_DAILY_BENEFITS_URL)
            dy_page.wait_for_selector('.byted-date-picker', timeout=30000)
            # time.sleep(0.3)

            douyin_set_date(dy_page, target)

            for dy_idx, dy_config in enumerate(DOUYIN_STORE_CONFIG):
                dy_store_short = dy_config["store_short"]
                desc = f"抖音-{dy_store_short}"

                def _do_douyin(cfg=dy_config, idx=dy_idx, ss=dy_store_short):
                    logger.info(f"{TAG}  门店 {idx + 1}/{len(DOUYIN_STORE_CONFIG)}: {ss}")
                    douyin_select_store(dy_page, cfg)
                    dy_page.wait_for_selector('#submodule', timeout=15000)
                    # dy_page.wait_for_selector('[class*="container--"]', timeout=15000)
                    # dy_page.wait_for_load_state("networkidle", timeout=15000)
                    dy_data = scrape_douyin_daily_benefits(dy_page)
                    save_douyin_daily_benefits_csv(dy_data, ss, date_label, output_dir)

                retry_until_success(_do_douyin, desc)

                # wade
                time.sleep(5)

        finally:
            if dy_page:
                try:
                    dy_page.close()
                except Exception:
                    pass
            if dy_browser:
                try:
                    dy_browser.close()
                except Exception:
                    pass

    logger.info(f"{TAG}  抖音下载完成！")


def _run_koubei_task(args, target, target_str, date_label, output_dir):
    TAG = "[口碑]"
    from playwright.sync_api import sync_playwright

    kb_browser = None
    kb_page = None

    with sync_playwright() as pw:
        try:
            kb_browser, kb_page = _connect_cdp_with_retry(pw, 9226)

            logger.info(f"{TAG}  [导航] 前往高德口碑账单汇总页...")
            kb_page.goto(KOUBEI_BILL_URL)
            # kb_page.wait_for_load_state("networkidle", timeout=60000)
            kb_page.wait_for_selector('#ice-container', timeout=30000)
            dismiss_koubei_popup(kb_page)

            for kb_idx, kb_config in enumerate(KOUBEI_STORE_CONFIG):
                kb_store_short = kb_config["store_short"]
                desc = f"口碑-{kb_store_short}"

                def _do_koubei(cfg=kb_config, idx=kb_idx, ss=kb_store_short):
                    logger.info(f"{TAG}  门店 {idx + 1}/{len(KOUBEI_STORE_CONFIG)}: {ss}")
                    if idx > 0:
                        kb_page.goto(KOUBEI_BILL_URL)
                        # kb_page.wait_for_load_state("networkidle", timeout=60000)
                        kb_page.wait_for_selector('#ice-container', timeout=30000)
                        dismiss_koubei_popup(kb_page)
                    koubei_set_date(kb_page, target)
                    koubei_select_query_type(kb_page)
                    koubei_select_store(kb_page, cfg)
                    logger.info(f"{TAG}  [查询] 点击查询...")
                    kb_page.locator('button.aamf-btn-primary:has-text("查 询")').click()
                    kb_page.wait_for_selector('.name--pnzqmgDZ', timeout=30000)
                    time.sleep(3)
                    kb_data = scrape_koubei_bill(kb_page)
                    save_koubei_bill_csv(kb_data, ss, date_label, output_dir)

                retry_until_success(_do_koubei, desc)

        finally:
            if kb_page:
                try:
                    kb_page.close()
                except Exception:
                    pass
            if kb_browser:
                try:
                    kb_browser.close()
                except Exception:
                    pass

    logger.info(f"{TAG}  口碑下载完成！")


def run_shared_chrome_tasks(args, target, target_str, date_label, output_dir):
    logger.info("  启动共享 Chrome (port=9226)...")
    chrome_process = _launch_chrome(9226, r"C:\ChromeDebug_MTJYB", args.headless)
    try:
        retry_until_success(
            lambda: _run_jyb_task(args, target, target_str, date_label, output_dir),
            "经营宝"
        )
        retry_until_success(
            lambda: _run_zhaohang_task(args, target, target_str, date_label, output_dir),
            "招行"
        )
        retry_until_success(
            lambda: _run_douyin_task(args, target, target_str, date_label, output_dir),
            "抖音"
        )
        retry_until_success(
            lambda: _run_koubei_task(args, target, target_str, date_label, output_dir),
            "口碑"
        )
    finally:
        logger.info("  关闭共享 Chrome (port=9226)...")
        if chrome_process:
            _kill_process_tree(chrome_process)


# ─── Group D: 格式化写入 ────────────────────────────────────────────────────


def run_formatting(target, output_dir, date_label):
    logger.info(f"{'─' * 55}")
    logger.info(f"  格式化数据并写入月度统计表")
    logger.info(f"{'─' * 55}")

    monthly_file, created = create_or_open_monthly_file(
        STORES, target, TEMPLATE_FILE, OUTPUT_DIR
    )

    for i, store in enumerate(STORES):
        logger.info(f"  门店 {i + 1}/{len(STORES)}: {store['template_name']}")
        fill_daily_data(monthly_file, target, store, output_dir, store_index=i)

    logger.info(f"{'=' * 55}")
    logger.info(f"  麦安研营业统计格式化全部完成！")
    logger.info(f"{'=' * 55}\n")


# ─── 主入口 ──────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(description="麦安研营业统计自动化脚本")
    parser.add_argument("--days", type=int, default=0, help="日期偏移量：0=今天，-1=昨天（默认0）")
    parser.add_argument("--date", type=str, help="指定目标日期，格式 YYYY.MM.DD")
    parser.add_argument("--headless", action="store_true", help="无头模式（不显示浏览器窗口）")
    args = parser.parse_args()

    target = get_target_date(days=args.days, date_str=args.date)
    target_str = target.strftime("%Y.%m.%d")
    date_label = target.strftime("%Y-%m-%d")

    logger.info(f"{'=' * 55}")
    logger.info(f"  麦安研营业统计（并行模式）")
    logger.info(f"  目标日期：{target_str}")
    logger.info(f"  输出根目录：{OUTPUT_DIR}")
    logger.info(f"{'=' * 55}")

    try:
        from playwright.sync_api import sync_playwright  # noqa: F401
    except ImportError:
        logger.error("未安装 playwright，请先运行: pip install playwright && playwright install chromium")
        sys.exit(1)

    logger.info("  检查并清理残留浏览器进程...")
    cleanup_stale_browsers()

    output_dir = OUTPUT_DIR / "原始下载" / date_label
    output_dir.mkdir(parents=True, exist_ok=True)

    # ── 三大组并行执行，失败的组重试直到全部成功 ──────────────────
    group_tasks = {
        "银豹系统": lambda: run_pospal_tasks(args, target, target_str, date_label, output_dir),
        "美团外卖": lambda: run_meituan_waimai_all(args, target, target_str, date_label, output_dir),
        "共享Chrome任务": lambda: run_shared_chrome_tasks(args, target, target_str, date_label, output_dir),
    }
    pending_groups = dict(group_tasks)
    group_round = 0

    while pending_groups:
        group_round += 1
        if group_round > 1:
            logger.info(f"{'=' * 55}")
            logger.info(f"  第 {group_round} 轮重试，剩余 {len(pending_groups)} 个任务组: {list(pending_groups.keys())}")
            logger.info(f"{'=' * 55}")
            time.sleep(5)

        failed_groups = {}

        with ThreadPoolExecutor(max_workers=len(pending_groups)) as executor:
            futures = {
                executor.submit(fn): name
                for name, fn in pending_groups.items()
            }
            for future in as_completed(futures):
                name = futures[future]
                try:
                    future.result()
                    logger.info(f"  >>> {name} 全部完成")
                except Exception as e:
                    logger.error(f"  >>> {name} 失败: {e}")
                    failed_groups[name] = group_tasks[name]

        pending_groups = failed_groups

    # ── 格式化写入（等待所有下载完成后执行）──────────────────
    retry_until_success(
        lambda: run_formatting(target, output_dir, date_label),
        "格式化写入"
    )

    logger.info(f"  全部任务成功完成！")


if __name__ == "__main__":
    main()
