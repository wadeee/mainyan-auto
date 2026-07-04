"""
工厂配送麦安研门店每日销售报表 - 自动化脚本
=====================================
依赖：pip install playwright openpyxl && playwright install chromium

自动登录 Pospal 后台，导出仓库配送商品门店对比表并生成每日格式化报表。

用法：
    python factory_delivery_mainyan_daily.py                    # 导出今天的数据
    python factory_delivery_mainyan_daily.py --days -1          # 导出昨天的数据
    python factory_delivery_mainyan_daily.py --date 2026.06.20  # 指定日期
    python factory_delivery_mainyan_daily.py --headless          # 无头模式（不显示浏览器）
"""

import argparse
import copy
import logging
import logging.handlers
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ─── 日志配置 ───────────────────────────────────────────────────────────────────
LOG_DIR = Path(__file__).resolve().parent / "log"
LOG_DIR.mkdir(exist_ok=True)

_file_handler = logging.handlers.TimedRotatingFileHandler(
    LOG_DIR / "factory_delivery_mainyan_daily.log",
    when="midnight",
    backupCount=30,
    encoding="utf-8",
)
_file_handler.suffix = "%Y-%m-%d"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[_file_handler, logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# ─── 配置区（按需修改）─────────────────────────────────────────────────────────
ACCOUNT = "huomimayzb"
WORKER_ID = "M006"
PASSWORD = "tusijia88"

LOGIN_URL = "https://beta69.pospal.cn/"
REPORT_URL = "https://css69.pospal.cn/ChainStoreSupplyReport/WarehouseDeliveryProductStoreComparison"
DELIVERY_PRODUCT_COMPARISON_URL = "https://css69.pospal.cn/EnterpriseReport/DeliveryProductComparison"

OUTPUT_DIR = Path(__file__).resolve().parent / "工厂配送麦安研门店每日销售报表"
TEMPLATE_FILE = Path(__file__).resolve().parent / "工厂配送麦安研门店每日销售报表_格式化模板.xlsx"

STORE_NAME_MAP = {
    "麦安研（顺德杏坛店）": "杏坛",
    "麦安研（东站宝泰店）": "宝泰",
    "麦安研（佛山创产店）": "创产",
    "麦安研（顺德龙江店）": "龙江",
    "麦安研（佛山万民金海城店）": "金海城",
    "麦安研门店裱花间": "裱花间",
    "中央工厂冷加工间": "冷加工间",
}

BMX_STORE_ORDER = ["广钢", "滨江", "南州路"]
BMX_CUSTOMER_NAME_MAP = {
    "焙满香广钢店": "广钢",
    "焙满香滨江店": "滨江",
    "焙满香南州路店": "南州路",
}

TARGET_CATEGORIES = [
    "配送费",
    "包材耗材",
    "工衣模具",
    "慕斯+饼干+饮品+其他",
    "原料铺料",
    "冷冻面团",
    "蛋糕及面包成品及饼干类",
]

CATEGORY_SORT_ORDER = [
    "冷冻面团",
    "成品面包类", "饼干类",
    "蛋糕类",
    "热销类", "慕斯类/外", "冷冻肉类", "冷冻馅料类", "冷藏馅料类", "油脂类", "粉类", "糖类", "常温馅料类",
    "干果类", "饼干类/外", "饮品类/外", "其他/外", "专版包材类", "公版包材类", "工衣工帽围裙", "模具",
    "保洁用品", "配送费",
]

_CATEGORY_PRIORITY = {cat: idx for idx, cat in enumerate(CATEGORY_SORT_ORDER)}

HEADER_ROW = 1
TOTALS_ROW = 2
DATA_START_ROW = 3
SAMPLE_ROW = 3


# ─── 工具函数 ──────────────────────────────────────────────────────────────────


def _to_num(val):
    if val is None:
        return None
    try:
        f = float(val)
        if f == 0:
            return None
        if f == int(f):
            return int(f)
        return f
    except (ValueError, TypeError):
        return val


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def clear_cell(cell):
    from openpyxl.styles import Font, PatternFill, Border, Alignment
    cell.value = None
    cell.font = Font()
    cell.fill = PatternFill()
    cell.border = Border()
    cell.alignment = Alignment()
    cell.number_format = "General"


def _style_ref_idx(col_idx, sample_len):
    """列号 → 样式参照索引。1~6 直接映射；7+ 按 G/H 两列一组循环。"""
    if col_idx <= 6:
        return min(col_idx - 1, sample_len - 1)
    offset = (col_idx - 7) % 2
    ref = 6 + offset
    return min(ref, sample_len - 1)


def _get_short_name(full_name):
    short = STORE_NAME_MAP.get(full_name)
    if short:
        return short
    m = re.search(r'[（(](.+?)[）)]', full_name)
    if m:
        return m.group(1).rstrip('店')
    return full_name


# ─── 数据读取 ──────────────────────────────────────────────────────────────────


def read_downloaded_data(data_file):
    """读取仓库配送商品门店对比表，返回 (store_order, products)。

    store_order: 活跃门店简称列表
    products: 商品明细列表，每项包含 stores 字典（{简称: {数量, 金额}}）
    """
    wb = load_workbook(data_file, data_only=True)
    ws = wb.active

    # 扫描门店列组：每组 3 列（数量、成本金额、出库金额），跳过合计
    store_groups = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h is None:
            continue
        h_str = str(h).strip()
        if h_str.endswith("_数量"):
            name_part = h_str.rsplit("_", 1)[0]
            if name_part != "合计":
                store_groups.append((name_part, c, c + 2))

    # 过滤：仅保留三列中存在非零数据的门店
    active_stores = []
    for full_name, qty_col, amt_col in store_groups:
        has_data = False
        for r in range(2, ws.max_row + 1):
            for col in [qty_col, qty_col + 1, amt_col]:
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    try:
                        if float(v) != 0:
                            has_data = True
                            break
                    except (ValueError, TypeError):
                        pass
            if has_data:
                break
        if has_data:
            short_name = _get_short_name(full_name)
            active_stores.append((short_name, qty_col, amt_col))

    logger.info(f"  活跃门店 ({len(active_stores)}): {', '.join(s[0] for s in active_stores)}")

    # 读取商品数据（来源列：D=商品大类, E=商品分类, F=商品名称, I=规格, J=单位）
    products = []
    for r in range(2, ws.max_row + 1):
        seq = ws.cell(row=r, column=1).value
        if seq is None:
            break
        product = {
            "商品大类": str(ws.cell(row=r, column=4).value or "").strip(),
            "商品分类": str(ws.cell(row=r, column=5).value or "").strip(),
            "商品名称": str(ws.cell(row=r, column=6).value or "").strip(),
            "规格": ws.cell(row=r, column=9).value,
            "单位": ws.cell(row=r, column=10).value,
            "stores": {},
        }
        for short_name, qty_col, amt_col in active_stores:
            product["stores"][short_name] = {
                "数量": _to_num(ws.cell(row=r, column=qty_col).value),
                "金额": _to_num(ws.cell(row=r, column=amt_col).value),
            }
        products.append(product)

    wb.close()
    store_order = [s[0] for s in active_stores]
    logger.info(f"  共 {len(products)} 条商品数据")
    return store_order, products


def read_bmx_store_data(data_file):
    """读取仓库配送商品大客户对比表，提取焙满香三家门店数据，按商品名称透视。

    返回 {商品名称: {"info": {商品大类, 商品分类, 规格, 单位}, "stores": {store_short: {数量, 金额}}}}
    """
    wb = load_workbook(data_file, data_only=True)
    ws = wb.active

    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val:
            headers[str(val).strip()] = c

    pivot = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=headers["大客户名称"]).value
        if name is None:
            break
        name = str(name).strip()
        if name == "合计" or name == "-":
            break

        store_short = BMX_CUSTOMER_NAME_MAP.get(name)
        if not store_short:
            continue

        product_name = str(ws.cell(row=r, column=headers["商品名称"]).value or "").strip()
        if not product_name:
            continue

        if product_name not in pivot:
            pivot[product_name] = {
                "info": {
                    "商品大类": str(ws.cell(row=r, column=headers["商品大类"]).value or "").strip(),
                    "商品分类": str(ws.cell(row=r, column=headers["商品分类"]).value or "").strip(),
                    "规格": ws.cell(row=r, column=headers["规格"]).value,
                    "单位": ws.cell(row=r, column=headers["单位"]).value,
                },
                "stores": {},
            }
        if store_short not in pivot[product_name]["stores"]:
            pivot[product_name]["stores"][store_short] = {"数量": 0.0, "金额": 0.0}

        qty = ws.cell(row=r, column=headers["实际销售量"]).value
        amt = ws.cell(row=r, column=headers["实际销售金额"]).value
        if qty:
            try:
                pivot[product_name]["stores"][store_short]["数量"] += float(qty)
            except (ValueError, TypeError):
                pass
        if amt:
            try:
                pivot[product_name]["stores"][store_short]["金额"] += float(amt)
            except (ValueError, TypeError):
                pass

    wb.close()

    for pname in pivot:
        for store in pivot[pname]["stores"]:
            pivot[pname]["stores"][store]["数量"] = _to_num(pivot[pname]["stores"][store]["数量"])
            pivot[pname]["stores"][store]["金额"] = _to_num(pivot[pname]["stores"][store]["金额"])

    logger.info(f"  焙满香门店数据: {len(pivot)} 个商品")
    return pivot


def merge_products_with_bmx(products, bmx_data):
    """将焙满香门店数据合并到产品列表中。

    - 已存在的商品：将焙满香门店数据追加到 stores 字典
    - 仅焙满香有的商品：新建产品条目
    返回 (合并后的产品列表, 活跃焙满香门店列表)
    """
    # 计算活跃的焙满香门店
    bmx_active_order = []
    for short_name in BMX_STORE_ORDER:
        has_data = any(
            (stores["stores"].get(short_name, {}).get("数量") is not None or
             stores["stores"].get(short_name, {}).get("金额") is not None)
            for stores in bmx_data.values()
        )
        if has_data:
            bmx_active_order.append(short_name)

    if not bmx_active_order:
        return products, []

    logger.info(f"  焙满香活跃门店 ({len(bmx_active_order)}): {', '.join(bmx_active_order)}")

    # 为已有产品追加焙满香门店数据
    product_name_set = set()
    for p in products:
        pname = p["商品名称"]
        product_name_set.add(pname)
        if pname in bmx_data:
            for store_short in bmx_active_order:
                sd = bmx_data[pname]["stores"].get(store_short)
                if sd:
                    p["stores"][store_short] = sd

    # 添加仅焙满香有的商品
    bmx_only_count = 0
    for pname, data in bmx_data.items():
        if pname in product_name_set:
            continue
        stores = {}
        for store_short in bmx_active_order:
            sd = data["stores"].get(store_short)
            if sd:
                stores[store_short] = sd
        if stores:
            products.append({
                "商品大类": data["info"]["商品大类"],
                "商品分类": data["info"]["商品分类"],
                "商品名称": pname,
                "规格": data["info"]["规格"],
                "单位": data["info"]["单位"],
                "stores": stores,
            })
            bmx_only_count += 1

    if bmx_only_count:
        logger.info(f"  焙满香独有商品: {bmx_only_count} 条（已合并）")

    return products, bmx_active_order


# ─── 排序 ──────────────────────────────────────────────────────────────────────


def _sort_key(product):
    """按商品分类在 EXPORT_CATEGORY_MAP 中的顺序排序，同分类内保持原始顺序。"""
    cat = product.get("商品分类", "")
    return _CATEGORY_PRIORITY.get(cat, len(CATEGORY_SORT_ORDER))


# ─── 填充模板 ─────────────────────────────────────────────────────────────────


def fill_sheet(wb, sorted_products, store_order, target_date):
    ws = wb.worksheets[0]
    original_max_col = ws.max_column

    store_count = len(store_order)
    max_col = 6 + store_count * 2

    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    sample_cells = list(ws.iter_rows(min_row=SAMPLE_ROW, max_row=SAMPLE_ROW))[0]
    sample_height = ws.row_dimensions[SAMPLE_ROW].height

    sample_count = ws.max_row - SAMPLE_ROW + 1
    ws.delete_rows(SAMPLE_ROW, sample_count)
    data_count = len(sorted_products)
    ws.insert_rows(DATA_START_ROW, data_count)

    # 建立门店列映射
    store_columns = {}
    for idx, short_name in enumerate(store_order):
        qty_col = 7 + idx * 2
        amt_col = 8 + idx * 2
        store_columns[short_name] = (qty_col, amt_col)

    # 写入门店表头
    for short_name, (qty_col, amt_col) in store_columns.items():
        ws.cell(row=HEADER_ROW, column=qty_col).value = f"{short_name}数量"
        ws.cell(row=HEADER_ROW, column=amt_col).value = f"{short_name}金额"

    # 填充数据行
    for i, product in enumerate(sorted_products):
        r = DATA_START_ROW + i
        ws.cell(row=r, column=1).value = f"=ROW()-{TOTALS_ROW}"
        ws.cell(row=r, column=2).value = product["商品大类"]
        ws.cell(row=r, column=3).value = product["商品分类"]
        ws.cell(row=r, column=4).value = product["商品名称"]
        ws.cell(row=r, column=5).value = product["规格"]
        ws.cell(row=r, column=6).value = product["单位"]

        for short_name, (qty_col, amt_col) in store_columns.items():
            store_data = product["stores"].get(short_name, {})
            qty_val = store_data.get("数量")
            amt_val = store_data.get("金额")
            if qty_val is not None:
                ws.cell(row=r, column=qty_col).value = qty_val
            if amt_val is not None:
                ws.cell(row=r, column=amt_col).value = amt_val

        for col_idx in range(1, max_col + 1):
            src_idx = _style_ref_idx(col_idx, len(sample_cells))
            copy_cell_style(sample_cells[src_idx], ws.cell(row=r, column=col_idx))
        if sample_height:
            ws.row_dimensions[r].height = sample_height

    # 更新合计行公式
    last_data_row = DATA_START_ROW + data_count - 1
    ws.cell(row=TOTALS_ROW, column=1).value = "合计"
    for col in range(7, max_col + 1):
        letter = get_column_letter(col)
        ws.cell(row=TOTALS_ROW, column=col).value = (
            f"=SUM({letter}{DATA_START_ROW}:{letter}{last_data_row})"
        )

    # 表头和合计行门店列样式
    for hr in [HEADER_ROW, TOTALS_ROW]:
        ref_qty = ws.cell(row=hr, column=7)
        ref_amt = ws.cell(row=hr, column=8)
        for c in range(7, max_col + 1):
            ref = ref_qty if (c - 7) % 2 == 0 else ref_amt
            copy_cell_style(ref, ws.cell(row=hr, column=c))

    # 列宽
    g_width = ws.column_dimensions["G"].width
    h_width = ws.column_dimensions["H"].width
    if g_width and h_width:
        for idx in range(store_count):
            ws.column_dimensions[get_column_letter(7 + idx * 2)].width = g_width
            ws.column_dimensions[get_column_letter(8 + idx * 2)].width = h_width

    # 清除多余列
    if max_col < original_max_col:
        for r in range(1, ws.max_row + 1):
            for c in range(max_col + 1, original_max_col + 1):
                clear_cell(ws.cell(row=r, column=c))

    # 合并单元格
    ws.merge_cells(f"A{TOTALS_ROW}:F{TOTALS_ROW}")

    # 更新 sheet 名称为日期
    target_dt = datetime.strptime(target_date, "%Y.%m.%d")
    ws.title = f"{target_dt.month}.{target_dt.day}"


# ─── 格式化合并主函数 ──────────────────────────────────────────────────────────


def merge_into_template(data_file, template_file, output_file, target_date, bmx_data=None):
    logger.info("读取下载数据...")
    store_order, products = read_downloaded_data(data_file)

    # 合并焙满香门店数据到统一的产品列表
    bmx_active_order = []
    if bmx_data:
        products, bmx_active_order = merge_products_with_bmx(products, bmx_data)

    if not products:
        logger.warning("没有数据，跳过格式化")
        return

    # 所有门店统一过滤和排序
    all_store_order = store_order + bmx_active_order
    active_products = [
        p for p in products
        if any((s.get("数量") or 0) != 0 or (s.get("金额") or 0) != 0 for s in p["stores"].values())
    ]
    sorted_products = sorted(active_products, key=_sort_key)

    logger.info(f"加载模板: {template_file}")
    wb = load_workbook(template_file)

    logger.info(f"填充数据, {len(sorted_products)} 条商品, {len(all_store_order)} 个门店...")
    fill_sheet(wb, sorted_products, all_store_order, target_date)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    logger.info(f"已生成格式化文件: {output_file}")


# ─── 浏览器自动化函数 ────────────────────────────────────────────────────────


def set_date(page, placeholder, value):
    page.evaluate(f"""
        (function() {{
            var inp = document.querySelector('input.timeInput.hasDatepicker[placeholder="{placeholder}"]');
            if (!inp) return 'not found';
            inp.value = "{value}";
            inp.dispatchEvent(new Event('input',  {{bubbles: true}}));
            inp.dispatchEvent(new Event('change', {{bubbles: true}}));
            inp.dispatchEvent(new Event('blur',   {{bubbles: true}}));
            return inp.value;
        }})()
    """)


def click_by_text(page, text, desc=""):
    result = page.evaluate(f"""
        (function() {{
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {{
                if (els[i].textContent.trim() === "{text}" && els[i].children.length === 0) {{
                    els[i].click();
                    return 'clicked tag=' + els[i].tagName + ' class=' + els[i].className;
                }}
            }}
            return 'not found';
        }})()
    """)
    tag = f"[{desc}] " if desc else ""
    logger.info(f"  {tag}→ {result}")
    return "clicked" in result


def login(page):
    logger.info("[1/4] 打开登录页面...")
    page.goto(LOGIN_URL)
    page.wait_for_load_state("networkidle")

    logger.info("[2/4] 切换到工号登录模式...")
    page.evaluate("""
        (function() {
            var els = document.querySelectorAll('div,span,a,button');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '工号登录' && els[i].children.length === 0) {
                    els[i].click();
                    return 'clicked';
                }
            }
            return 'not found';
        })()
    """)
    time.sleep(1.5)

    placeholder = page.evaluate("""
        (function() {
            var inputs = document.querySelectorAll('input');
            var r = [];
            for (var i = 0; i < inputs.length; i++) {
                if (inputs[i].placeholder) r.push(inputs[i].placeholder);
            }
            return r.join(' | ');
        })()
    """)
    if "员工工号" not in placeholder:
        raise RuntimeError(f"工号登录切换失败，当前输入框：{placeholder}")
    logger.info(f"  表单已切换 → {placeholder}")

    logger.info("[3/4] 填入账号/工号/密码...")
    page.evaluate(f"""
        (function() {{
            var inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].placeholder === '请输入账号')     inputs[i].value = "{ACCOUNT}";
                if (inputs[i].placeholder === '请输入员工工号') inputs[i].value = "{WORKER_ID}";
                if (inputs[i].placeholder === '请输入工号密码') inputs[i].value = "{PASSWORD}";
            }}
            inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].value) {{
                    inputs[i].dispatchEvent(new Event('input', {{bubbles: true}}));
                }}
            }}
        }})()
    """)

    time.sleep(1.5)

    logger.info("[4/4] 点击登录按钮...")
    click_by_text(page, "登 录", "登录")
    page.wait_for_load_state("networkidle", timeout=120_000)

    time.sleep(2)


# ─── 主函数 ────────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(description="Pospal 工厂配送麦安研门店每日销售报表自动化脚本")
    parser.add_argument("--days", type=int, default=0, help="日期偏移：0=今天，-1=昨天（默认0）")
    parser.add_argument("--date", type=str, help="指定日期，格式 YYYY.MM.DD")
    parser.add_argument("--headless", action="store_true", help="无头模式（不显示浏览器窗口）")
    args = parser.parse_args()

    if args.date:
        target_date = args.date
    else:
        target_dt = datetime.now() + timedelta(days=args.days)
        target_date = target_dt.strftime("%Y.%m.%d")

    target_dt = datetime.strptime(target_date, "%Y.%m.%d")
    date_str = target_dt.strftime("%Y-%m-%d")

    logger.info(f"{'=' * 55}")
    logger.info(f"  Pospal 工厂配送麦安研门店每日销售报表")
    logger.info(f"  目标日期：{target_date}")
    logger.info(f"  输出根目录：{OUTPUT_DIR}")
    logger.info(f"{'=' * 55}")

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        logger.error("未安装 playwright，请先运行: pip install playwright && playwright install chromium")
        sys.exit(1)

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=args.headless,
            args=["--start-maximized"],
        )
        context = browser.new_context(
            viewport={"width": 1600, "height": 900},
            accept_downloads=True,
        )
        page = context.new_page()
        page.set_default_timeout(120000)
        page.set_default_navigation_timeout(120000)

        try:
            login(page)

            # ── 步骤 1：下载仓库配送商品门店对比表 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：下载仓库配送商品门店对比表")
            logger.info(f"{'─' * 55}")

            logger.info("  [导航] 前往仓库配送商品门店对比表...")
            page.goto(REPORT_URL)
            page.wait_for_load_state("networkidle", timeout=120_000)
            logger.info(f"  已到达 → {page.url}")

            logger.info(f"  → 设置时间: {target_date}...")
            set_date(page, "开始日期", f"{target_date} 00:00")
            set_date(page, "结束日期", f"{target_date} 23:59")

            logger.info("  → 点击更多搜索...")
            more_search_result = page.evaluate("""
                (function() {
                    var btn = document.getElementById('advancedBtn');
                    if (btn) { btn.click(); return 'clicked advancedBtn'; }
                    var els = document.querySelectorAll('*');
                    for (var i = 0; i < els.length; i++) {
                        var t = els[i].textContent.trim();
                        if ((t === '更多搜索' || t === '更多搜索▼') && els[i].children.length === 0) {
                            els[i].click();
                            return 'clicked tag=' + els[i].tagName + ' class=' + els[i].className;
                        }
                    }
                    return 'not found';
                })()
            """)
            logger.info(f"  [更多搜索] → {more_search_result}")
            time.sleep(0.5)

            logger.info("  → 打开分类选择弹框...")
            select_cat_result = page.evaluate("""
                (function() {
                    var btn = document.getElementById('selectCategory');
                    if (btn) { btn.click(); return 'clicked selectCategory'; }
                    var els = document.querySelectorAll('*');
                    for (var i = 0; i < els.length; i++) {
                        if (els[i].textContent.trim() === '选择分类' && els[i].children.length === 0) {
                            els[i].click();
                            return 'clicked tag=' + els[i].tagName;
                        }
                    }
                    return 'not found';
                })()
            """)
            logger.info(f"  [选择分类] → {select_cat_result}")
            time.sleep(1.0)

            logger.info(f"  → 勾选 {len(TARGET_CATEGORIES)} 个分类...")
            categories_js = str(TARGET_CATEGORIES).replace("'", '"')
            page.evaluate(f"""
                (function() {{
                    var targets = {categories_js};
                    var divs = document.querySelectorAll('.checkBoxDiv');
                    divs.forEach(function(d) {{
                        var span = d.querySelector('span');
                        if (span && targets.indexOf(span.textContent.trim()) >= 0) {{
                            if (!d.classList.contains('on')) {{
                                d.click();
                            }}
                        }}
                    }});
                }})()
            """)

            check_result = page.evaluate(f"""
                (function() {{
                    var targets = {categories_js};
                    var divs = document.querySelectorAll('.checkBoxDiv');
                    var r = [];
                    divs.forEach(function(d) {{
                        var s = d.querySelector('span');
                        if (s && targets.indexOf(s.textContent.trim()) >= 0) {{
                            r.push(s.textContent.trim() + ':' + d.classList.contains('on'));
                        }}
                    }});
                    return r.join(' | ');
                }})()
            """)
            logger.info(f"    勾选验证: {check_result}")

            if "false" in check_result:
                logger.warning("部分分类未勾选，尝试重试...")
                page.evaluate(f"""
                    (function() {{
                        var targets = {categories_js};
                        var divs = document.querySelectorAll('.checkBoxDiv');
                        divs.forEach(function(d) {{
                            var span = d.querySelector('span');
                            if (span && targets.indexOf(span.textContent.trim()) >= 0 && !d.classList.contains('on')) {{
                                d.click();
                            }}
                        }});
                    }})()
                """)

            logger.info("  → 点击「确定」关闭弹框...")
            click_by_text(page, "确定", "确定弹框")
            time.sleep(0.5)

            logger.info("  [查询] 执行查询...")
            click_by_text(page, "查询", "查询")
            page.wait_for_load_state("networkidle", timeout=150_000)
            time.sleep(3)

            logger.info("  [导出] 导出文件...")
            output_dir = OUTPUT_DIR / date_str / "原始下载"
            output_dir.mkdir(parents=True, exist_ok=True)

            with page.expect_download(timeout=180_000) as dl_info:
                result = page.evaluate("""
                    (function() {
                        var els = document.querySelectorAll('*');
                        for (var i = 0; i < els.length; i++) {
                            if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                                els[i].click();
                                return 'ok';
                            }
                        }
                        return 'not found';
                    })()
                """)
                logger.info(f"  点击导出: {result}")
                if result == "not found":
                    raise RuntimeError("未找到「导出」按钮")

            download = dl_info.value
            logger.info(f"  下载文件名: {download.suggested_filename}")

            dest = output_dir / f"仓库配送商品门店对比表_{date_str}.xlsx"
            download.save_as(dest)
            logger.info(f"  已保存到: {dest}")

            # ── 步骤 2：下载仓库配送商品大客户对比表 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：下载仓库配送商品大客户对比表")
            logger.info(f"{'─' * 55}")

            logger.info("  [导航] 前往仓库配送商品大客户对比表...")
            page.goto(DELIVERY_PRODUCT_COMPARISON_URL)
            page.wait_for_load_state("networkidle", timeout=120_000)
            logger.info(f"  已到达 → {page.url}")

            logger.info(f"  → 设置时间: {target_date}...")
            set_date(page, "开始日期", f"{target_date} 00:00")
            set_date(page, "结束日期", f"{target_date} 23:59")

            logger.info("  [查询] 执行查询...")
            click_by_text(page, "查询", "查询")
            page.wait_for_load_state("networkidle", timeout=150_000)
            time.sleep(3)

            logger.info("  [导出] 导出文件...")
            with page.expect_download(timeout=180_000) as dl_info2:
                result = page.evaluate("""
                    (function() {
                        var els = document.querySelectorAll('*');
                        for (var i = 0; i < els.length; i++) {
                            if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                                els[i].click();
                                return 'ok';
                            }
                        }
                        return 'not found';
                    })()
                """)
                logger.info(f"  点击导出: {result}")
                if result == "not found":
                    raise RuntimeError("未找到「导出」按钮")

            download2 = dl_info2.value
            logger.info(f"  下载文件名: {download2.suggested_filename}")

            dest2 = output_dir / f"仓库配送商品大客户对比表_{date_str}.xlsx"
            download2.save_as(dest2)
            logger.info(f"  已保存到: {dest2}")

            # ── 步骤 3：格式化数据并生成报表 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：格式化数据并生成报表")
            logger.info(f"{'─' * 55}")

            logger.info("  读取焙满香门店数据...")
            bmx_data = read_bmx_store_data(dest2)

            formatted_output = OUTPUT_DIR / date_str / f"工厂配送麦安研门店每日销售报表_{date_str}.xlsx"
            merge_into_template(dest, TEMPLATE_FILE, formatted_output, target_date, bmx_data)

            logger.info(f"{'=' * 55}")
            logger.info(f"  工厂配送麦安研门店每日销售报表全部完成！")
            logger.info(f"  仓库配送商品门店对比表 → {dest}")
            logger.info(f"  仓库配送商品大客户对比表 → {dest2}")
            logger.info(f"  格式化输出 → {formatted_output}")
            logger.info(f"{'=' * 55}\n")

        except Exception as e:
            logger.error(f"任务失败: {e}")
            try:
                screenshot = OUTPUT_DIR / "error_screenshot.png"
                page.screenshot(path=str(screenshot))
                logger.info(f"  错误截图已保存: {screenshot}")
            except Exception:
                pass
            raise
        finally:
            context.close()
            browser.close()


if __name__ == "__main__":
    main()
