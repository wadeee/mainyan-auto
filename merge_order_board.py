"""
订货商品汇总看板 - 数据填入格式模板脚本
用途：将每日导出的订货数据 Excel 填入格式模板，删除示例行，更新汇总公式
依赖：pip install openpyxl
"""

import copy
from pathlib import Path
from datetime import date

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ============================================================
# 配置区 - 按需修改
# ============================================================
BASE_DIR = Path("C:/Users/Wadec/Desktop/订货商品汇总看板/2026-05-27")

# 数据源文件（每日导出的原始看板）
# today_str = date.today().strftime("%Y-%m-%d")
DATA_FILE = BASE_DIR / "原始下载" / f"订货商品汇总看板_2026-05-27.xlsx"

# 明细数据源（用于计算 G2-G5 等汇总区的分类金额）
DETAIL_FILE = BASE_DIR / "原始下载" / f"订货商品明细看板_2026-05-27.xlsx"

# 汇总区行号 → 商品分类映射（行号对应模板中的 row 2-5）
ROW_CATEGORY_MAP = {
    2: ["冷冻面团"],
    3: ["蛋糕类", "成品面包类", "饼干类"],
    4: ["专版包材类", "公版包材类", "工衣工帽围裙", "模具", "保洁用品", "饼干类/外", "慕斯类/外", "饮品类/外",
        "其他/外", "热销类", "冷冻馅料类", "肉类", "油脂类", "冷藏馅料类", "粉类", "糖类", "常温馅料类", "干果类"],
    5: ["配送费"],
}

# 格式模板文件（固定，不会变）
TEMPLATE_FILE = Path("订货商品汇总看板_格式化模板.xlsx")

# 输出文件
OUTPUT_FILE = BASE_DIR / f"订货商品汇总看板_格式化_2026-05-27.xlsx"

# 模板中示例数据所在行（第几行开始、共几行）
SAMPLE_ROW_START = 8  # 示例数据起始行
SAMPLE_ROW_COUNT = 2  # 示例数据行数

# 数据写入起始行（删除示例行后，从此行开始写入真实数据）
DATA_START_ROW = 8

# 表头行号（用于计算序号偏移量：序号 = ROW() - HEADER_ROW）
HEADER_ROW = 7

# 固定列映射（数据源 0-based 索引 → 模板列号 1-based）
# 门店列（G列起）由数据源中"合计订货量"之后的列动态决定
FIXED_COLUMN_MAP = [
    (1, 2),  # 商品分类 -> B
    (3, 3),  # 商品名称 -> C
    (4, 4),  # 规格 -> D
    (5, 5),  # 单位 -> E
]

TOTAL_ORDER_HEADER = "合计订货量"


# ============================================================


def copy_cell_style(src_cell, dst_cell):
    """将源单元格的样式（字体/填充/边框/对齐/数字格式）复制到目标单元格"""
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def compute_category_sums(detail_file: Path, store_names: list[str]) -> dict:
    """从明细看板读取数据，按(行号, 门店名)分组求和订货金额。

    返回 {(row_num, store_name): float}，row_num 为模板行号 2-5。
    """
    wb = load_workbook(detail_file, data_only=True)
    ws = wb.active

    cat_to_row = {}
    for row_num, categories in ROW_CATEGORY_MAP.items():
        for cat in categories:
            cat_to_row[cat] = row_num

    sums: dict[tuple[int, str], float] = {}
    for row in ws.iter_rows(min_row=2):
        category = row[1].value  # B列：商品分类
        org = row[5].value  # F列：订货组织
        amount = row[13].value  # N列：订货金额

        if category is None or org is None or amount is None:
            continue

        category_str = str(category).strip()
        org_str = str(org).strip()

        row_num = cat_to_row.get(category_str)
        if row_num is None:
            continue

        if org_str not in store_names:
            continue

        key = (row_num, org_str)
        sums[key] = sums.get(key, 0) + float(amount)

    wb.close()
    return sums


def read_data_file(data_file: Path):
    """从数据源文件读取表头信息和数据行，动态识别门店列"""
    wb = load_workbook(data_file)
    ws = wb.active

    headers = [cell.value for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]

    total_col_idx = None
    for i, h in enumerate(headers):
        if h and TOTAL_ORDER_HEADER in str(h):
            total_col_idx = i
            break
    if total_col_idx is None:
        raise ValueError(f"数据源中找不到'{TOTAL_ORDER_HEADER}'列")

    store_columns = []
    for i in range(total_col_idx + 1, len(headers)):
        if headers[i] is not None and str(headers[i]).strip():
            store_columns.append((i, str(headers[i]).strip()))

    rows = []
    for row in ws.iter_rows(min_row=2):
        vals = [cell.value for cell in row]
        if any(v is not None and str(v).strip() != "" for v in vals):
            rows.append(vals)

    return total_col_idx, store_columns, rows


def merge_into_template(data_rows, total_col_idx, store_columns, template_file, output_file,
                        category_sums=None):
    wb = load_workbook(template_file)
    ws = wb.active

    store_count = len(store_columns)
    last_col = 6 + store_count
    template_max_col = ws.max_column

    column_map = list(FIXED_COLUMN_MAP)
    column_map.append((total_col_idx, 6))
    for offset, (src_idx, _) in enumerate(store_columns):
        column_map.append((src_idx, 7 + offset))

    # 更新第7行门店列标题
    for offset, (_, name) in enumerate(store_columns):
        ws.cell(row=HEADER_ROW, column=7 + offset).value = name

    # 清除模板中多余的门店列
    for col in range(last_col + 1, template_max_col + 1):
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.font = copy.copy(cell.font)
            cell.border = copy.copy(cell.border)

    # 如果数据门店列多于模板，从模板最后一个门店列复制样式到新列
    if last_col > template_max_col:
        for row in range(1, ws.max_row + 1):
            src = ws.cell(row=row, column=template_max_col)
            for new_col in range(template_max_col + 1, last_col + 1):
                copy_cell_style(src, ws.cell(row=row, column=new_col))

    # 更新顶部汇总区公式
    first_store = get_column_letter(7)
    last_store = get_column_letter(last_col)
    for r in range(2, 6):
        ws.cell(row=r, column=6).value = f"=SUM({first_store}{r}:{last_store}{r})"
    ws.cell(row=6, column=6).value = f"=SUM({first_store}6:{last_store}6)"

    # 保存示例行样式
    sample_cells = list(ws.iter_rows(
        min_row=SAMPLE_ROW_START,
        max_row=SAMPLE_ROW_START
    ))[0]

    ws.delete_rows(SAMPLE_ROW_START, SAMPLE_ROW_COUNT)
    ws.insert_rows(DATA_START_ROW, len(data_rows))

    for i, row_data in enumerate(data_rows):
        excel_row = DATA_START_ROW + i

        ws.cell(row=excel_row, column=1).value = f"=ROW()-{HEADER_ROW}"

        for src_idx, dst_col in column_map:
            value = row_data[src_idx] if src_idx < len(row_data) else None
            ws.cell(row=excel_row, column=dst_col).value = value

        for col_idx in range(1, last_col + 1):
            if col_idx <= len(sample_cells):
                src = sample_cells[col_idx - 1]
            else:
                src = sample_cells[len(sample_cells) - 1]
            copy_cell_style(src, ws.cell(row=excel_row, column=col_idx))

    last_data_row = DATA_START_ROW + len(data_rows) - 1

    # 门店列 rows 2-5：填入分类汇总金额
    for offset, (_, store_name) in enumerate(store_columns):
        col_num = 7 + offset
        for r in range(2, 6):
            if category_sums:
                ws.cell(row=r, column=col_num).value = category_sums.get((r, store_name), 0)
            else:
                ws.cell(row=r, column=col_num).value = 0

    # 更新第6行门店列公式：统一格式 =SUM({col}2:{col}5)
    for col_num in range(7, last_col + 1):
        letter = get_column_letter(col_num)
        ws[f"{letter}6"] = f"=SUM({letter}2:{letter}5)"

    # G列之后的每列宽度与G列保持一致
    g_width = ws.column_dimensions["G"].width
    if g_width:
        for col_num in range(8, last_col + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = g_width

    wb.save(output_file)
    return last_data_row


def main():
    print(f"数据源: {DATA_FILE}")
    print(f"模板  : {TEMPLATE_FILE}")
    print(f"输出  : {OUTPUT_FILE}")
    print()

    if not DATA_FILE.exists():
        raise FileNotFoundError(f"数据源文件不存在: {DATA_FILE}")
    if not TEMPLATE_FILE.exists():
        raise FileNotFoundError(f"模板文件不存在: {TEMPLATE_FILE}")

    print("读取数据源...")
    total_col_idx, store_columns, data_rows = read_data_file(DATA_FILE)
    print(f"  共 {len(data_rows)} 行数据")
    print(f"  合计订货量列: {get_column_letter(total_col_idx + 1)} (索引 {total_col_idx})")
    print(f"  门店列 ({len(store_columns)}): {', '.join(name for _, name in store_columns)}")

    category_sums = {}
    if DETAIL_FILE.exists():
        store_names = {name for _, name in store_columns}
        print("读取明细数据，计算分类汇总金额...")
        category_sums = compute_category_sums(DETAIL_FILE, store_names)
        for (r, store), val in sorted(category_sums.items()):
            print(f"  Row {r} / {store}: {val}")
    else:
        print(f"  明细文件不存在，汇总区将填入 0: {DETAIL_FILE}")

    print("填入模板...")
    last_row = merge_into_template(data_rows, total_col_idx, store_columns, TEMPLATE_FILE, OUTPUT_FILE,
                                   category_sums)
    last_letter = get_column_letter(6 + len(store_columns))
    print(f"  数据区: A{DATA_START_ROW}:{last_letter}{last_row}")

    print()
    print(f"已生成: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
