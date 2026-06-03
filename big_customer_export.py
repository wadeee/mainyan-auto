"""
大客户导出 - 自动导出 & 格式化脚本
=====================================
依赖：pip install playwright openpyxl && playwright install chromium

自动登录后导出大客户订购商品统计表 (OrderProductReport)，
并填入格式化模板生成统计表。

用法：
    python big_customer_export.py                    # 导出后天的数据
    python big_customer_export.py --date 2026.05.30  # 指定日期
    python big_customer_export.py --days 2           # N天后（默认2=后天）
    python big_customer_export.py --headless         # 无头模式（不显示浏览器）
"""

import argparse
import copy
import logging
import logging.handlers
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

# ─── 日志配置 ───────────────────────────────────────────────────────────────────
LOG_DIR = Path(__file__).resolve().parent / "log"
LOG_DIR.mkdir(exist_ok=True)

_file_handler = logging.handlers.TimedRotatingFileHandler(
    LOG_DIR / "big_customer_export.log",
    when="midnight",
    backupCount=30,
    encoding="utf-8",
)
_file_handler.suffix = "%Y-%m-%d"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[_file_handler, logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# ─── 配置区（按需修改）─────────────────────────────────────────────────────────
ACCOUNT = "huomimayzb"
WORKER_ID = "M006"
PASSWORD = "tusijia88"

LOGIN_URL = "https://beta69.pospal.cn/"
ORDER_PRODUCT_REPORT_URL = "https://beta69.pospal.cn/EnterpriseReport/OrderProductReport"

OUTPUT_DIR = Path(__file__).resolve().parent / "大客户订购商品统计表"

TARGET_CATEGORIES = [
    "配送费",
    "包材耗材",
    "工衣模具",
    "慕斯+饼干+饮品+其他",
    "原料铺料",
    "冷冻面团",
    "蛋糕及面包成品及饼干类",
]

REPORT_EXCLUDED_STORES = {"焙满香滨江店", "焙满香广钢店"}

TEMPLATE_FILE = Path(__file__).resolve().parent / "大客户订购商品统计表_格式化模板.xlsx"

EXPORT_CATEGORY_MAP = {
    "面团": ["冷冻面团"],
    "成品面包饼干": ["成品面包类", "饼干类"],
    "蛋糕": ["蛋糕类"],
    "物料包材": ["热销类", "冷冻肉类", "冷冻馅料类", "冷藏馅料类", "油脂类", "粉类", "糖类", "常温馅料类", "干果类", "饼干类/外",
         "慕斯类/外", "饮品类/外", "其他/外", "专版包材类", "公版包材类", "工衣工帽围裙", "模具", "保洁用品", "配送费"],
}

HEADER_ROW = 2
SAMPLE_ROW = 3
TOTALS_ROW = 4
DATA_START_ROW = 3


# ─── 格式化合并函数 ──────────────────────────────────────────────────────────


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def read_report_data(report_file: Path):
    """读取大客户订购商品统计表（2行表头，数据从第3行开始，门店在J/L/N...每隔2列）"""
    wb = load_workbook(report_file, data_only=True)
    ws = wb.active

    stores = []
    for c in range(10, ws.max_column + 1, 2):
        name = ws.cell(row=1, column=c).value
        if name and str(name).strip():
            stores.append((c, str(name).strip()))
    stores = [(c, name) for c, name in stores if name not in REPORT_EXCLUDED_STORES]

    rows = []
    for r in range(3, ws.max_row + 1):
        first_col = ws.cell(row=r, column=1).value
        if first_col is not None and str(first_col).strip() == "合计":
            break
        name = ws.cell(row=r, column=2).value
        if name is None:
            continue
        quantities = {}
        for src_col, store_name in stores:
            val = ws.cell(row=r, column=src_col).value
            try:
                quantities[store_name] = float(val) if val else 0
            except (ValueError, TypeError):
                quantities[store_name] = 0
        rows.append({
            "name": ws.cell(row=r, column=2).value,
            "category": ws.cell(row=r, column=3).value,
            "spec": ws.cell(row=r, column=6).value,
            "unit": ws.cell(row=r, column=7).value,
            "quantities": quantities,
        })

    wb.close()
    logger.info(f"  大客户订购商品统计表: {len(rows)} 行, {len(stores)} 个门店")
    return [name for _, name in stores], rows


def merge_into_template(report_stores, report_rows,
                        template_file: Path, output_file: Path, target_date: str = None):
    wb = load_workbook(template_file)
    ws = wb.active

    all_stores = report_stores
    store_count = len(all_stores)
    last_col = 6 + store_count
    template_max_col = ws.max_column

    sample_cells = list(ws.iter_rows(min_row=SAMPLE_ROW, max_row=SAMPLE_ROW))[0]
    sample_height = ws.row_dimensions[SAMPLE_ROW].height
    totals_cells = list(ws.iter_rows(min_row=TOTALS_ROW, max_row=TOTALS_ROW))[0]
    totals_height = ws.row_dimensions[TOTALS_ROW].height

    all_rows = report_rows
    data_count = len(all_rows)

    ws.delete_rows(SAMPLE_ROW, 2)
    ws.insert_rows(DATA_START_ROW, data_count + 1)

    header_store_col = min(8, template_max_col)
    for r in range(1, DATA_START_ROW):
        src = ws.cell(row=r, column=header_store_col)
        for col in range(header_store_col + 1, last_col + 1):
            copy_cell_style(src, ws.cell(row=r, column=col))

    for col in range(last_col + 1, template_max_col + 1):
        for r in range(1, DATA_START_ROW):
            cell = ws.cell(row=r, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = None

    for i, store_name in enumerate(all_stores):
        cell = ws.cell(row=HEADER_ROW, column=7 + i)
        m = re.search(r'[（(](.+?)[）)]', str(store_name))
        short_name = m.group(1) if m else store_name
        short_name = re.sub(r'兔司家|甜麦面包屋', '', short_name).strip()
        cell.value = short_name

    for i, row_data in enumerate(all_rows):
        r = DATA_START_ROW + i
        ws.cell(row=r, column=1).value = i + 1
        ws.cell(row=r, column=2).value = row_data["name"]
        ws.cell(row=r, column=3).value = row_data["category"]
        ws.cell(row=r, column=4).value = row_data["spec"]
        ws.cell(row=r, column=5).value = row_data["unit"]

        first_letter = get_column_letter(7)
        last_letter = get_column_letter(last_col)
        ws.cell(row=r, column=6).value = f"=SUM({first_letter}{r}:{last_letter}{r})"

        for j, store_name in enumerate(all_stores):
            qty = row_data["quantities"].get(store_name, 0)
            col = 7 + j
            if qty and float(qty) != 0:
                val = float(qty)
                if val == int(val):
                    val = int(val)
                ws.cell(row=r, column=col).value = val

        for col_idx in range(1, last_col + 1):
            if col_idx >= 7:
                src_idx = min(7, len(sample_cells) - 1)
            else:
                src_idx = min(col_idx - 1, len(sample_cells) - 1)
            copy_cell_style(sample_cells[src_idx], ws.cell(row=r, column=col_idx))
        if sample_height:
            ws.row_dimensions[r].height = sample_height

    tr = DATA_START_ROW + data_count
    ws.cell(row=tr, column=1).value = data_count
    for col in [2, 3, 4, 5]:
        ws.cell(row=tr, column=col).value = "-"
    for col in range(6, last_col + 1):
        letter = get_column_letter(col)
        ws.cell(row=tr, column=col).value = f"=SUM({letter}{DATA_START_ROW}:{letter}{tr - 1})"

    for col_idx in range(1, last_col + 1):
        if col_idx >= 7:
            src_idx = min(7, len(totals_cells) - 1)
        else:
            src_idx = min(col_idx - 1, len(totals_cells) - 1)
        copy_cell_style(totals_cells[src_idx], ws.cell(row=tr, column=col_idx))
    if totals_height:
        ws.row_dimensions[tr].height = totals_height

    g_width = ws.column_dimensions["G"].width
    if g_width:
        for col_num in range(8, last_col + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = g_width

    for merge in list(ws.merged_cells.ranges):
        if merge.min_row == 1 and merge.max_row == 1:
            ws.unmerge_cells(str(merge))
    if last_col > 5:
        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=last_col)

    if data_count == 0:
        for col in [7, 8]:  # G, H
            for r in range(1, ws.max_row + 1):
                cell = ws.cell(row=r, column=col)
                if not isinstance(cell, MergedCell):
                    cell.value = None
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=3, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = None

    if target_date:
        target_dt = datetime.strptime(target_date, "%Y.%m.%d")
        prev_day = (target_dt - timedelta(days=2)).strftime("%Y.%m.%d")
        for cell_ref, new_date in [("A1", prev_day), ("C1", target_date)]:
            cell = ws[cell_ref]
            if cell.value:
                cell.value = re.sub(r"\d{4}\.\d{2}\.\d{2}", new_date, str(cell.value))

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    logger.info(f"  已生成格式化文件: {output_file}")


# ─── 浏览器自动化函数 ────────────────────────────────────────────────────────


def set_date(page, placeholder: str, value: str):
    page.evaluate(f"""
        (function() {{
            var inp = document.querySelector('input.timeInput.hasDatepicker[placeholder="{placeholder}"]');
            if (!inp) return 'not found';
            inp.value = "{value}";
            inp.dispatchEvent(new Event('input',  {{bubbles: true}}));
            inp.dispatchEvent(new Event('change', {{bubbles: true}}));
            inp.dispatchEvent(new Event('blur',   {{bubbles: true}}));
            return inp.value;
        }})()
    """)


def verify_dates(page, expected_date: str) -> bool:
    result = page.evaluate("""
        (function() {
            var r = [];
            document.querySelectorAll('input.timeInput.hasDatepicker').forEach(function(inp) {
                r.push(inp.placeholder + '=' + inp.value);
            });
            return r.join(' | ');
        })()
    """)
    logger.info(f"  [日期验证] {result}")
    return expected_date in result and result.count(expected_date) == 2


def click_by_text(page, text: str, desc: str = ""):
    result = page.evaluate(f"""
        (function() {{
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {{
                if (els[i].textContent.trim() === "{text}" && els[i].children.length === 0) {{
                    els[i].click();
                    return 'clicked tag=' + els[i].tagName + ' class=' + els[i].className;
                }}
            }}
            return 'not found';
        }})()
    """)
    tag = f"[{desc}] " if desc else ""
    logger.info(f"  {tag}→ {result}")
    return "clicked" in result


def login(page):
    logger.info("[1/4] 打开登录页面...")
    page.goto(LOGIN_URL)
    page.wait_for_load_state("networkidle")

    logger.info("[2/4] 切换到工号登录模式...")
    page.evaluate("""
        (function() {
            var els = document.querySelectorAll('div,span,a,button');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '工号登录' && els[i].children.length === 0) {
                    els[i].click();
                    return 'clicked';
                }
            }
            return 'not found';
        })()
    """)
    time.sleep(1.5)

    placeholder = page.evaluate("""
        (function() {
            var inputs = document.querySelectorAll('input');
            var r = [];
            for (var i = 0; i < inputs.length; i++) {
                if (inputs[i].placeholder) r.push(inputs[i].placeholder);
            }
            return r.join(' | ');
        })()
    """)
    if "员工工号" not in placeholder:
        raise RuntimeError(f"工号登录切换失败，当前输入框：{placeholder}")
    logger.info(f"  表单已切换 → {placeholder}")

    logger.info("[3/4] 填入账号/工号/密码...")
    page.evaluate(f"""
        (function() {{
            var inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].placeholder === '请输入账号')     inputs[i].value = "{ACCOUNT}";
                if (inputs[i].placeholder === '请输入员工工号') inputs[i].value = "{WORKER_ID}";
                if (inputs[i].placeholder === '请输入工号密码') inputs[i].value = "{PASSWORD}";
            }}
            inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].value) {{
                    inputs[i].dispatchEvent(new Event('input', {{bubbles: true}}));
                }}
            }}
        }})()
    """)

    time.sleep(1.5)

    logger.info("[4/4] 点击登录按钮...")
    click_by_text(page, "登 录", "登录")
    page.wait_for_load_state("networkidle", timeout=30_000)

    time.sleep(2)


def navigate_to_board(page, board_url: str, board_name: str):
    logger.info(f"  [导航] 前往{board_name}...")
    page.goto(board_url)
    page.wait_for_load_state("networkidle", timeout=30_000)
    logger.info(f"  已到达 → {page.url}")


def setup_filters(page, target_date: str):
    logger.info("  [筛选] 设置筛选条件...")

    logger.info("  → 切换到「期望到货时间」标签...")
    result = page.evaluate("""
        (function() {
            var lis = document.querySelectorAll('li');
            for (var i = 0; i < lis.length; i++) {
                if (lis[i].textContent.trim() === '期望到货时间') {
                    lis[i].click();
                    return 'clicked li[' + i + ']';
                }
            }
            return 'not found';
        })()
    """)
    logger.info(f"    {result}")
    if result == "not found":
        raise RuntimeError("未找到「期望到货时间」标签")

    tab_status = page.evaluate("""
        (function() {
            var r = [];
            document.querySelectorAll('li').forEach(function(li) {
                var t = li.textContent.trim();
                if (t === '订货时间' || t === '期望到货时间') {
                    r.push(t + ':' + li.className);
                }
            });
            return r.join(' | ');
        })()
    """)
    logger.info(f"    标签状态: {tab_status}")

    logger.info("  → 展开高级搜索面板...")
    page.evaluate("document.getElementById('advancedBtn').click()")
    time.sleep(0.5)

    logger.info("  → 打开分类选择弹框...")
    page.evaluate("document.getElementById('selectCategory').click()")
    time.sleep(1.0)

    logger.info(f"  → 勾选 {len(TARGET_CATEGORIES)} 个分类...")
    categories_js = str(TARGET_CATEGORIES).replace("'", '"')
    page.evaluate(f"""
        (function() {{
            var targets = {categories_js};
            var divs = document.querySelectorAll('.checkBoxDiv');
            divs.forEach(function(d) {{
                var span = d.querySelector('span');
                if (span && targets.indexOf(span.textContent.trim()) >= 0) {{
                    d.click();
                }}
            }});
        }})()
    """)

    check_result = page.evaluate(f"""
        (function() {{
            var targets = {categories_js};
            var divs = document.querySelectorAll('.checkBoxDiv');
            var r = [];
            divs.forEach(function(d) {{
                var s = d.querySelector('span');
                if (s && targets.indexOf(s.textContent.trim()) >= 0) {{
                    r.push(s.textContent.trim() + ':' + d.classList.contains('on'));
                }}
            }});
            return r.join(' | ');
        }})()
    """)
    logger.info(f"    勾选验证: {check_result}")
    if "false" in check_result or check_result.count(":true") < len(TARGET_CATEGORIES):
        logger.warning("部分分类未勾选，尝试重试...")
        page.evaluate(f"""
            (function() {{
                var targets = {categories_js};
                var divs = document.querySelectorAll('.checkBoxDiv');
                divs.forEach(function(d) {{
                    var span = d.querySelector('span');
                    if (span && targets.indexOf(span.textContent.trim()) >= 0 && !d.classList.contains('on')) {{
                        d.click();
                    }}
                }});
            }})()
        """)

    logger.info("  → 点击「确定」关闭弹框...")
    click_by_text(page, "确定", "确定弹框")
    time.sleep(0.5)

    logger.info(f"  → 设置日期: {target_date}...")
    set_date(page, "开始日期", f"{target_date} 00:00")
    set_date(page, "结束日期", f"{target_date} 23:59")


def search_and_count_rows(page, target_date: str, btn_id: str = None, max_retries: int = 3) -> int:
    logger.info("  [查询] 执行查询...")

    for attempt in range(1, max_retries + 1):
        logger.info(f"  查询第 {attempt} 次...")
        if btn_id:
            page.evaluate(f"document.getElementById('{btn_id}').click()")
        else:
            click_by_text(page, "查询", "查询")
        page.wait_for_load_state("networkidle", timeout=90_000)

        result = page.evaluate("""
            (function() {
                var rows = document.querySelectorAll('table tbody tr');
                var r = 'rows:' + rows.length + ' | ';
                document.querySelectorAll('input.timeInput.hasDatepicker').forEach(function(inp) {
                    r += inp.placeholder + '=' + inp.value + ' | ';
                });
                return r;
            })()
        """)
        logger.info(f"    结果: {result}")

        if target_date in result:
            m = re.search(r"rows:(\d+)", result)
            row_count = int(m.group(1)) if m else 0
            logger.info(f"  查询成功，共 {row_count} 行数据")
            return row_count
        else:
            logger.warning("日期被重置！重新设置日期...")
            set_date(page, "开始日期", f"{target_date} 00:00")
            set_date(page, "结束日期", f"{target_date} 23:59")
            if not verify_dates(page, target_date):
                logger.warning("日期验证失败")

    raise RuntimeError(f"查询 {max_retries} 次后日期仍然不正确，请手动检查")


def export_and_save(page, target_date: str, file_prefix: str, *, confirm_after_export: bool = False) -> Path:
    logger.info("  [导出] 导出文件...")
    time.sleep(3)

    date_str = target_date.replace(".", "-")
    daily_output_dir = OUTPUT_DIR / date_str / "原始下载"
    daily_output_dir.mkdir(parents=True, exist_ok=True)
    logger.info(f"  → 输出至: {daily_output_dir}")

    with page.expect_download(timeout=60_000) as dl_info:
        result = page.evaluate("""
            (function() {
                var els = document.querySelectorAll('*');
                for (var i = 0; i < els.length; i++) {
                    if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                        els[i].click();
                        return 'ok';
                    }
                }
                return 'not found';
            })()
        """)
        logger.info(f"  点击导出: {result}")
        if result == "not found":
            raise RuntimeError("未找到「导出」按钮")

        if confirm_after_export:
            time.sleep(2)
            logger.info("  → 点击「确定」确认导出...")
            confirm_result = page.evaluate("""
                (function() {
                    var candidates = [];
                    var els = document.querySelectorAll('*');
                    for (var i = 0; i < els.length; i++) {
                        if (els[i].textContent.trim() === '确定' && els[i].children.length === 0 && els[i].offsetParent !== null) {
                            candidates.push(els[i]);
                        }
                    }
                    if (candidates.length > 0) {
                        var btn = candidates[candidates.length - 1];
                        btn.click();
                        return 'clicked last visible confirm tag=' + btn.tagName + ' class=' + btn.className;
                    }
                    return 'not found';
                })()
            """)
            logger.info(f"  [确定导出] → {confirm_result}")

    download = dl_info.value
    logger.info(f"  下载文件名: {download.suggested_filename}")

    dest = daily_output_dir / f"{file_prefix}_{date_str}.xlsx"
    download.save_as(dest)
    logger.info(f"  已保存到: {dest}")
    return dest


def main():
    parser = argparse.ArgumentParser(description="Pospal 大客户自动导出")
    parser.add_argument("--date", type=str, help="指定日期，格式 YYYY.MM.DD，如 2026.05.30")
    parser.add_argument("--days", type=int, default=2, help="今天之后第N天（默认2=后天）")
    parser.add_argument("--headless", action="store_true", help="无头模式（不显示浏览器窗口）")
    args = parser.parse_args()

    if args.date:
        target_date = args.date
    else:
        target_dt = datetime.now() + timedelta(days=args.days)
        target_date = target_dt.strftime("%Y.%m.%d")

    logger.info(f"{'=' * 55}")
    logger.info(f"  Pospal 大客户导出")
    logger.info(f"  目标日期：{target_date}")
    logger.info(f"  输出根目录：{OUTPUT_DIR}")
    logger.info(f"{'=' * 55}")

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        logger.error("未安装 playwright，请先运行: pip install playwright && playwright install chromium")
        sys.exit(1)

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=args.headless,
            args=["--start-maximized"],
        )
        context = browser.new_context(
            viewport={"width": 1600, "height": 900},
            accept_downloads=True,
        )
        page = context.new_page()

        try:
            login(page)

            # ── 任务 1：大客户订购商品统计表 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  任务 1/2：下载大客户订购商品统计表")
            logger.info(f"{'─' * 55}")

            navigate_to_board(page, ORDER_PRODUCT_REPORT_URL, "大客户订购商品统计表")
            setup_filters(page, target_date)
            report_row_count = search_and_count_rows(page, target_date)
            report_path = export_and_save(page, target_date, "大客户订购商品统计表", confirm_after_export=True)

            logger.info(f"{'=' * 55}")
            logger.info(f"  下载完成！")
            logger.info(f"  大客户订购商品统计表：{report_row_count} 行 → {report_path}")
            logger.info(f"{'=' * 55}\n")

            # ── 任务 2：格式化（按分类分表）──
            logger.info(f"{'─' * 55}")
            logger.info(f"  任务 2/2：生成格式化统计表（{len(EXPORT_CATEGORY_MAP) + 1} 个分表）")
            logger.info(f"{'─' * 55}")

            date_str = target_date.replace(".", "-")
            report_stores, report_rows = read_report_data(report_path)

            all_output = OUTPUT_DIR / date_str / f"大客户订购商品统计表_格式化_全部_{date_str}.xlsx"
            merge_into_template(
                report_stores, report_rows,
                TEMPLATE_FILE, all_output,
                target_date=target_date,
            )
            logger.info(f"  [全部] {len(report_rows)} 行 → {all_output}")

            for export_name, export_cats in EXPORT_CATEGORY_MAP.items():
                export_cat_set = set(export_cats)
                filtered_rows = [
                    row for row in report_rows
                    if row["category"] is not None and str(row["category"]).strip() in export_cat_set
                ]
                logger.info(f"  [{export_name}] 匹配 {len(filtered_rows)} 行")
                if not filtered_rows:
                    logger.info(f"  [{export_name}] 无数据，跳过")
                    continue

                output_file = OUTPUT_DIR / date_str / f"大客户订购商品统计表_{export_name}_{date_str}.xlsx"
                merge_into_template(
                    report_stores, filtered_rows,
                    TEMPLATE_FILE, output_file,
                    target_date=target_date,
                )

                if export_name == "面团":
                    wb_tmp = load_workbook(output_file)
                    ws_tmp = wb_tmp.active
                    for r in range(DATA_START_ROW, ws_tmp.max_row + 1):
                        cell = ws_tmp.cell(row=r, column=2)
                        if cell.value is not None and str(cell.value).strip() != "-":
                            cell.font = cell.font.copy(size=12)
                    wb_tmp.save(output_file)
                    wb_tmp.close()

                logger.info(f"  [{export_name}] → {output_file}")

            logger.info(f"{'=' * 55}")
            logger.info(f"  全部完成！")
            logger.info(f"{'=' * 55}\n")

        except Exception as e:
            logger.error(f"任务失败: {e}")
            try:
                screenshot = OUTPUT_DIR / "error_screenshot.png"
                page.screenshot(path=str(screenshot))
                logger.info(f"  错误截图已保存: {screenshot}")
            except Exception:
                pass
            raise
        finally:
            context.close()
            browser.close()


if __name__ == "__main__":
    main()
