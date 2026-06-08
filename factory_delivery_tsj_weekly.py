"""
工厂配送兔司家门店周度销售报表 - 自动化脚本
=====================================
依赖：pip install playwright && playwright install chromium

自动登录 Pospal 后台，依次完成工厂配送兔司家门店周度销售报表所需的数据导出与处理。

用法：
    python factory_delivery_tsj_weekly.py                    # 导出本周数据（周一到周日）
    python factory_delivery_tsj_weekly.py --weeks -1         # 导出上周数据
    python factory_delivery_tsj_weekly.py --date 2026.06.03  # 根据日期推断所在周
    python factory_delivery_tsj_weekly.py --headless          # 无头模式（不显示浏览器）
"""

import argparse
import copy
import logging
import logging.handlers
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.utils import get_column_letter

# ─── 日志配置 ───────────────────────────────────────────────────────────────────
LOG_DIR = Path(__file__).resolve().parent / "log"
LOG_DIR.mkdir(exist_ok=True)

_file_handler = logging.handlers.TimedRotatingFileHandler(
    LOG_DIR / "factory_delivery_tsj_weekly.log",
    when="midnight",
    backupCount=30,
    encoding="utf-8",
)
_file_handler.suffix = "%Y-%m-%d"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[_file_handler, logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# ─── 配置区（按需修改）─────────────────────────────────────────────────────────
ACCOUNT = "huomimayzb"
WORKER_ID = "M006"
PASSWORD = "tusijia88"

LOGIN_URL = "https://beta69.pospal.cn/"
DELIVERY_COMPARISON_URL = "https://css69.pospal.cn/EnterpriseReport/DeliveryComparison"
DELIVERY_PRODUCT_COMPARISON_URL = "https://css69.pospal.cn/EnterpriseReport/DeliveryProductComparison"

OUTPUT_DIR = Path(__file__).resolve().parent / "工厂配送兔司家门店周度销售报表"
TEMPLATE_FILE = Path(__file__).resolve().parent / "工厂配送兔司家门店周度销售报表_格式化模板.xlsx"

HEADER_ROW = 2
TOTALS_ROW = 3
DATA_START_ROW = 4
SAMPLE_ROW = 4

CUSTOMER_NAME_MAP = {
    "焙满香广钢店": "广钢店",
    "焙满香滨江店": "滨江店",
    "兔司家东莞寮厦店": "寮厦店",
    "兔司家海悦城店": "海悦城店",
    "兔司家-西樵凰樵圣堡店": "圣堡店",
    "兔司家佛山禾仰广场店": "禾仰店",
    "兔司家东莞凤岗店": "凤岗店",
    "兔司家-信步闲庭店": "信步店",
    "兔司家-萝岗和苑店": "萝岗和苑店",
    "兔司家金碧新城店": "金碧店",
    "兔司家广园新村店": "广园店",
    "兔司家白云尚城店": "尚城店",
    "南安市场甜麦面包屋": "南安市场店",
    "南村店番禺区": "南村店",
    "花都保利城": "花都保利城店",
    "兔司家花都花侨店": "花都花侨店",
    "增城包多多店": "包多多店",
    "包多多2店": "包多多2店",
    "兔司家佛山同济广场店": "同济店",
    "兔司家花海湾店": "花海湾店",
    "兔司家高志大厦店": "高志店",
    "兔司家珠江花城店": "珠江花城店",
    "兔司家白云同和店": "同和店",
    "兔司家保利罗兰店": "保利罗兰店",
    "兔司家白云永泰店": "永泰店",
    "兔司家海珠南洲店": "海珠南洲店",
    "兔司家荔湾国际城店": "国际城店",
}

EXPORT_CATEGORY_MAP = {
    "面团": ["冷冻面团", "面团现烤类", "面团包装类", "面团丹麦类", "面团吐司类", "面团定制款类"],
    "成品面包饼干": ["成品面包类", "饼干类"],
    "蛋糕": ["蛋糕类"],
    "物料包材": ["热销类", "慕斯类/外", "冷冻肉类", "冷冻馅料类", "冷藏馅料类", "油脂类", "粉类", "糖类",
                 "常温馅料类", "干果类", "饼干类/外", "饮品类/外", "其他/外", "专版包材类", "公版包材类",
                 "工衣工帽围裙", "模具", "保洁用品", "配送费"],
}

S2_TOTALS_ROW = 4
S2_DATA_START_ROW = 5
S2_SAMPLE_ROW = 5

S2_CATEGORY_ORDER = [
    "冷冻面团",
    "蛋糕及面包成品及饼干类",
    "慕斯+饼干+饮品+其他",
    "原料铺料",
    "包材耗材",
    "工衣模具",
    "配送费",
]

# ─── 格式化合并函数 ──────────────────────────────────────────────────────────


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def _read_headers(ws):
    """读取第一行作为表头，返回 {列名: 列号}。"""
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val:
            headers[str(val).strip()] = c
    return headers


def _is_stop_name(name):
    """汇总行（None / '-' / '合计'）停止读取。"""
    if name is None:
        return True
    s = str(name).strip()
    return s in ("-", "合计")


def _read_name_rows(ws, headers, extra_cols=None):
    """逐行读取大客户名称，遇到汇总行停止。
    返回 list[dict]，每行包含 大客户名称 及 extra_cols 指定的列（值已转为数字）。
    """
    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=headers["大客户名称"]).value
        if _is_stop_name(name):
            break
        row = {"大客户名称": str(name).strip()}
        if extra_cols:
            for col_name in extra_cols:
                row[col_name] = _to_num(ws.cell(row=r, column=headers[col_name]).value)
        rows.append(row)
    return rows


def read_product_comparison(data_file: Path):
    """读取「仓库配送商品大客户对比表」，按大客户聚合，返回 list[dict]。

    聚合后的字段对应目标模板：
        实际销售金额 → C(实际出库金额/含运费)
        销售量       → K(出库量)
        退货量       → L(退货量)
        实际销售量   → M(实际出库量)
        销售金额     → N(出库金额)
        退货金额     → O(退货金额)
    """
    wb = load_workbook(data_file, data_only=True)
    ws = wb.active
    headers = _read_headers(ws)

    sum_cols = ["销售量", "退货量", "实际销售量", "销售金额", "退货金额", "实际销售金额"]

    customer_data = {}  # name -> {col: sum}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=headers["大客户名称"]).value
        if _is_stop_name(name):
            break
        name = str(name).strip()
        if name not in customer_data:
            customer_data[name] = {c: 0.0 for c in sum_cols}
        for col_name in sum_cols:
            val = ws.cell(row=r, column=headers[col_name]).value
            if val:
                try:
                    customer_data[name][col_name] += float(val)
                except (ValueError, TypeError):
                    pass
    wb.close()

    rows = []
    for name, d in customer_data.items():
        rows.append({
            "大客户名称": name,
            "实际出库金额": d["实际销售金额"],  # → C
            "出库量": d["销售量"],              # → K
            "退货量": d["退货量"],              # → L
            "实际出库量": d["实际销售量"],      # → M
            "出库金额": d["销售金额"],          # → N
            "退货金额": d["退货金额"],          # → O
        })
    return rows


def read_product_comparison_detail(data_file: Path):
    """读取「仓库配送商品大客户对比表」原始明细行，不做聚合。

    返回 list[dict]，每行包含：
        大客户名称, 商品名称, 商品大类, 商品分类, 规格, 单位, 实际销售量, 实际销售金额
    """
    wb = load_workbook(data_file, data_only=True)
    ws = wb.active
    headers = _read_headers(ws)

    detail_cols = ["大客户名称", "商品名称", "商品大类", "商品分类", "规格", "单位", "实际销售量", "实际销售金额"]
    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=headers["大客户名称"]).value
        if _is_stop_name(name):
            break
        row = {}
        for col_name in detail_cols:
            val = ws.cell(row=r, column=headers[col_name]).value
            row[col_name] = val
        row["大客户名称"] = str(row["大客户名称"]).strip()
        row["商品名称"] = str(row["商品名称"]).strip()
        rows.append(row)
    wb.close()
    return rows


def read_delivery_comparison(data_file: Path):
    """读取「仓库配送大客户对比表」，提取 销售出库单数(I) 和 销售退货单数(J)。"""
    wb = load_workbook(data_file, data_only=True)
    ws = wb.active
    headers = _read_headers(ws)
    rows = _read_name_rows(ws, headers, extra_cols=["销售出库单数", "销售退货单数"])
    wb.close()
    return rows


def read_delivery_fee(data_file: Path):
    """读取「仓库配送大客户对比表_配送费」，返回 {大客户名称: {配送次数, 运费金额}}。

    配送次数(E) ← 出库量, 运费金额(F) ← 实际出库金额
    """
    wb = load_workbook(data_file, data_only=True)
    ws = wb.active
    headers = _read_headers(ws)

    fee_map = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=headers["大客户名称"]).value
        if _is_stop_name(name):
            break
        fee_map[str(name).strip()] = {
            "配送次数": _to_num(ws.cell(row=r, column=headers["出库量"]).value),     # → E
            "运费金额": _to_num(ws.cell(row=r, column=headers["实际出库金额"]).value),  # → F
        }
    wb.close()
    return fee_map


def read_self_product(data_file: Path):
    """读取「仓库配送大客户对比表_自产品」，返回 {大客户名称: 自产品金额}。

    自产品金额(H) ← 实际出库金额
    """
    wb = load_workbook(data_file, data_only=True)
    ws = wb.active
    headers = _read_headers(ws)

    sp_map = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=headers["大客户名称"]).value
        if _is_stop_name(name):
            break
        sp_map[str(name).strip()] = _to_num(ws.cell(row=r, column=headers["实际出库金额"]).value)
    wb.close()
    return sp_map


def _get_category_sort_key(product_category):
    """返回 (大类序号, 子分类序号) 用于按 EXPORT_CATEGORY_MAP 排序。"""
    for group_idx, (_group, subcats) in enumerate(EXPORT_CATEGORY_MAP.items()):
        for sub_idx, sub in enumerate(subcats):
            if sub == product_category:
                return (group_idx, sub_idx)
    return (999, 0)


def _get_store_order_from_s1(wb):
    """从 Sheet 1 数据行中按顺序提取门店列表（使用简称），返回有序列表。"""
    ws = wb[wb.sheetnames[0]]
    store_order = []
    seen = set()
    for r in range(DATA_START_ROW, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if name is None:
            continue
        name = str(name).strip()
        if not name or name == "合计":
            break
        short = CUSTOMER_NAME_MAP.get(name, name)
        if short not in seen:
            seen.add(short)
            store_order.append(short)
    return store_order


def fill_product_comparison_sheet(wb, product_detail_rows, monday, sunday, store_order):
    """将 B 表明细数据透视后写入模板的第二个 sheet。
    store_order: 门店简称有序列表，决定从 I 列开始的列布局。
    返回 sorted_products 列表供 S3 使用。
    """
    ws = wb[wb.sheetnames[1]]
    original_s2_max_col = ws.max_column

    s2_merged = list(ws.merged_cells.ranges)
    for mr in s2_merged:
        ws.unmerge_cells(str(mr))

    # 根据 store_order 动态构建列映射: {简称: (销量列, 销售额列)}
    store_columns = {}
    for idx, short_name in enumerate(store_order):
        qty_col = 9 + idx * 2
        amt_col = 10 + idx * 2
        store_columns[short_name] = (qty_col, amt_col)

    store_count = len(store_order)
    s2_max_col = 8 + store_count * 2  # 最后一列

    pivot = {}
    for row in product_detail_rows:
        pname = row["商品名称"]
        store_short = CUSTOMER_NAME_MAP.get(row["大客户名称"])
        if not store_short:
            continue

        if pname not in pivot:
            pivot[pname] = {
                "info": {
                    "商品名称": pname,
                    "商品大类": row.get("商品大类"),
                    "商品分类": row.get("商品分类"),
                    "规格": row.get("规格"),
                    "单位": row.get("单位"),
                },
                "stores": {},
            }
        stores = pivot[pname]["stores"]
        if store_short not in stores:
            stores[store_short] = {"销量": 0.0, "销售额": 0.0}
        qty = row.get("实际销售量")
        amt = row.get("实际销售金额")
        if qty:
            try:
                stores[store_short]["销量"] += float(qty)
            except (ValueError, TypeError):
                pass
        if amt:
            try:
                stores[store_short]["销售额"] += float(amt)
            except (ValueError, TypeError):
                pass

    def _s2_sort_key(p):
        cat = p["info"].get("商品大类", "") or ""
        try:
            cat_idx = S2_CATEGORY_ORDER.index(cat)
        except ValueError:
            cat_idx = len(S2_CATEGORY_ORDER)
        total_amt = sum(s["销售额"] for s in p["stores"].values())
        return (cat_idx, -total_amt)

    sorted_products = sorted(pivot.values(), key=_s2_sort_key)

    data_count = len(sorted_products)
    if data_count == 0:
        return []

    sample_row_num = S2_SAMPLE_ROW
    sample_cells = list(ws.iter_rows(min_row=sample_row_num, max_row=sample_row_num))[0]
    sample_height = ws.row_dimensions[sample_row_num].height

    sample_count = ws.max_row - S2_SAMPLE_ROW + 1
    ws.delete_rows(S2_SAMPLE_ROW, sample_count)
    ws.insert_rows(S2_DATA_START_ROW, data_count)

    # 动态销量/销售额列字母列表
    qty_letters = [get_column_letter(9 + i * 2) for i in range(store_count)]
    amt_letters = [get_column_letter(10 + i * 2) for i in range(store_count)]

    for i, product in enumerate(sorted_products):
        r = S2_DATA_START_ROW + i
        info = product["info"]

        ws.cell(row=r, column=1).value = f"=ROW()-4"
        ws.cell(row=r, column=2).value = info["商品名称"]
        ws.cell(row=r, column=3).value = info.get("商品大类")
        ws.cell(row=r, column=4).value = info.get("商品分类")
        ws.cell(row=r, column=5).value = info.get("规格")
        ws.cell(row=r, column=6).value = info.get("单位")

        ws.cell(row=r, column=7).value = f"=SUM({','.join(c + str(r) for c in qty_letters)})"
        ws.cell(row=r, column=8).value = f"=SUM({','.join(c + str(r) for c in amt_letters)})"

        for store_short, (qty_col, amt_col) in store_columns.items():
            store_data = product["stores"].get(store_short)
            if store_data:
                qty_val = _to_num(store_data["销量"])
                amt_val = _to_num(store_data["销售额"])
                if qty_val is not None:
                    ws.cell(row=r, column=qty_col).value = qty_val
                if amt_val is not None:
                    ws.cell(row=r, column=amt_col).value = amt_val

        for col_idx in range(1, s2_max_col + 1):
            src_idx = min(col_idx - 1, len(sample_cells) - 1)
            copy_cell_style(sample_cells[src_idx], ws.cell(row=r, column=col_idx))
        if sample_height:
            ws.row_dimensions[r].height = sample_height

    last_data_row = S2_DATA_START_ROW + data_count - 1
    ws.cell(row=S2_TOTALS_ROW, column=1).value = "合计"
    for col in range(7, s2_max_col + 1):
        letter = get_column_letter(col)
        ws.cell(row=S2_TOTALS_ROW, column=col).value = (
            f"=SUM({letter}{S2_DATA_START_ROW}:{letter}{last_data_row})"
        )

    # 写入门店简称到表头 (第2行对应各门店名称，第3行对应数量/销售额)
    for short_name, (qty_col, amt_col) in store_columns.items():
        ws.cell(row=2, column=qty_col).value = short_name
        ws.cell(row=3, column=qty_col).value = "数量"
        ws.cell(row=3, column=amt_col).value = "销售额"

    # 恢复 Sheet 2 合并单元格（动态范围）
    last_col_letter = get_column_letter(s2_max_col)
    ws.merge_cells(f"A1:{last_col_letter}1")
    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        ws.merge_cells(f"{col_letter}2:{col_letter}3")
    ws.merge_cells("G2:H2")
    for idx in range(store_count):
        start_col = 9 + idx * 2
        sl = get_column_letter(start_col)
        el = get_column_letter(start_col + 1)
        ws.merge_cells(f"{sl}2:{el}2")
    ws.merge_cells("A4:F4")

    # 清除模板中多余的门店列（当实际门店数少于模板预留时）
    if s2_max_col < original_s2_max_col:
        for r in range(1, ws.max_row + 1):
            for c in range(s2_max_col + 1, original_s2_max_col + 1):
                ws.cell(row=r, column=c).value = None

    return sorted_products


S3_TOTALS_ROW = 3
S3_DATA_START_ROW = 4
S3_SAMPLE_ROW = 4


def fill_sales_ranking_sheet(wb, sorted_products):
    """将 S2 透视数据按销售额降序写入 S3（门店产品销售排行表）。
    仅输出 A-H 列（序号、商品名称、商品大类、商品分类、规格、单位、合计数量、合计销售额）。
    """
    ws = wb[wb.sheetnames[2]]
    original_s3_max_col = ws.max_column

    s3_merged = list(ws.merged_cells.ranges)
    for mr in s3_merged:
        ws.unmerge_cells(str(mr))

    s3_max_col = 8  # 仅到 H 列

    data_count = len(sorted_products)
    if data_count == 0:
        return

    sample_cells = list(ws.iter_rows(min_row=S3_SAMPLE_ROW, max_row=S3_SAMPLE_ROW))[0]
    sample_height = ws.row_dimensions[S3_SAMPLE_ROW].height

    sample_count = ws.max_row - S3_SAMPLE_ROW + 1
    ws.delete_rows(S3_SAMPLE_ROW, sample_count)
    ws.insert_rows(S3_DATA_START_ROW, data_count)

    for i, product in enumerate(sorted_products):
        r = S3_DATA_START_ROW + i
        info = product["info"]

        ws.cell(row=r, column=1).value = f"=ROW()-3"
        ws.cell(row=r, column=2).value = info["商品名称"]
        ws.cell(row=r, column=3).value = info.get("商品大类")
        ws.cell(row=r, column=4).value = info.get("商品分类")
        ws.cell(row=r, column=5).value = info.get("规格")
        ws.cell(row=r, column=6).value = info.get("单位")

        # 合计数量和销售额直接从透视数据计算
        total_qty = sum(_to_num(s["销量"]) or 0 for s in product["stores"].values())
        total_amt = sum(_to_num(s["销售额"]) or 0 for s in product["stores"].values())
        ws.cell(row=r, column=7).value = _to_num(total_qty)
        ws.cell(row=r, column=8).value = _to_num(total_amt)

        for col_idx in range(1, s3_max_col + 1):
            src_idx = min(col_idx - 1, len(sample_cells) - 1)
            copy_cell_style(sample_cells[src_idx], ws.cell(row=r, column=col_idx))
        if sample_height:
            ws.row_dimensions[r].height = sample_height

    last_data_row = S3_DATA_START_ROW + data_count - 1
    ws.cell(row=S3_TOTALS_ROW, column=1).value = "合计"
    for col in range(7, s3_max_col + 1):
        letter = get_column_letter(col)
        ws.cell(row=S3_TOTALS_ROW, column=col).value = (
            f"=SUM({letter}{S3_DATA_START_ROW}:{letter}{last_data_row})"
        )

    # 清除模板中 I 列及之后的多余内容
    if original_s3_max_col > s3_max_col:
        for r in range(1, ws.max_row + 1):
            for c in range(s3_max_col + 1, original_s3_max_col + 1):
                ws.cell(row=r, column=c).value = None

    # 恢复 S3 合并单元格（仅 A-H 范围）
    last_col_letter = get_column_letter(s3_max_col)
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws.merge_cells("G2:H2")
    ws.merge_cells("A3:F3")


def merge_delivery_into_template(product_rows, delivery_rows, template_file: Path,
                                  output_file: Path, monday: datetime, sunday: datetime,
                                  fee_data: dict = None, self_product_data: dict = None,
                                  product_detail_rows: list = None):
    """将多个数据源合并写入模板。

    product_rows:        来自 仓库配送商品大客户对比表（聚合后），提供 C, K, L, M, N, O
    delivery_rows:       来自 仓库配送大客户对比表，提供 I(销售出库单数), J(销售退货单数)
    fee_data:            来自 仓库配送大客户对比表_配送费，提供 E(配送次数), F(运费金额)
    self_product_data:   来自 仓库配送大客户对比表_自产品，提供 H(自产品金额)
    product_detail_rows: 来自 仓库配送商品大客户对比表（明细），填入 Sheet 2
    """
    wb = load_workbook(template_file, rich_text=True)
    ws = wb.active

    max_col = 15

    # 解除所有合并单元格，避免 delete_rows/insert_rows 后 MergedCell 只读问题
    merged_ranges = list(ws.merged_cells.ranges)
    for mr in merged_ranges:
        ws.unmerge_cells(str(mr))

    sample_cells = list(ws.iter_rows(min_row=SAMPLE_ROW, max_row=SAMPLE_ROW))[0]
    sample_height = ws.row_dimensions[SAMPLE_ROW].height
    totals_cells = list(ws.iter_rows(min_row=TOTALS_ROW, max_row=TOTALS_ROW))[0]

    # 按 实际出库金额(C) 降序排列
    data_rows_sorted = sorted(product_rows,
                              key=lambda r: float(r.get("实际出库金额", 0) or 0),
                              reverse=True)
    data_count = len(data_rows_sorted)

    # 将 delivery_rows 转为 {name: row} 以便按名称查找 I, J
    delivery_map = {str(r["大客户名称"]).strip(): r for r in delivery_rows}

    ws.delete_rows(SAMPLE_ROW, 2)
    ws.insert_rows(DATA_START_ROW, data_count)

    for i, row_data in enumerate(data_rows_sorted):
        r = DATA_START_ROW + i
        name = str(row_data.get("大客户名称", "")).strip()
        ws.cell(row=r, column=1).value = f"=ROW()-3"
        ws.cell(row=r, column=2).value = name

        # C(3) = 实际出库金额/含运费  ← 商品对比表聚合
        ws.cell(row=r, column=3).value = _to_num(row_data.get("实际出库金额"))

        # E(5) = 配送次数, F(6) = 运费金额  ← 配送费表
        fee_info = fee_data.get(name) if fee_data else None
        if fee_info:
            ws.cell(row=r, column=5).value = fee_info.get("配送次数")
            ws.cell(row=r, column=6).value = fee_info.get("运费金额")

        # H(8) = 自产品金额  ← 自产品表
        if self_product_data:
            sp_val = self_product_data.get(name)
            if sp_val is not None:
                ws.cell(row=r, column=8).value = sp_val

        # D(4) = C - F（实际出库额/不含运费）
        ws.cell(row=r, column=4).value = f"=C{r}-F{r}"

        # G(7) = C - F - H（外购品金额）
        ws.cell(row=r, column=7).value = f"=C{r}-F{r}-H{r}"

        # I(9) = 销售出库单数, J(10) = 销售退货单数  ← 大客户对比表
        del_info = delivery_map.get(name)
        if del_info:
            ws.cell(row=r, column=9).value = del_info.get("销售出库单数")
            ws.cell(row=r, column=10).value = del_info.get("销售退货单数")

        # K(11)=出库量, L(12)=退货量, M(13)=实际出库量, N(14)=出库金额, O(15)=退货金额
        ws.cell(row=r, column=11).value = _to_num(row_data.get("出库量"))
        ws.cell(row=r, column=12).value = _to_num(row_data.get("退货量"))
        ws.cell(row=r, column=13).value = _to_num(row_data.get("实际出库量"))
        ws.cell(row=r, column=14).value = _to_num(row_data.get("出库金额"))
        ws.cell(row=r, column=15).value = _to_num(row_data.get("退货金额"))

        for col_idx in range(1, max_col + 1):
            src_idx = min(col_idx - 1, len(sample_cells) - 1)
            copy_cell_style(sample_cells[src_idx], ws.cell(row=r, column=col_idx))
        if sample_height:
            ws.row_dimensions[r].height = sample_height

    tr = TOTALS_ROW
    last_data_row = DATA_START_ROW + data_count - 1

    ws.cell(row=tr, column=1).value = "合计"
    for col in [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]:
        letter = get_column_letter(col)
        ws.cell(row=tr, column=col).value = f"=SUM({letter}{DATA_START_ROW}:{letter}{last_data_row})"

    for col_idx in range(1, max_col + 1):
        src_idx = min(col_idx - 1, len(totals_cells) - 1)
        copy_cell_style(totals_cells[src_idx], ws.cell(row=tr, column=col_idx))

    date_sheet = f"{monday.month}.{monday.day}-{sunday.month}.{sunday.day}"
    old_sheet_name = ws.title
    new_sheet_name = re.sub(r'\d{1,2}\.\d{1,2}-\d{1,2}\.\d{1,2}', date_sheet, old_sheet_name)
    ws.title = new_sheet_name

    cell_a1 = ws.cell(row=1, column=1)
    if cell_a1.value:
        m_start = monday.month
        d_start = monday.day
        m_end = sunday.month
        d_end = sunday.day
        y = monday.year

        date_pattern = r'\d{4}年\d{1,2}月\d{1,2}日至\d{1,2}月\d{1,2}日'
        date_replacement = f"{y}年{m_start}月{d_start}日至{m_end}月{d_end}日"

        if isinstance(cell_a1.value, CellRichText):
            new_blocks = []
            for block in cell_a1.value:
                if isinstance(block, TextBlock):
                    new_blocks.append(TextBlock(block.font, re.sub(date_pattern, date_replacement, block.text)))
                elif isinstance(block, str):
                    new_blocks.append(re.sub(date_pattern, date_replacement, block))
            cell_a1.value = CellRichText(*new_blocks)
        else:
            cell_a1.value = re.sub(date_pattern, date_replacement, str(cell_a1.value))

    # 恢复 Sheet 1 标题行合并
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

    # ── Sheet 2: 填充商品对比明细 ──
    sorted_products = []
    store_order = _get_store_order_from_s1(wb)
    if product_detail_rows:
        sorted_products = fill_product_comparison_sheet(wb, product_detail_rows, monday, sunday, store_order)

    # ── Sheet 3: 填充销售排行（数据来源于 S2）──
    if sorted_products:
        fill_sales_ranking_sheet(wb, sorted_products)

    # ── Sheet 2 & Sheet 3: 日期替换（sheet名 + A1标题）──
    date_sheet = f"{monday.month}.{monday.day}-{sunday.month}.{sunday.day}"
    date_pattern = r'\d{4}年\d{1,2}月\d{1,2}日至\d{1,2}月\d{1,2}日'
    date_replacement = f"{monday.year}年{monday.month}月{monday.day}日至{sunday.month}月{sunday.day}日"

    for sheet_idx in [1, 2]:
        if sheet_idx >= len(wb.sheetnames):
            break
        s = wb[wb.sheetnames[sheet_idx]]
        s.title = re.sub(r'\d{1,2}\.\d{1,2}-\d{1,2}\.\d{1,2}', date_sheet, s.title)
        cell = s.cell(row=1, column=1)
        if cell.value:
            if isinstance(cell.value, CellRichText):
                new_blocks = []
                for block in cell.value:
                    if isinstance(block, TextBlock):
                        new_blocks.append(TextBlock(block.font, re.sub(date_pattern, date_replacement, block.text)))
                    elif isinstance(block, str):
                        new_blocks.append(re.sub(date_pattern, date_replacement, block))
                cell.value = CellRichText(*new_blocks)
            else:
                cell.value = re.sub(date_pattern, date_replacement, str(cell.value))

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    logger.info(f"  已生成格式化文件: {output_file}")


def _to_num(val):
    if val is None:
        return None
    try:
        f = float(val)
        if f == 0:
            return None
        if f == int(f):
            return int(f)
        return f
    except (ValueError, TypeError):
        return val


# ─── 工具函数 ──────────────────────────────────────────────────────────────────


def get_week_range(weeks: int = 0, date_str: str = None):
    if date_str:
        base = datetime.strptime(date_str, "%Y.%m.%d")
    else:
        base = datetime.now()

    monday = base - timedelta(days=base.weekday())
    monday = monday + timedelta(weeks=weeks)
    sunday = monday + timedelta(days=6)

    return monday, sunday


# ─── 浏览器自动化函数 ────────────────────────────────────────────────────────


def set_date(page, placeholder: str, value: str):
    page.evaluate(f"""
        (function() {{
            var inp = document.querySelector('input.timeInput.hasDatepicker[placeholder="{placeholder}"]');
            if (!inp) return 'not found';
            inp.value = "{value}";
            inp.dispatchEvent(new Event('input',  {{bubbles: true}}));
            inp.dispatchEvent(new Event('change', {{bubbles: true}}));
            inp.dispatchEvent(new Event('blur',   {{bubbles: true}}));
            return inp.value;
        }})()
    """)


def click_by_text(page, text: str, desc: str = ""):
    result = page.evaluate(f"""
        (function() {{
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {{
                if (els[i].textContent.trim() === "{text}" && els[i].children.length === 0) {{
                    els[i].click();
                    return 'clicked tag=' + els[i].tagName + ' class=' + els[i].className;
                }}
            }}
            return 'not found';
        }})()
    """)
    tag = f"[{desc}] " if desc else ""
    logger.info(f"  {tag}→ {result}")
    return "clicked" in result


def login(page):
    logger.info("[1/4] 打开登录页面...")
    page.goto(LOGIN_URL)
    page.wait_for_load_state("networkidle")

    logger.info("[2/4] 切换到工号登录模式...")
    page.evaluate("""
        (function() {
            var els = document.querySelectorAll('div,span,a,button');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '工号登录' && els[i].children.length === 0) {
                    els[i].click();
                    return 'clicked';
                }
            }
            return 'not found';
        })()
    """)
    time.sleep(1.5)

    placeholder = page.evaluate("""
        (function() {
            var inputs = document.querySelectorAll('input');
            var r = [];
            for (var i = 0; i < inputs.length; i++) {
                if (inputs[i].placeholder) r.push(inputs[i].placeholder);
            }
            return r.join(' | ');
        })()
    """)
    if "员工工号" not in placeholder:
        raise RuntimeError(f"工号登录切换失败，当前输入框：{placeholder}")
    logger.info(f"  表单已切换 → {placeholder}")

    logger.info("[3/4] 填入账号/工号/密码...")
    page.evaluate(f"""
        (function() {{
            var inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].placeholder === '请输入账号')     inputs[i].value = "{ACCOUNT}";
                if (inputs[i].placeholder === '请输入员工工号') inputs[i].value = "{WORKER_ID}";
                if (inputs[i].placeholder === '请输入工号密码') inputs[i].value = "{PASSWORD}";
            }}
            inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].value) {{
                    inputs[i].dispatchEvent(new Event('input', {{bubbles: true}}));
                }}
            }}
        }})()
    """)

    time.sleep(1.5)

    logger.info("[4/4] 点击登录按钮...")
    click_by_text(page, "登 录", "登录")
    page.wait_for_load_state("networkidle", timeout=120_000)

    time.sleep(2)


def main():
    parser = argparse.ArgumentParser(description="Pospal 工厂配送兔司家门店周度销售报表自动化脚本")
    parser.add_argument("--weeks", type=int, default=0, help="周偏移量：0=本周，-1=上周（默认0）")
    parser.add_argument("--date", type=str, help="指定日期推断所在周，格式 YYYY.MM.DD")
    parser.add_argument("--headless", action="store_true", help="无头模式（不显示浏览器窗口）")
    args = parser.parse_args()

    if args.date:
        monday, sunday = get_week_range(date_str=args.date)
    else:
        monday, sunday = get_week_range(weeks=args.weeks)

    monday_str = monday.strftime("%Y.%m.%d")
    sunday_str = sunday.strftime("%Y.%m.%d")
    date_range_str = f"{monday.strftime('%Y-%m-%d')}~{sunday.strftime('%Y-%m-%d')}"

    logger.info(f"{'=' * 55}")
    logger.info(f"  Pospal 工厂配送兔司家门店周度销售报表")
    logger.info(f"  目标周：{monday_str} ~ {sunday_str}")
    logger.info(f"  输出根目录：{OUTPUT_DIR}")
    logger.info(f"{'=' * 55}")

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        logger.error("未安装 playwright，请先运行: pip install playwright && playwright install chromium")
        sys.exit(1)

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=args.headless,
            args=["--start-maximized"],
        )
        context = browser.new_context(
            viewport={"width": 1600, "height": 900},
            accept_downloads=True,
        )
        page = context.new_page()
        page.set_default_timeout(120000)
        page.set_default_navigation_timeout(120000)

        try:
            login(page)

            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：下载仓库配送大客户对比表")
            logger.info(f"{'─' * 55}")

            logger.info("  [导航] 前往仓库配送大客户对比表...")
            page.goto(DELIVERY_COMPARISON_URL)
            page.wait_for_load_state("networkidle", timeout=120_000)
            logger.info(f"  已到达 → {page.url}")

            logger.info(f"  → 设置完成时间: {monday_str} ~ {sunday_str}...")
            set_date(page, "开始日期", f"{monday_str} 00:00")
            set_date(page, "结束日期", f"{sunday_str} 23:59")

            logger.info("  [查询] 执行查询...")
            click_by_text(page, "查询", "查询")
            page.wait_for_load_state("networkidle", timeout=150_000)
            time.sleep(3)

            logger.info("  [导出] 导出文件...")
            output_dir = OUTPUT_DIR / date_range_str / "原始下载"
            output_dir.mkdir(parents=True, exist_ok=True)

            with page.expect_download(timeout=180_000) as dl_info:
                result = page.evaluate("""
                    (function() {
                        var els = document.querySelectorAll('*');
                        for (var i = 0; i < els.length; i++) {
                            if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                                els[i].click();
                                return 'ok';
                            }
                        }
                        return 'not found';
                    })()
                """)
                logger.info(f"  点击导出: {result}")
                if result == "not found":
                    raise RuntimeError("未找到「导出」按钮")

            download = dl_info.value
            logger.info(f"  下载文件名: {download.suggested_filename}")

            dest = output_dir / f"仓库配送大客户对比表_{date_range_str}.xlsx"
            download.save_as(dest)
            logger.info(f"  已保存到: {dest}")

            # ── 步骤 1.5：仓库配送大客户对比表_配送费 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：下载仓库配送大客户对比表_配送费")
            logger.info(f"{'─' * 55}")

            logger.info("  → 点击更多搜索...")
            more_search_result = page.evaluate("""
                (function() {
                    var els = document.querySelectorAll('*');
                    for (var i = 0; i < els.length; i++) {
                        var t = els[i].textContent.trim();
                        if ((t === '更多搜索' || t === '更多搜索▼') && els[i].children.length === 0) {
                            els[i].click();
                            return 'clicked tag=' + els[i].tagName + ' class=' + els[i].className;
                        }
                    }
                    var btn = document.getElementById('advancedBtn');
                    if (btn) { btn.click(); return 'clicked advancedBtn'; }
                    return 'not found';
                })()
            """)
            logger.info(f"  [更多搜索] → {more_search_result}")
            time.sleep(0.5)

            logger.info("  → 打开分类选择弹框...")
            select_cat_result = page.evaluate("""
                (function() {
                    var btn = document.getElementById('selectCategory');
                    if (btn) { btn.click(); return 'clicked selectCategory'; }
                    var els = document.querySelectorAll('*');
                    for (var i = 0; i < els.length; i++) {
                        if (els[i].textContent.trim() === '选择分类' && els[i].children.length === 0) {
                            els[i].click();
                            return 'clicked tag=' + els[i].tagName;
                        }
                    }
                    return 'not found';
                })()
            """)
            logger.info(f"  [选择分类] → {select_cat_result}")
            time.sleep(1.0)

            logger.info("  → 勾选「配送费」...")
            page.evaluate("""
                (function() {
                    var divs = document.querySelectorAll('.checkBoxDiv');
                    divs.forEach(function(d) {
                        var span = d.querySelector('span');
                        if (span && span.textContent.trim() === '配送费') {
                            if (!d.classList.contains('on')) {
                                d.click();
                            }
                        }
                    });
                })()
            """)

            check_result = page.evaluate("""
                (function() {
                    var divs = document.querySelectorAll('.checkBoxDiv');
                    var r = [];
                    divs.forEach(function(d) {
                        var s = d.querySelector('span');
                        if (s && s.textContent.trim() === '配送费') {
                            r.push(s.textContent.trim() + ':' + d.classList.contains('on'));
                        }
                    });
                    return r.join(' | ');
                })()
            """)
            logger.info(f"    勾选验证: {check_result}")

            logger.info("  → 点击「确定」关闭弹框...")
            click_by_text(page, "确定", "确定弹框")
            time.sleep(0.5)

            logger.info("  [查询] 执行查询...")
            click_by_text(page, "查询", "查询")
            page.wait_for_load_state("networkidle", timeout=150_000)
            time.sleep(3)

            logger.info("  [导出] 导出配送费文件...")
            with page.expect_download(timeout=180_000) as dl_info_fee:
                result = page.evaluate("""
                    (function() {
                        var els = document.querySelectorAll('*');
                        for (var i = 0; i < els.length; i++) {
                            if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                                els[i].click();
                                return 'ok';
                            }
                        }
                        return 'not found';
                    })()
                """)
                logger.info(f"  点击导出: {result}")
                if result == "not found":
                    raise RuntimeError("未找到「导出」按钮")

            download_fee = dl_info_fee.value
            logger.info(f"  下载文件名: {download_fee.suggested_filename}")

            dest_fee = output_dir / f"仓库配送大客户对比表_配送费_{date_range_str}.xlsx"
            download_fee.save_as(dest_fee)
            logger.info(f"  已保存到: {dest_fee}")

            # ── 步骤 1.7：仓库配送大客户对比表_自产品 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：下载仓库配送大客户对比表_自产品")
            logger.info(f"{'─' * 55}")

            logger.info("  → 点击重新搜索...")
            click_by_text(page, "重新搜索", "重新搜索")
            time.sleep(1.0)

            logger.info("  → 打开分类选择弹框...")
            select_cat_result = page.evaluate("""
                (function() {
                    var btn = document.getElementById('selectCategory');
                    if (btn) { btn.click(); return 'clicked selectCategory'; }
                    var els = document.querySelectorAll('*');
                    for (var i = 0; i < els.length; i++) {
                        if (els[i].textContent.trim() === '选择分类' && els[i].children.length === 0) {
                            els[i].click();
                            return 'clicked tag=' + els[i].tagName;
                        }
                    }
                    return 'not found';
                })()
            """)
            logger.info(f"  [选择分类] → {select_cat_result}")
            time.sleep(1.0)

            logger.info("  → 取消「配送费」，勾选「冷冻面团」和「蛋糕及面包成品及饼干类」...")
            page.evaluate("""
                (function() {
                    var divs = document.querySelectorAll('.checkBoxDiv');
                    divs.forEach(function(d) {
                        var span = d.querySelector('span');
                        if (!span) return;
                        var name = span.textContent.trim();
                        if (name === '配送费') {
                            if (d.classList.contains('on')) {
                                d.click();
                            }
                        }
                        if (name === '冷冻面团' || name === '蛋糕及面包成品及饼干类') {
                            if (!d.classList.contains('on')) {
                                d.click();
                            }
                        }
                    });
                })()
            """)

            check_result = page.evaluate("""
                (function() {
                    var divs = document.querySelectorAll('.checkBoxDiv');
                    var r = [];
                    divs.forEach(function(d) {
                        var s = d.querySelector('span');
                        if (s) {
                            var name = s.textContent.trim();
                            if (name === '配送费' || name === '冷冻面团' || name === '蛋糕及面包成品及饼干类') {
                                r.push(name + ':' + d.classList.contains('on'));
                            }
                        }
                    });
                    return r.join(' | ');
                })()
            """)
            logger.info(f"    勾选验证: {check_result}")

            logger.info("  → 点击「确定」关闭弹框...")
            click_by_text(page, "确定", "确定弹框")
            time.sleep(0.5)

            logger.info("  [查询] 执行查询...")
            click_by_text(page, "查询", "查询")
            page.wait_for_load_state("networkidle", timeout=150_000)
            time.sleep(3)

            logger.info("  [导出] 导出自产品文件...")
            with page.expect_download(timeout=180_000) as dl_info_self:
                result = page.evaluate("""
                    (function() {
                        var els = document.querySelectorAll('*');
                        for (var i = 0; i < els.length; i++) {
                            if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                                els[i].click();
                                return 'ok';
                            }
                        }
                        return 'not found';
                    })()
                """)
                logger.info(f"  点击导出: {result}")
                if result == "not found":
                    raise RuntimeError("未找到「导出」按钮")

            download_self = dl_info_self.value
            logger.info(f"  下载文件名: {download_self.suggested_filename}")

            dest_self = output_dir / f"仓库配送大客户对比表_自产品_{date_range_str}.xlsx"
            download_self.save_as(dest_self)
            logger.info(f"  已保存到: {dest_self}")

            # ── 步骤 2：仓库配送商品大客户对比表 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：下载仓库配送商品大客户对比表")
            logger.info(f"{'─' * 55}")

            logger.info("  [导航] 前往仓库配送商品大客户对比表...")
            page.goto(DELIVERY_PRODUCT_COMPARISON_URL)
            page.wait_for_load_state("networkidle", timeout=120_000)
            logger.info(f"  已到达 → {page.url}")

            logger.info(f"  → 设置完成时间: {monday_str} ~ {sunday_str}...")
            set_date(page, "开始日期", f"{monday_str} 00:00")
            set_date(page, "结束日期", f"{sunday_str} 23:59")

            logger.info("  [查询] 执行查询...")
            click_by_text(page, "查询", "查询")
            page.wait_for_load_state("networkidle", timeout=150_000)
            time.sleep(3)

            logger.info("  [导出] 导出文件...")
            with page.expect_download(timeout=180_000) as dl_info2:
                result = page.evaluate("""
                    (function() {
                        var els = document.querySelectorAll('*');
                        for (var i = 0; i < els.length; i++) {
                            if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                                els[i].click();
                                return 'ok';
                            }
                        }
                        return 'not found';
                    })()
                """)
                logger.info(f"  点击导出: {result}")
                if result == "not found":
                    raise RuntimeError("未找到「导出」按钮")

            download2 = dl_info2.value
            logger.info(f"  下载文件名: {download2.suggested_filename}")

            dest2 = output_dir / f"仓库配送商品大客户对比表_{date_range_str}.xlsx"
            download2.save_as(dest2)
            logger.info(f"  已保存到: {dest2}")

            logger.info(f"{'=' * 55}")
            logger.info(f"  工厂配送兔司家门店周度销售报表下载完成！")
            logger.info(f"  仓库配送大客户对比表 → {dest}")
            logger.info(f"  仓库配送大客户对比表_配送费 → {dest_fee}")
            logger.info(f"  仓库配送大客户对比表_自产品 → {dest_self}")
            logger.info(f"  仓库配送商品大客户对比表 → {dest2}")
            logger.info(f"{'=' * 55}\n")

            # ── 步骤 3：格式化仓库配送商品大客户对比表 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：格式化数据并生成报表")
            logger.info(f"{'─' * 55}")

            logger.info("  读取仓库配送商品大客户对比表（聚合）...")
            product_rows = read_product_comparison(dest2)
            logger.info(f"  共 {len(product_rows)} 个大客户（商品聚合）")

            logger.info("  读取仓库配送大客户对比表（订单数）...")
            delivery_rows = read_delivery_comparison(dest)
            logger.info(f"  共 {len(delivery_rows)} 个大客户")

            logger.info("  读取配送费数据...")
            fee_data = read_delivery_fee(dest_fee)
            logger.info(f"  共 {len(fee_data)} 条配送费数据")

            logger.info("  读取自产品数据...")
            self_product_data = read_self_product(dest_self)
            logger.info(f"  共 {len(self_product_data)} 条自产品数据")

            logger.info("  读取商品大客户对比表明细...")
            product_detail_rows = read_product_comparison_detail(dest2)
            logger.info(f"  共 {len(product_detail_rows)} 条明细行")

            formatted_output = OUTPUT_DIR / date_range_str / f"工厂配送兔司家门店周度销售报表_{date_range_str}.xlsx"
            merge_delivery_into_template(product_rows, delivery_rows, TEMPLATE_FILE,
                                         formatted_output, monday, sunday, fee_data,
                                         self_product_data, product_detail_rows)
            logger.info(f"  格式化输出 → {formatted_output}")

            logger.info(f"{'=' * 55}")
            logger.info(f"  工厂配送兔司家门店周度销售报表全部完成！")
            logger.info(f"{'=' * 55}\n")

        except Exception as e:
            logger.error(f"任务失败: {e}")
            try:
                screenshot = OUTPUT_DIR / "error_screenshot.png"
                page.screenshot(path=str(screenshot))
                logger.info(f"  错误截图已保存: {screenshot}")
            except Exception:
                pass
            raise
        finally:
            context.close()
            browser.close()


if __name__ == "__main__":
    main()
