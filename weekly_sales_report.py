"""
周度销售报表 - 自动化脚本
=====================================
依赖：pip install playwright && playwright install chromium

自动登录 Pospal 后台，依次完成周度销售报表所需的数据导出与处理。

用法：
    python weekly_sales_report.py                    # 导出本周数据（周一到周日）
    python weekly_sales_report.py --weeks -1         # 导出上周数据
    python weekly_sales_report.py --date 2026.06.03  # 根据日期推断所在周
    python weekly_sales_report.py --headless          # 无头模式（不显示浏览器）
"""

import argparse
import copy
import logging
import logging.handlers
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ─── 日志配置 ───────────────────────────────────────────────────────────────────
LOG_DIR = Path(__file__).resolve().parent / "log"
LOG_DIR.mkdir(exist_ok=True)

_file_handler = logging.handlers.TimedRotatingFileHandler(
    LOG_DIR / "weekly_sales_report.log",
    when="midnight",
    backupCount=30,
    encoding="utf-8",
)
_file_handler.suffix = "%Y-%m-%d"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[_file_handler, logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# ─── 配置区（按需修改）─────────────────────────────────────────────────────────
ACCOUNT = "huomimayzb"
WORKER_ID = "M006"
PASSWORD = "tusijia88"

LOGIN_URL = "https://beta69.pospal.cn/"
DELIVERY_COMPARISON_URL = "https://css69.pospal.cn/EnterpriseReport/DeliveryComparison"
DELIVERY_PRODUCT_COMPARISON_URL = "https://css69.pospal.cn/EnterpriseReport/DeliveryProductComparison"

OUTPUT_DIR = Path(__file__).resolve().parent / "周度销售报表"
TEMPLATE_FILE = Path(__file__).resolve().parent / "周度_工厂配送兔司家门店货品对比表及销售排行表_格式化模板.xlsx"

HEADER_ROW = 2
TOTALS_ROW = 3
DATA_START_ROW = 4
SAMPLE_ROW = 4

SRC_COL_MAP = {
    "大客户名称": "B",
    "实际出库成本": "C",
    "销售出库单数": "I",
    "销售退货单数": "J",
    "出库量": "K",
    "退货量": "L",
    "实际出库量": "M",
    "出库金额": "N",
    "退货金额": "O",
}


# ─── 格式化合并函数 ──────────────────────────────────────────────────────────


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def read_delivery_comparison(data_file: Path):
    wb = load_workbook(data_file, data_only=True)
    ws = wb.active

    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val:
            headers[str(val).strip()] = c

    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=headers["大客户名称"]).value
        if name is None or str(name).strip() == "-":
            break
        row_data = {}
        for h, c in headers.items():
            row_data[h] = ws.cell(row=r, column=c).value
        rows.append(row_data)

    wb.close()
    return rows


def merge_delivery_into_template(data_rows, template_file: Path, output_file: Path,
                                  monday: datetime, sunday: datetime):
    wb = load_workbook(template_file)
    ws = wb.active

    max_col = 15

    sample_cells = list(ws.iter_rows(min_row=SAMPLE_ROW, max_row=SAMPLE_ROW))[0]
    sample_height = ws.row_dimensions[SAMPLE_ROW].height
    totals_cells = list(ws.iter_rows(min_row=TOTALS_ROW, max_row=TOTALS_ROW))[0]

    data_rows_sorted = sorted(
        data_rows,
        key=lambda r: float(r.get("实际出库成本", 0) or 0),
        reverse=True,
    )
    data_count = len(data_rows_sorted)

    for merge in list(ws.merged_cells.ranges):
        if merge.min_col in (7, 8):
            ws.unmerge_cells(str(merge))
    ws.cell(row=TOTALS_ROW, column=7).value = None
    ws.cell(row=TOTALS_ROW, column=8).value = None

    ws.delete_rows(SAMPLE_ROW, 2)
    ws.insert_rows(DATA_START_ROW, data_count)

    for i, row_data in enumerate(data_rows_sorted):
        r = DATA_START_ROW + i
        ws.cell(row=r, column=1).value = f"=ROW()-3"
        ws.cell(row=r, column=2).value = row_data.get("大客户名称")

        cost = row_data.get("实际出库成本")
        ws.cell(row=r, column=3).value = float(cost) if cost else None

        c_letter = "C"
        f_letter = "F"
        ws.cell(row=r, column=4).value = f"={c_letter}{r}-{f_letter}{r}"

        ws.cell(row=r, column=9).value = _to_num(row_data.get("销售出库单数"))
        ws.cell(row=r, column=10).value = _to_num(row_data.get("销售退货单数"))
        ws.cell(row=r, column=11).value = _to_num(row_data.get("出库量"))
        ws.cell(row=r, column=12).value = _to_num(row_data.get("退货量"))
        ws.cell(row=r, column=13).value = _to_num(row_data.get("实际出库量"))
        ws.cell(row=r, column=14).value = _to_num(row_data.get("出库金额"))
        ws.cell(row=r, column=15).value = _to_num(row_data.get("退货金额"))

        for col_idx in range(1, max_col + 1):
            src_idx = min(col_idx - 1, len(sample_cells) - 1)
            copy_cell_style(sample_cells[src_idx], ws.cell(row=r, column=col_idx))
        if sample_height:
            ws.row_dimensions[r].height = sample_height

    tr = TOTALS_ROW
    last_data_row = DATA_START_ROW + data_count - 1

    ws.cell(row=tr, column=1).value = "合计"
    for col in [3, 4, 5, 6, 9, 10, 11, 12, 13, 14, 15]:
        letter = get_column_letter(col)
        ws.cell(row=tr, column=col).value = f"=SUM({letter}{DATA_START_ROW}:{letter}{last_data_row})"

    for col_idx in range(1, max_col + 1):
        src_idx = min(col_idx - 1, len(totals_cells) - 1)
        copy_cell_style(totals_cells[src_idx], ws.cell(row=tr, column=col_idx))

    cell_a1 = ws.cell(row=1, column=1)
    if cell_a1.value:
        old_text = str(cell_a1.value)
        m_start = monday.month
        d_start = monday.day
        m_end = sunday.month
        d_end = sunday.day
        y = monday.year

        new_text = re.sub(
            r'\d{4}年\d{1,2}月\d{1,2}日至\d{1,2}月\d{1,2}日',
            f"{y}年{m_start}月{d_start}日至{m_end}月{d_end}日",
            old_text,
        )
        cell_a1.value = new_text

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    logger.info(f"  已生成格式化文件: {output_file}")


def _to_num(val):
    if val is None:
        return None
    try:
        f = float(val)
        if f == 0:
            return None
        if f == int(f):
            return int(f)
        return f
    except (ValueError, TypeError):
        return val


# ─── 工具函数 ──────────────────────────────────────────────────────────────────


def get_week_range(weeks: int = 0, date_str: str = None):
    if date_str:
        base = datetime.strptime(date_str, "%Y.%m.%d")
    else:
        base = datetime.now()

    monday = base - timedelta(days=base.weekday())
    monday = monday + timedelta(weeks=weeks)
    sunday = monday + timedelta(days=6)

    return monday, sunday


# ─── 浏览器自动化函数 ────────────────────────────────────────────────────────


def set_date(page, placeholder: str, value: str):
    page.evaluate(f"""
        (function() {{
            var inp = document.querySelector('input.timeInput.hasDatepicker[placeholder="{placeholder}"]');
            if (!inp) return 'not found';
            inp.value = "{value}";
            inp.dispatchEvent(new Event('input',  {{bubbles: true}}));
            inp.dispatchEvent(new Event('change', {{bubbles: true}}));
            inp.dispatchEvent(new Event('blur',   {{bubbles: true}}));
            return inp.value;
        }})()
    """)


def click_by_text(page, text: str, desc: str = ""):
    result = page.evaluate(f"""
        (function() {{
            var els = document.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {{
                if (els[i].textContent.trim() === "{text}" && els[i].children.length === 0) {{
                    els[i].click();
                    return 'clicked tag=' + els[i].tagName + ' class=' + els[i].className;
                }}
            }}
            return 'not found';
        }})()
    """)
    tag = f"[{desc}] " if desc else ""
    logger.info(f"  {tag}→ {result}")
    return "clicked" in result


def login(page):
    logger.info("[1/4] 打开登录页面...")
    page.goto(LOGIN_URL)
    page.wait_for_load_state("networkidle")

    logger.info("[2/4] 切换到工号登录模式...")
    page.evaluate("""
        (function() {
            var els = document.querySelectorAll('div,span,a,button');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '工号登录' && els[i].children.length === 0) {
                    els[i].click();
                    return 'clicked';
                }
            }
            return 'not found';
        })()
    """)
    time.sleep(1.5)

    placeholder = page.evaluate("""
        (function() {
            var inputs = document.querySelectorAll('input');
            var r = [];
            for (var i = 0; i < inputs.length; i++) {
                if (inputs[i].placeholder) r.push(inputs[i].placeholder);
            }
            return r.join(' | ');
        })()
    """)
    if "员工工号" not in placeholder:
        raise RuntimeError(f"工号登录切换失败，当前输入框：{placeholder}")
    logger.info(f"  表单已切换 → {placeholder}")

    logger.info("[3/4] 填入账号/工号/密码...")
    page.evaluate(f"""
        (function() {{
            var inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].placeholder === '请输入账号')     inputs[i].value = "{ACCOUNT}";
                if (inputs[i].placeholder === '请输入员工工号') inputs[i].value = "{WORKER_ID}";
                if (inputs[i].placeholder === '请输入工号密码') inputs[i].value = "{PASSWORD}";
            }}
            inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].value) {{
                    inputs[i].dispatchEvent(new Event('input', {{bubbles: true}}));
                }}
            }}
        }})()
    """)

    time.sleep(1.5)

    logger.info("[4/4] 点击登录按钮...")
    click_by_text(page, "登 录", "登录")
    page.wait_for_load_state("networkidle", timeout=30_000)

    time.sleep(2)


def main():
    parser = argparse.ArgumentParser(description="Pospal 周度销售报表自动化脚本")
    parser.add_argument("--weeks", type=int, default=0, help="周偏移量：0=本周，-1=上周（默认0）")
    parser.add_argument("--date", type=str, help="指定日期推断所在周，格式 YYYY.MM.DD")
    parser.add_argument("--headless", action="store_true", help="无头模式（不显示浏览器窗口）")
    args = parser.parse_args()

    if args.date:
        monday, sunday = get_week_range(date_str=args.date)
    else:
        monday, sunday = get_week_range(weeks=args.weeks)

    monday_str = monday.strftime("%Y.%m.%d")
    sunday_str = sunday.strftime("%Y.%m.%d")
    date_range_str = f"{monday.strftime('%Y-%m-%d')}~{sunday.strftime('%Y-%m-%d')}"

    logger.info(f"{'=' * 55}")
    logger.info(f"  Pospal 周度销售报表")
    logger.info(f"  目标周：{monday_str} ~ {sunday_str}")
    logger.info(f"  输出根目录：{OUTPUT_DIR}")
    logger.info(f"{'=' * 55}")

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        logger.error("未安装 playwright，请先运行: pip install playwright && playwright install chromium")
        sys.exit(1)

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=args.headless,
            args=["--start-maximized"],
        )
        context = browser.new_context(
            viewport={"width": 1600, "height": 900},
            accept_downloads=True,
        )
        page = context.new_page()

        try:
            login(page)

            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：下载仓库配送大客户对比表")
            logger.info(f"{'─' * 55}")

            logger.info("  [导航] 前往仓库配送大客户对比表...")
            page.goto(DELIVERY_COMPARISON_URL)
            page.wait_for_load_state("networkidle", timeout=30_000)
            logger.info(f"  已到达 → {page.url}")

            logger.info(f"  → 设置完成时间: {monday_str} ~ {sunday_str}...")
            set_date(page, "开始日期", f"{monday_str} 00:00")
            set_date(page, "结束日期", f"{sunday_str} 23:59")

            logger.info("  [查询] 执行查询...")
            click_by_text(page, "查询", "查询")
            page.wait_for_load_state("networkidle", timeout=90_000)
            time.sleep(3)

            logger.info("  [导出] 导出文件...")
            output_dir = OUTPUT_DIR / date_range_str / "原始下载"
            output_dir.mkdir(parents=True, exist_ok=True)

            with page.expect_download(timeout=60_000) as dl_info:
                result = page.evaluate("""
                    (function() {
                        var els = document.querySelectorAll('*');
                        for (var i = 0; i < els.length; i++) {
                            if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                                els[i].click();
                                return 'ok';
                            }
                        }
                        return 'not found';
                    })()
                """)
                logger.info(f"  点击导出: {result}")
                if result == "not found":
                    raise RuntimeError("未找到「导出」按钮")

            download = dl_info.value
            logger.info(f"  下载文件名: {download.suggested_filename}")

            dest = output_dir / f"仓库配送大客户对比表_{date_range_str}.xlsx"
            download.save_as(dest)
            logger.info(f"  已保存到: {dest}")

            # ── 步骤 2：仓库配送商品大客户对比表 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：下载仓库配送商品大客户对比表")
            logger.info(f"{'─' * 55}")

            logger.info("  [导航] 前往仓库配送商品大客户对比表...")
            page.goto(DELIVERY_PRODUCT_COMPARISON_URL)
            page.wait_for_load_state("networkidle", timeout=30_000)
            logger.info(f"  已到达 → {page.url}")

            logger.info(f"  → 设置完成时间: {monday_str} ~ {sunday_str}...")
            set_date(page, "开始日期", f"{monday_str} 00:00")
            set_date(page, "结束日期", f"{sunday_str} 23:59")

            logger.info("  [查询] 执行查询...")
            click_by_text(page, "查询", "查询")
            page.wait_for_load_state("networkidle", timeout=90_000)
            time.sleep(3)

            logger.info("  [导出] 导出文件...")
            with page.expect_download(timeout=60_000) as dl_info2:
                result = page.evaluate("""
                    (function() {
                        var els = document.querySelectorAll('*');
                        for (var i = 0; i < els.length; i++) {
                            if (els[i].textContent.trim() === '导出' && els[i].children.length === 0) {
                                els[i].click();
                                return 'ok';
                            }
                        }
                        return 'not found';
                    })()
                """)
                logger.info(f"  点击导出: {result}")
                if result == "not found":
                    raise RuntimeError("未找到「导出」按钮")

            download2 = dl_info2.value
            logger.info(f"  下载文件名: {download2.suggested_filename}")

            dest2 = output_dir / f"仓库配送商品大客户对比表_{date_range_str}.xlsx"
            download2.save_as(dest2)
            logger.info(f"  已保存到: {dest2}")

            logger.info(f"{'=' * 55}")
            logger.info(f"  周度销售报表下载完成！")
            logger.info(f"  仓库配送大客户对比表 → {dest}")
            logger.info(f"  仓库配送商品大客户对比表 → {dest2}")
            logger.info(f"{'=' * 55}\n")

            # ── 步骤 3：格式化仓库配送大客户对比表 ──
            logger.info(f"{'─' * 55}")
            logger.info(f"  步骤：格式化仓库配送大客户对比表")
            logger.info(f"{'─' * 55}")

            logger.info("  读取仓库配送大客户对比表...")
            delivery_rows = read_delivery_comparison(dest)
            logger.info(f"  共 {len(delivery_rows)} 行大客户数据")

            formatted_output = OUTPUT_DIR / date_range_str / f"工厂配送兔司家门店货品对比表_{date_range_str}.xlsx"
            merge_delivery_into_template(delivery_rows, TEMPLATE_FILE, formatted_output, monday, sunday)
            logger.info(f"  格式化输出 → {formatted_output}")

            logger.info(f"{'=' * 55}")
            logger.info(f"  周度销售报表全部完成！")
            logger.info(f"{'=' * 55}\n")

        except Exception as e:
            logger.error(f"任务失败: {e}")
            try:
                screenshot = OUTPUT_DIR / "error_screenshot.png"
                page.screenshot(path=str(screenshot))
                logger.info(f"  错误截图已保存: {screenshot}")
            except Exception:
                pass
            raise
        finally:
            context.close()
            browser.close()


if __name__ == "__main__":
    main()
